"""Safety business workflows."""

import json
import logging
import os
import uuid
from datetime import datetime

from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.safety.ai_workflow_configs import SAFETY_WORKFLOW_CONFIGS
from app.modules.safety.models import (
    SpecialOperationReport,
)
from app.modules.safety.repository import SafetyRepository
from app.modules.safety.schemas import (
    SpecialOperationReportCreate,
    SpecialOperationReportUpdate,
)
from app.platform.integrations.ai.client import AIOutputError, AIService
from app.platform.integrations.ai.prompts import (
    STANDALONE_WORKFLOW_CONFIG,
    build_prompt,
)

logger = logging.getLogger(__name__)


class SpecialOperationReportService:
    """八大特殊作业报备业务服务"""

    def __init__(self, session: AsyncSession):
        self.session = session
        self.repo = SafetyRepository(session)

    # ── CRUD ──

    async def get_reports(
        self,
        skip: int = 0,
        limit: int = 20,
        status: str | None = None,
        operation_type: str | None = None,
        operation_level: str | None = None,
        risk_level: str | None = None,
        department: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        keyword: str | None = None,
        is_critical: bool | None = None,
    ) -> tuple[list[SpecialOperationReport], int]:
        """获取特殊作业报备列表"""
        return await self.repo.get_special_operation_reports(
            skip, limit, status, operation_type, operation_level,
            risk_level, department, date_from, date_to, keyword, is_critical,
        )

    async def get_report(self, report_id: uuid.UUID) -> SpecialOperationReport | None:
        """获取报备详情"""
        return await self.repo.get_special_operation_report_by_id(report_id)

    async def create_report(
        self, data: SpecialOperationReportCreate
    ) -> SpecialOperationReport:
        """创建报备"""
        return await self.repo.create_special_operation_report(data.model_dump())

    async def update_report(
        self, report_id: uuid.UUID, data: SpecialOperationReportUpdate
    ) -> SpecialOperationReport | None:
        """更新报备"""
        update_data = {k: v for k, v in data.model_dump().items() if v is not None}
        return await self.repo.update_special_operation_report(report_id, update_data)

    async def delete_report(self, report_id: uuid.UUID) -> bool:
        """删除报备"""
        return await self.repo.delete_special_operation_report(report_id)

    # ── 工作流 ──

    async def submit_report(
        self, report_id: uuid.UUID
    ) -> SpecialOperationReport | None:
        """提交报备（草稿→已提交），并自动判定是否为关键作业"""
        report = await self.repo.get_special_operation_report_by_id(report_id)
        if not report or report.status != "draft":
            return None

        # AI 自动判定关键作业
        is_critical, critical_reason = await self._identify_critical(report)

        return await self.repo.update_special_operation_report(
            report_id, {
                "status": "submitted",
                "is_critical": is_critical,
                "is_critical_reason": critical_reason,
            }
        )

    async def approve_report(
        self, report_id: uuid.UUID
    ) -> SpecialOperationReport | None:
        """审批报备（已提交→已审批）"""
        report = await self.repo.get_special_operation_report_by_id(report_id)
        if not report or report.status != "submitted":
            return None
        return await self.repo.update_special_operation_report(
            report_id,
            {"status": "approved", "approved_at": datetime.now()},
        )

    async def reject_report(
        self, report_id: uuid.UUID, reason: str
    ) -> SpecialOperationReport | None:
        """驳回报备（已提交→已驳回）"""
        report = await self.repo.get_special_operation_report_by_id(report_id)
        if not report or report.status != "submitted":
            return None
        return await self.repo.update_special_operation_report(
            report_id,
            {"status": "rejected", "rejection_reason": reason},
        )

    async def set_critical_manual(
        self, report_id: uuid.UUID, is_critical: bool, reason: str | None, updated_by: str | None
    ) -> SpecialOperationReport | None:
        """手动修改关键作业标记"""
        report = await self.repo.get_special_operation_report_by_id(report_id)
        if not report:
            return None
        return await self.repo.update_special_operation_report(
            report_id, {
                "is_critical": is_critical,
                "is_critical_reason": reason,
                "is_critical_updated_by": updated_by,
            }
        )

    # ── AI 判定 ──

    async def _get_workflow_config(self, module_code: str) -> dict:
        """获取独立工作流配置（DB 优先，fallback 到硬编码 STANDALONE_WORKFLOW_CONFIG）"""
        workflow = await self.repo.get_ai_workflow_config_by_module(module_code)
        if workflow and workflow.is_enabled and workflow.script_configs:
            scripts = workflow.script_configs
            if isinstance(scripts, list) and len(scripts) > 0:
                first = scripts[0]
                if first.get("is_enabled", True):
                    logger.debug(
                        "使用 DB 工作流配置: %s — %s", module_code, first.get("name")
                    )
                    return first
        # Fallback：安全模块配置优先，再到平台全局配置
        config = SAFETY_WORKFLOW_CONFIGS.get(module_code) or STANDALONE_WORKFLOW_CONFIG.get(module_code, {})
        if config:
            logger.debug("使用硬编码工作流配置: %s — %s", module_code, config.get("name"))
        return config

    async def _get_ai_service(self) -> "AIService":
        """获取文本模型 AIService（硬编码配置）"""
        from app.modules.safety.service.config import create_ai_service
        return create_ai_service("text")

    async def _identify_critical(
        self, report: "SpecialOperationReport"
    ) -> tuple[bool, str | None]:
        """判定报备是否为关键作业（AI 优先，失败时基于规则 fallback）"""
        try:
            ai = await self._get_ai_service()
            return await self._ai_identify_critical(ai, report)
        except Exception as e:
            logger.warning("AI 关键作业判定失败，使用规则 fallback: %s", e)
            return self._rule_based_identify_critical(report)

    async def _ai_identify_critical(
        self, ai: "AIService", report: "SpecialOperationReport"
    ) -> tuple[bool, str | None]:
        """使用 AI 判定关键作业（提示词由工作流配置提供）"""
        OP_TYPE_LABELS = {
            "hot_work": "动火作业", "confined_space": "受限空间",
            "height_work": "高处作业", "temporary_electricity": "临时用电",
            "blind_plate": "盲板抽堵", "excavation": "动土作业",
            "lifting": "起重吊装", "road_breaking": "断路作业",
        }
        op_label = OP_TYPE_LABELS.get(report.operation_type, report.operation_type)

        context = (
            f"作业类型：{op_label}\n"
            f"作业级别：{report.operation_level or '未指定'}\n"
            f"作业地点：{report.location or '未指定'}\n"
            f"作业部门：{report.department or '未指定'}\n"
            f"作业内容：{report.work_description or '未指定'}\n"
            f"风险等级：{report.risk_level or '未指定'}\n"
            f"安全措施：{report.safety_measures or '未指定'}\n"
            f"风险评估：{report.risk_assessment or '未指定'}\n"
            f"应急消防器材：{report.emergency_equipment or '未指定'}"
        )

        # 从工作流配置构建提示词（DB 优先，fallback 到硬编码）
        wf_config = await self._get_workflow_config("special-ops-critical")
        if wf_config:
            prompt = build_prompt(wf_config) + "\n\n## 本次判定输入\n" + context
        else:
            prompt = context  # 极端情况：无任何配置可用

        messages = [
            {"role": "system", "content": "你是一名化工安全专家，严格按照 GB 30871-2022 标准判定。只返回 JSON。"},
            {"role": "user", "content": prompt},
        ]

        response_text = await ai.chat(messages, response_format="json_object")
        result = json.loads(response_text)
        return result.get("is_critical", False), result.get("reason")

    def _rule_based_identify_critical(
        self, report: "SpecialOperationReport"
    ) -> tuple[bool, str | None]:
        """基于规则的 fallback 关键作业判定"""
        high_risk_types = {"hot_work", "confined_space", "height_work", "lifting"}
        critical_reasons: list[str] = []

        # 特级/一级作业
        if report.operation_level in ("special", "grade1"):
            critical_reasons.append(f"作业级别为{report.operation_level}")

        # 高风险等级
        if report.risk_level in ("level_1", "level_2"):
            critical_reasons.append(f"风险等级为{report.risk_level}")

        # 高风险作业类型 + 特级/一级
        if report.operation_type in high_risk_types and report.operation_level in ("special", "grade1"):
            if not critical_reasons:
                critical_reasons.append(
                    f"{report.operation_type} 作业类型属于高风险作业"
                )

        if critical_reasons:
            return True, "；".join(critical_reasons)
        return False, None

    # ── 台账查询 ──

    async def get_ledger(
        self,
        skip: int = 0,
        limit: int = 20,
        status_list: list[str] | None = None,
        operation_type: str | None = None,
        operation_level: str | None = None,
        risk_level: str | None = None,
        department: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        keyword: str | None = None,
        is_critical: bool | None = None,
    ) -> tuple[list[SpecialOperationReport], int]:
        """获取特殊作业台账列表"""
        return await self.repo.get_special_operation_ledger(
            skip=skip,
            limit=limit,
            status_list=status_list,
            operation_type=operation_type,
            operation_level=operation_level,
            risk_level=risk_level,
            department=department,
            date_from=date_from,
            date_to=date_to,
            keyword=keyword,
            is_critical=is_critical,
        )

    async def get_ledger_stats(
        self, status_list: list[str] | None = None
    ) -> list[dict]:
        """获取台账统计"""
        return await self.repo.get_special_operation_ledger_stats(status_list)

    # ── AI 导出 ──

    async def parse_natural_query(
        self, natural_query: str
    ) -> dict:
        """使用 AI 将自然语言筛选条件解析为结构化参数（提示词由工作流配置提供）"""
        OP_TYPE_LABELS = {
            "动火作业": "hot_work", "受限空间": "confined_space",
            "高处作业": "height_work", "临时用电": "temporary_electricity",
            "盲板抽堵": "blind_plate", "动土作业": "excavation",
            "起重吊装": "lifting", "断路作业": "road_breaking",
        }

        # 从工作流配置构建提示词（DB 优先，fallback 到硬编码）
        wf_config = await self._get_workflow_config("special-ops-export")
        if wf_config:
            prompt = build_prompt(wf_config) + "\n\n用户查询：" + natural_query
        else:
            prompt = natural_query

        try:
            ai = await self._get_ai_service()
            messages = [
                {"role": "system", "content": "你是一个数据库查询助手。只返回 JSON。"},
                {"role": "user", "content": prompt},
            ]
            response_text = await ai.chat(messages, response_format="json_object")
            import json
            result = json.loads(response_text)
            # 验证 operation_type 值
            if result.get("operation_type"):
                op_type = result["operation_type"]
                if op_type in OP_TYPE_LABELS:
                    result["operation_type"] = OP_TYPE_LABELS[op_type]
            # 清除 None 值
            return {k: v for k, v in result.items() if v is not None}
        except Exception as e:
            logger.warning("AI 自然语言解析失败: %s", e)
            return {"explanation": f"AI 解析失败，将使用原始查询: {natural_query}", "keyword": natural_query}

    async def export_ledger_excel(
        self,
        operation_type: str | None = None,
        operation_level: str | None = None,
        risk_level: str | None = None,
        department: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        keyword: str | None = None,
        is_critical: bool | None = None,
    ) -> bytes:
        """导出台账为 Excel 文件"""
        import io

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        items, _ = await self.get_ledger(
            skip=0, limit=10000,  # 导出全部（上限1万条）
            operation_type=operation_type, operation_level=operation_level,
            risk_level=risk_level, department=department,
            date_from=date_from, date_to=date_to,
            keyword=keyword, is_critical=is_critical,
        )

        OP_TYPE_LABELS = {
            "hot_work": "动火作业", "confined_space": "受限空间",
            "height_work": "高处作业", "temporary_electricity": "临时用电",
            "blind_plate": "盲板抽堵", "excavation": "动土作业",
            "lifting": "起重吊装", "road_breaking": "断路作业",
        }
        OP_LEVEL_LABELS = {"special": "特级", "grade1": "一级", "grade2": "二级", "not_applicable": "不涉及"}
        STATUS_LABELS = {"draft": "草稿", "submitted": "审批中", "approved": "已审批", "rejected": "已驳回"}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "特殊作业台账"

        # 标题样式
        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="5645D4", end_color="5645D4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="微软雅黑", size=10)
        cell_alignment = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        critical_fill = PatternFill(start_color="FDE0EC", end_color="FDE0EC", fill_type="solid")

        # 表头
        headers = [
            "序号", "报备编号", "作业类型", "作业级别", "作业地点",
            "作业内容", "作业部门", "计划开始", "计划结束",
            "报备人", "审批人", "审批时间",
            "状态", "是否关键作业", "关键作业判定理由", "备注",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        for row_idx, item in enumerate(items, 2):
            values = [
                row_idx - 1,
                item.report_no or "",
                OP_TYPE_LABELS.get(item.operation_type, item.operation_type or ""),
                OP_LEVEL_LABELS.get(item.operation_level, item.operation_level or ""),
                item.location or "",
                item.work_description or "",
                item.department or "",
                item.planned_start_time.strftime("%Y-%m-%d %H:%M") if item.planned_start_time else "",
                item.planned_end_time.strftime("%Y-%m-%d %H:%M") if item.planned_end_time else "",
                item.applicant_name or "",
                item.approver_name or "",
                item.approved_at.strftime("%Y-%m-%d %H:%M") if item.approved_at else "",
                STATUS_LABELS.get(item.status, item.status or ""),
                "是" if item.is_critical else "否",
                item.is_critical_reason or "",
                item.notes or "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                # 关键作业行高亮
                if item.is_critical:
                    cell.fill = critical_fill

        # 列宽
        col_widths = [6, 14, 10, 8, 14, 24, 12, 16, 16, 8, 8, 16, 8, 10, 28, 20]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


