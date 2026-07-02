import asyncio
import os
from urllib.parse import quote

from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.response import success_response
from app.modules.administration.schemas import VehicleCreate, VehicleResponse, VehicleUpdate
from app.modules.administration.service import VehicleService
from app.shared.schemas import PageParams

router = APIRouter()


def get_vehicle_service(session: AsyncSession = Depends(get_db)) -> VehicleService:
    return VehicleService(session)


@router.get("/vehicles", summary="车辆列表")
async def list_vehicles(
    keyword: str | None = None,
    status: str | None = None,
    page_params: PageParams = Depends(),
    service: VehicleService = Depends(get_vehicle_service),
):
    vehicles, total = await service.list_vehicles(
        keyword=keyword,
        status=status,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [VehicleResponse.model_validate(v).model_dump(mode="json") for v in vehicles]
    return {
        "code": 200,
        "message": "success",
        "data": data,
        "meta": {
            "page": page_params.page,
            "page_size": page_params.page_size,
            "total": total,
        },
    }


@router.post("/vehicles", summary="创建车辆")
async def create_vehicle(
    payload: VehicleCreate,
    service: VehicleService = Depends(get_vehicle_service),
):
    vehicle = await service.create_vehicle(payload)
    return success_response(
        data=VehicleResponse.model_validate(vehicle).model_dump(mode="json"),
        message="车辆创建成功",
        status_code=201,
    )


@router.get("/vehicles/{vehicle_id}", summary="车辆详情")
async def get_vehicle(
    vehicle_id: str,
    service: VehicleService = Depends(get_vehicle_service),
):
    vehicle = await service.get_vehicle(vehicle_id)
    return success_response(
        data=VehicleResponse.model_validate(vehicle).model_dump(mode="json"),
    )


@router.put("/vehicles/{vehicle_id}", summary="更新车辆")
async def update_vehicle(
    vehicle_id: str,
    payload: VehicleUpdate,
    service: VehicleService = Depends(get_vehicle_service),
):
    vehicle = await service.update_vehicle(vehicle_id, payload)
    return success_response(
        data=VehicleResponse.model_validate(vehicle).model_dump(mode="json"),
        message="车辆更新成功",
    )


@router.delete("/vehicles/{vehicle_id}", summary="删除车辆")
async def delete_vehicle(
    vehicle_id: str,
    service: VehicleService = Depends(get_vehicle_service),
):
    await service.delete_vehicle(vehicle_id)
    return success_response(message="车辆删除成功")


@router.get("/vehicles/template/download", summary="下载车辆导入模板")
async def download_vehicle_template():
    """提供车辆批量导入模板下载."""
    template_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "..", "..", "scripts", "车辆信息批量导入模板.xlsx"
    )

    if not os.path.exists(template_path):
        # 如果模板不存在，使用内存中的模板
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '车辆导入模板'

        # 样式定义
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # 标题行
        ws.merge_cells('A1:I1')
        ws['A1'] = '车辆信息批量导入模板'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # 说明行
        ws.merge_cells('A2:I2')
        ws['A2'] = '说明：请按以下格式填写数据，带 * 的为必填项。图片列支持填写图片URL或留空（导入后可在页面单独上传）。'
        ws['A2'].font = Font(size=10, color='666666')
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.row_dimensions[2].height = 35

        # 列标题
        headers = ['*车牌号', '品牌', '型号', '颜色', '行驶里程', '状态', '所属部门', '图片', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        ws.row_dimensions[3].height = 25

        # 示例数据
        sample_data = [
            ['粤B12345', '丰田', '凯美瑞', '黑色', 50000, '可用', '生产技术部', '', ''],
            ['粤B67890', '本田', '雅阁', '白色', 35000, '可用', '质量管理部', '', ''],
            ['粤B11111', '大众', '帕萨特', '银色', 80000, '维修中', '设备工程部', '', '待大修'],
        ]

        for row_idx, row_data in enumerate(sample_data, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # 列宽
        col_widths = [14, 12, 14, 10, 12, 10, 16, 20, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        ws.freeze_panes = 'A4'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f"attachment; filename*=UTF-8''{quote('车辆信息批量导入模板.xlsx', safe='')}"
            }
        )

    return FileResponse(
        template_path,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename='车辆信息批量导入模板.xlsx'
    )
