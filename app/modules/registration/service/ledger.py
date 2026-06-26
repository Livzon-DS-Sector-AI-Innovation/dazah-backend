"""Registration ledger service."""

import io
from datetime import date, datetime
from typing import List, Type

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.registration.models.copp_certificate import CoppCertificate
from app.modules.registration.models.domestic_approval import DomesticApproval
from app.modules.registration.models.international_review import InternationalReview
from app.modules.registration.models.overseas_approval import OverseasApproval
from app.modules.registration.models.wc_certificate import WcCertificate
from app.modules.registration.schemas.ledger import (
    CoppCertificateCreate,
    DomesticApprovalCreate,
    InternationalReviewCreate,
    LedgerSummary,
    OverseasApprovalCreate,
    WcCertificateCreate,
)



# ── Excel Parsing Helpers ──────────────────────────────────────────

def _parse_date(value) -> date | None:
    """Parse date from Excel cell (datetime, date, string, or None)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # 处理非标准日期格式（如"长期有效"）
        try:
            return date.fromisoformat(value)
        except (ValueError, TypeError):
            # 无法解析的日期格式，返回 None
            return None
    return None


def _parse_str(value) -> str | None:
    """Parse string from Excel cell, converting numbers to str."""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    return str(value)


def _parse_int(value) -> int | None:
    """Parse int from Excel cell (int, float, string, or None)."""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        return int(float(value))
    return None



def _find_header_row(ws, expected_headers: list[str], max_scan: int = 10) -> tuple[int, dict[str, int]]:
    """
    自动寻找表头行，返回 (header_row_index, {field_name: column_index})
    expected_headers: 期望的表头名称列表
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        row_values = [str(cell).strip() if cell else "" for cell in row]
        
        # 检查这一行是否包含预期的表头
        matches = sum(1 for h in expected_headers if h in row_values)
        if matches >= len(expected_headers) * 0.5:  # 至少匹配 50%
            # 构建字段映射
            field_map = {}
            for col_idx, val in enumerate(row_values):
                if val:
                    field_map[val] = col_idx
            return row_idx, field_map
    
    # 如果没找到，返回默认值
    return 1, {}


def _get_cell_value(row: tuple, field_map: dict[str, int], field_name: str, default=None):
    """从行数据中根据字段名获取值"""
    if field_name not in field_map:
        return default
    col_idx = field_map[field_name]
    if col_idx >= len(row):
        return default
    return row[col_idx]


def _fill_merged_cells(ws):
    """处理合并单元格，向下填充空值"""
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
        # 获取合并区域的第一个单元格的值
        top_left_value = ws.cell(row=min_row, column=min_col).value
        # 填充整个合并区域
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


# ── CRUD Operations ───────────────────────────────────────────────

async def list_domestic_approvals(db: AsyncSession, skip: int = 0, limit: int = 100) -> List[DomesticApproval]:
    stmt = select(DomesticApproval).where(DomesticApproval.is_deleted == False).offset(skip).limit(limit)
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def create_domestic_approval(db: AsyncSession, data: DomesticApprovalCreate) -> DomesticApproval:
    obj = DomesticApproval(**data.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


async def list_overseas_approvals(db: AsyncSession, skip: int = 0, limit: int = 100) -> List[OverseasApproval]:
    stmt = select(OverseasApproval).where(OverseasApproval.is_deleted == False).offset(skip).limit(limit)
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def create_overseas_approval(db: AsyncSession, data: OverseasApprovalCreate) -> OverseasApproval:
    obj = OverseasApproval(**data.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


async def list_international_reviews(db: AsyncSession, skip: int = 0, limit: int = 100) -> List[InternationalReview]:
    stmt = select(InternationalReview).where(InternationalReview.is_deleted == False).offset(skip).limit(limit)
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def create_international_review(db: AsyncSession, data: InternationalReviewCreate) -> InternationalReview:
    obj = InternationalReview(**data.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


async def list_copp_certificates(db: AsyncSession, skip: int = 0, limit: int = 100) -> List[CoppCertificate]:
    stmt = select(CoppCertificate).where(CoppCertificate.is_deleted == False).offset(skip).limit(limit)
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def create_copp_certificate(db: AsyncSession, data: CoppCertificateCreate) -> CoppCertificate:
    obj = CoppCertificate(**data.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


async def list_wc_certificates(db: AsyncSession, skip: int = 0, limit: int = 100) -> List[WcCertificate]:
    stmt = select(WcCertificate).where(WcCertificate.is_deleted == False).offset(skip).limit(limit)
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def create_wc_certificate(db: AsyncSession, data: WcCertificateCreate) -> WcCertificate:
    obj = WcCertificate(**data.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


# ── Dashboard Summary ──────────────────────────────────────────────

async def get_ledger_summary(db: AsyncSession) -> LedgerSummary:
    # 国内已获批
    stmt = select(func.count()).select_from(DomesticApproval).where(DomesticApproval.is_deleted == False)
    domestic_count = (await db.execute(stmt)).scalar() or 0

    # 国外已获批
    stmt = select(func.count()).select_from(OverseasApproval).where(OverseasApproval.is_deleted == False)
    overseas_count = (await db.execute(stmt)).scalar() or 0

    # 国外获批国家数
    stmt = select(func.count(func.distinct(OverseasApproval.issuing_authority))).select_from(OverseasApproval).where(
        OverseasApproval.is_deleted == False,
        OverseasApproval.issuing_authority.isnot(None),
    )
    overseas_countries = (await db.execute(stmt)).scalar() or 0

    # 国际关联审评
    stmt = select(func.count()).select_from(InternationalReview).where(InternationalReview.is_deleted == False)
    international_review_count = (await db.execute(stmt)).scalar() or 0

    # COPP证书
    stmt = select(func.count()).select_from(CoppCertificate).where(CoppCertificate.is_deleted == False)
    copp_count = (await db.execute(stmt)).scalar() or 0

    # WC证书
    stmt = select(func.count()).select_from(WcCertificate).where(WcCertificate.is_deleted == False)
    wc_count = (await db.execute(stmt)).scalar() or 0

    # 审评中（从drugs表统计）
    from app.modules.registration.models.drug import Drug
    stmt = select(func.count()).select_from(Drug).where(Drug.is_deleted == False)
    reviewing_count = (await db.execute(stmt)).scalar() or 0

    # 计划申报（暂时为0）
    planned_count = 0

    return LedgerSummary(
        domestic_count=domestic_count,
        overseas_count=overseas_count,
        overseas_countries=overseas_countries,
        international_review_count=international_review_count,
        copp_count=copp_count,
        wc_count=wc_count,
        reviewing_count=reviewing_count,
        planned_count=planned_count,
    )


# ── Excel Import/Export ────────────────────────────────────────────

def export_domestic_approvals_to_excel(data: List[DomesticApproval]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "国内已获批"
    
    headers = ["品名", "证书名称", "批件号", "国家/发证机关", "发证日期", "证书有效期至", 
               "产品范围", "质量标准", "登记号", "证书是否过期", "生产车间", "产品有效期", "贮存条件"]
    ws.append(headers)
    
    for item in data:
        ws.append([
            item.product_name,
            item.certificate_name or "",
            item.batch_no or "",
            item.issuing_authority or "",
            item.issue_date.isoformat() if item.issue_date else "",
            item.valid_until.isoformat() if item.valid_until else "",
            item.product_scope or "",
            item.quality_standard or "",
            item.registration_no or "",
            item.is_expired or "",
            item.production_workshop or "",
            item.product_validity or "",
            item.storage_condition or "",
        ])
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def import_domestic_approvals_from_excel(file_content: bytes) -> dict:
    """
    智能解析 Excel，返回详细结果
    """
    wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
    ws = wb.active
    
    # 处理合并单元格
    _fill_merged_cells(ws)
    
    # 期望的表头
    expected_headers = ["品名", "证书名称", "批件号", "国家/发证机关", "发证日期", "证书有效期至",
                        "产品范围", "质量标准", "登记号", "证书是否过期", "生产车间", "产品有效期", "贮存条件"]
    
    # 自动寻找表头行
    header_row, field_map = _find_header_row(ws, expected_headers)
    
    result = {
        "total_rows": 0,
        "success_count": 0,
        "skipped_count": 0,
        "error_count": 0,
        "errors": [],
        "items": []
    }
    
    if not field_map:
        result["errors"].append("未找到有效表头，请确保 Excel 包含品名等字段")
        return result
    
    # 从表头下一行开始读取数据
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        result["total_rows"] += 1
        
        # 跳过全空行
        if not any(row):
            result["skipped_count"] += 1
            continue
        
        # 获取必填字段
        product_name = _parse_str(_get_cell_value(row, field_map, "品名"))
        if not product_name:
            result["skipped_count"] += 1
            continue
        
        try:
            item = DomesticApprovalCreate(
                product_name=product_name,
                certificate_name=_parse_str(_get_cell_value(row, field_map, "证书名称")),
                batch_no=_parse_str(_get_cell_value(row, field_map, "批件号")),
                issuing_authority=_parse_str(_get_cell_value(row, field_map, "国家/发证机关")),
                issue_date=_parse_date(_get_cell_value(row, field_map, "发证日期")),
                valid_until=_parse_date(_get_cell_value(row, field_map, "证书有效期至")),
                product_scope=_parse_str(_get_cell_value(row, field_map, "产品范围")),
                quality_standard=_parse_str(_get_cell_value(row, field_map, "质量标准")),
                registration_no=_parse_str(_get_cell_value(row, field_map, "登记号")),
                is_expired=_parse_str(_get_cell_value(row, field_map, "证书是否过期")),
                production_workshop=_parse_str(_get_cell_value(row, field_map, "生产车间")),
                product_validity=_parse_str(_get_cell_value(row, field_map, "产品有效期")),
                storage_condition=_parse_str(_get_cell_value(row, field_map, "贮存条件")),
            )
            result["items"].append(item)
            result["success_count"] += 1
        except Exception as e:
            result["error_count"] += 1
            if len(result["errors"]) < 10:  # 最多记录 10 条错误
                result["errors"].append(f"第{row_idx}行: {str(e)}")
    
    return result


def export_overseas_approvals_to_excel(data: List[OverseasApproval]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "国外已获批"
    
    headers = ["品名", "证书名称", "批件号", "国家/发证机关", "发证日期", "证书有效期至",
               "产品范围", "质量标准", "证书是否过期", "生产车间", "产品有效期", "贮存条件"]
    ws.append(headers)
    
    for item in data:
        ws.append([
            item.product_name,
            item.certificate_name or "",
            item.batch_no or "",
            item.issuing_authority or "",
            item.issue_date.isoformat() if item.issue_date else "",
            item.valid_until.isoformat() if item.valid_until else "",
            item.product_scope or "",
            item.quality_standard or "",
            item.is_expired or "",
            item.production_workshop or "",
            item.product_validity or "",
            item.storage_condition or "",
        ])
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def import_overseas_approvals_from_excel(file_content: bytes) -> dict:
    wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
    ws = wb.active
    _fill_merged_cells(ws)
    
    expected_headers = ["品名", "证书名称", "批件号", "国家/发证机关", "发证日期", "证书有效期至",
                        "产品范围", "质量标准", "证书是否过期", "生产车间", "产品有效期", "贮存条件"]
    header_row, field_map = _find_header_row(ws, expected_headers)
    
    result = {"total_rows": 0, "success_count": 0, "skipped_count": 0, "error_count": 0, "errors": [], "items": []}
    
    if not field_map:
        result["errors"].append("未找到有效表头")
        return result
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        result["total_rows"] += 1
        if not any(row):
            result["skipped_count"] += 1
            continue
        
        product_name = _parse_str(_get_cell_value(row, field_map, "品名"))
        if not product_name:
            result["skipped_count"] += 1
            continue
        
        try:
            item = OverseasApprovalCreate(
                product_name=product_name,
                certificate_name=_parse_str(_get_cell_value(row, field_map, "证书名称")),
                batch_no=_parse_str(_get_cell_value(row, field_map, "批件号")),
                issuing_authority=_parse_str(_get_cell_value(row, field_map, "国家/发证机关")),
                issue_date=_parse_date(_get_cell_value(row, field_map, "发证日期")),
                valid_until=_parse_date(_get_cell_value(row, field_map, "证书有效期至")),
                product_scope=_parse_str(_get_cell_value(row, field_map, "产品范围")),
                quality_standard=_parse_str(_get_cell_value(row, field_map, "质量标准")),
                is_expired=_parse_str(_get_cell_value(row, field_map, "证书是否过期")),
                production_workshop=_parse_str(_get_cell_value(row, field_map, "生产车间")),
                product_validity=_parse_str(_get_cell_value(row, field_map, "产品有效期")),
                storage_condition=_parse_str(_get_cell_value(row, field_map, "贮存条件")),
            )
            result["items"].append(item)
            result["success_count"] += 1
        except Exception as e:
            result["error_count"] += 1
            if len(result["errors"]) < 10:
                result["errors"].append(f"第{row_idx}行: {str(e)}")
    
    return result


def export_international_reviews_to_excel(data: List[InternationalReview]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "国际关联审评"
    
    headers = ["品名", "获批国家", "获批国家数量", "获批客户", "获批客户数量",
               "审评中-国家", "审评中-国家数量", "审评中-客户", "审评中-客户数量"]
    ws.append(headers)
    
    for item in data:
        ws.append([
            item.product_name,
            item.approved_countries or "",
            item.approved_country_count or 0,
            item.approved_clients or "",
            item.approved_client_count or 0,
            item.reviewing_countries or "",
            item.reviewing_country_count or 0,
            item.reviewing_clients or "",
            item.reviewing_client_count or 0,
        ])
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def import_international_reviews_from_excel(file_content: bytes) -> dict:
    wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
    ws = wb.active
    _fill_merged_cells(ws)
    
    expected_headers = ["品名", "获批国家", "获批国家数量", "获批客户", "获批客户数量",
                        "审评中-国家", "审评中-国家数量", "审评中-客户", "审评中-客户数量"]
    header_row, field_map = _find_header_row(ws, expected_headers)
    
    result = {"total_rows": 0, "success_count": 0, "skipped_count": 0, "error_count": 0, "errors": [], "items": []}
    
    if not field_map:
        result["errors"].append("未找到有效表头")
        return result
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        result["total_rows"] += 1
        if not any(row):
            result["skipped_count"] += 1
            continue
        
        product_name = _parse_str(_get_cell_value(row, field_map, "品名"))
        if not product_name:
            result["skipped_count"] += 1
            continue
        
        try:
            item = InternationalReviewCreate(
                product_name=product_name,
                approved_countries=_parse_str(_get_cell_value(row, field_map, "获批国家")),
                approved_country_count=_parse_int(_get_cell_value(row, field_map, "获批国家数量")),
                approved_clients=_parse_str(_get_cell_value(row, field_map, "获批客户")),
                approved_client_count=_parse_int(_get_cell_value(row, field_map, "获批客户数量")),
                reviewing_countries=_parse_str(_get_cell_value(row, field_map, "审评中-国家")),
                reviewing_country_count=_parse_int(_get_cell_value(row, field_map, "审评中-国家数量")),
                reviewing_clients=_parse_str(_get_cell_value(row, field_map, "审评中-客户")),
                reviewing_client_count=_parse_int(_get_cell_value(row, field_map, "审评中-客户数量")),
            )
            result["items"].append(item)
            result["success_count"] += 1
        except Exception as e:
            result["error_count"] += 1
            if len(result["errors"]) < 10:
                result["errors"].append(f"第{row_idx}行: {str(e)}")
    
    return result


def export_copp_certificates_to_excel(data: List[CoppCertificate]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "COPP证书"
    
    headers = ["品名", "证书名称", "批件号", "国家/发证机关", "发证日期", "证书有效期至",
               "产品范围", "适用国家", "证书是否过期"]
    ws.append(headers)
    
    for item in data:
        ws.append([
            item.product_name,
            item.certificate_name or "",
            item.batch_no or "",
            item.issuing_authority or "",
            item.issue_date.isoformat() if item.issue_date else "",
            item.valid_until.isoformat() if item.valid_until else "",
            item.product_scope or "",
            item.applicable_countries or "",
            item.is_expired or "",
        ])
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def import_copp_certificates_from_excel(file_content: bytes) -> dict:
    wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
    ws = wb.active
    _fill_merged_cells(ws)
    
    expected_headers = ["品名", "证书名称", "批件号", "国家/发证机关", "发证日期", "证书有效期至",
                        "产品范围", "适用国家", "证书是否过期"]
    header_row, field_map = _find_header_row(ws, expected_headers)
    
    result = {"total_rows": 0, "success_count": 0, "skipped_count": 0, "error_count": 0, "errors": [], "items": []}
    
    if not field_map:
        result["errors"].append("未找到有效表头")
        return result
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        result["total_rows"] += 1
        if not any(row):
            result["skipped_count"] += 1
            continue
        
        product_name = _parse_str(_get_cell_value(row, field_map, "品名"))
        if not product_name:
            result["skipped_count"] += 1
            continue
        
        try:
            item = CoppCertificateCreate(
                product_name=product_name,
                certificate_name=_parse_str(_get_cell_value(row, field_map, "证书名称")),
                batch_no=_parse_str(_get_cell_value(row, field_map, "批件号")),
                issuing_authority=_parse_str(_get_cell_value(row, field_map, "国家/发证机关")),
                issue_date=_parse_date(_get_cell_value(row, field_map, "发证日期")),
                valid_until=_parse_date(_get_cell_value(row, field_map, "证书有效期至")),
                product_scope=_parse_str(_get_cell_value(row, field_map, "产品范围")),
                applicable_countries=_parse_str(_get_cell_value(row, field_map, "适用国家")),
                is_expired=_parse_str(_get_cell_value(row, field_map, "证书是否过期")),
            )
            result["items"].append(item)
            result["success_count"] += 1
        except Exception as e:
            result["error_count"] += 1
            if len(result["errors"]) < 10:
                result["errors"].append(f"第{row_idx}行: {str(e)}")
    
    return result


def export_wc_certificates_to_excel(data: List[WcCertificate]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "WC证书"
    
    headers = ["品名", "证书名称", "批件号", "国家/发证机关", "发证日期", "证书有效期至",
               "产品范围", "证书是否过期"]
    ws.append(headers)
    
    for item in data:
        ws.append([
            item.product_name,
            item.certificate_name or "",
            item.batch_no or "",
            item.issuing_authority or "",
            item.issue_date.isoformat() if item.issue_date else "",
            item.valid_until.isoformat() if item.valid_until else "",
            item.product_scope or "",
            item.is_expired or "",
        ])
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def import_wc_certificates_from_excel(file_content: bytes) -> dict:
    wb = load_workbook(filename=io.BytesIO(file_content), data_only=True)
    ws = wb.active
    _fill_merged_cells(ws)
    
    expected_headers = ["品名", "证书名称", "批件号", "国家/发证机关", "发证日期", "证书有效期至",
                        "产品范围", "证书是否过期"]
    header_row, field_map = _find_header_row(ws, expected_headers)
    
    result = {"total_rows": 0, "success_count": 0, "skipped_count": 0, "error_count": 0, "errors": [], "items": []}
    
    if not field_map:
        result["errors"].append("未找到有效表头")
        return result
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        result["total_rows"] += 1
        if not any(row):
            result["skipped_count"] += 1
            continue
        
        product_name = _parse_str(_get_cell_value(row, field_map, "品名"))
        if not product_name:
            result["skipped_count"] += 1
            continue
        
        try:
            item = WcCertificateCreate(
                product_name=product_name,
                certificate_name=_parse_str(_get_cell_value(row, field_map, "证书名称")),
                batch_no=_parse_str(_get_cell_value(row, field_map, "批件号")),
                issuing_authority=_parse_str(_get_cell_value(row, field_map, "国家/发证机关")),
                issue_date=_parse_date(_get_cell_value(row, field_map, "发证日期")),
                valid_until=_parse_date(_get_cell_value(row, field_map, "证书有效期至")),
                product_scope=_parse_str(_get_cell_value(row, field_map, "产品范围")),
                is_expired=_parse_str(_get_cell_value(row, field_map, "证书是否过期")),
            )
            result["items"].append(item)
            result["success_count"] += 1
        except Exception as e:
            result["error_count"] += 1
            if len(result["errors"]) < 10:
                result["errors"].append(f"第{row_idx}行: {str(e)}")
    
    return result
