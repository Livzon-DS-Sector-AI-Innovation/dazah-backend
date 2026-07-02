"""Generate training sign-in sheet documents from templates."""

import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

from app.modules.hr.schemas import TrainingSignInSheetInput


def _find_template(factory: str = "old") -> Path:
    """Locate the xlsx template by factory."""
    if factory == "new":
        template_name = "R-GN-2002 K 培训签到表.xlsx"
        dir_name = "新厂人员培训管理规程"
    else:
        template_name = "7.5培训签到表.xlsx"
        dir_name = "旧厂员工培训教育管理规程"
    base = Path(__file__).resolve().parent.parent.parent.parent
    candidates = [
        Path(dir_name) / template_name,
        Path("..") / dir_name / template_name,
        base / dir_name / template_name,
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f"模板文件未找到: {dir_name}/{template_name}")


def _save_with_images(wb: openpyxl.Workbook, template_path: Path) -> BytesIO:
    """Save workbook with the template's logo image re-added to the sheet.

    openpyxl discards images on load, so we extract the image from the
    original template and re-add it using openpyxl's add_image() API.
    """
    from openpyxl.drawing.image import Image as XLImage

    # Extract image from template
    with zipfile.ZipFile(template_path, "r") as z:
        img_data = z.read("xl/media/image1.jpeg")

    img = XLImage(BytesIO(img_data))
    ws = wb.active
    ws.add_image(img)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_training_sign_in_sheet(data: TrainingSignInSheetInput, factory: str = "old", page: int = 0) -> BytesIO:
    """Fill the training sign-in sheet template with form data.

    Each page holds up to 30 employees.  Returns a BytesIO buffer
    containing the generated xlsx for the requested page.
    """
    template_path = _find_template(factory)
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active

    if factory == "new":
        _fill_new_factory(ws, data)
    else:
        _fill_old_factory(ws, data)

    # Employee names list (max 30 per page)
    page_size = 30
    start = page * page_size
    page_names = data.employee_names[start : start + page_size]

    if factory == "new":
        # 新厂: 左栏A列 + 右栏C列
        for row in range(15, 30):
            for col in ["A", "C"]:
                cell = ws[f"{col}{row}"]
                if cell and cell.value and str(cell.value).strip().isdigit():
                    cell.value = ""
        for i, name in enumerate(page_names):
            if i < 15:
                row = 15 + i
                ws[f"A{row}"] = name
                ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
            else:
                row = 15 + (i - 15)
                ws[f"C{row}"] = name
                ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")
    else:
        # 旧厂: 左栏A列 + 右栏K列
        for row in range(15, 30):
            for col in ["A", "K"]:
                cell = ws[f"{col}{row}"]
                if cell and cell.value and str(cell.value).strip().isdigit():
                    cell.value = ""
        for i, name in enumerate(page_names):
            if i < 15:
                row = 15 + i
                ws[f"A{row}"] = name
                ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
            else:
                row = 15 + (i - 15)
                ws[f"K{row}"] = name
                ws[f"K{row}"].alignment = Alignment(horizontal="center", vertical="center")

    if factory == "new":
        buffer = _save_with_images(wb, template_path)
    else:
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
    return buffer


def _fill_old_factory(ws, data: TrainingSignInSheetInput) -> None:
    """旧厂签到表模板填充."""
    # Training date: D4=year, F4=month, H4=day
    if data.training_date:
        date_parts = str(data.training_date).split("-")
        if len(date_parts) == 3:
            ws["D4"] = date_parts[0]  # year
            ws["F4"] = date_parts[1]  # month
            ws["H4"] = date_parts[2]  # day

    # Department: D5:I5 merged
    if data.department:
        ws["D5"] = data.department

    # Training method: tick the checkbox in J5:N5
    if data.training_method:
        method_map = {
            "面授": "□面授",
            "函授": "□函授",
            "远程教育": "□远程教育",
            "自学": "□自学",
            "其他": "□其他：",
        }
        j5_value = ws["J5"].value or ""
        for label, placeholder in method_map.items():
            if label in data.training_method:
                j5_value = j5_value.replace(placeholder, f"☑{placeholder[1:]}")
        ws["J5"] = j5_value

    # Attendance count: J6:N6 merged
    total = len(data.employee_names)
    ws["J6"] = f"应受训人数：{total}人   实际受训人数合计：      人"

    # Training time: A8=start_hour, C8=start_min, E8=end_hour, G8=end_min
    if data.training_time_start and data.training_time_end:
        start_parts = data.training_time_start.split(":")
        end_parts = data.training_time_end.split(":")
        if len(start_parts) == 2:
            ws["A8"] = start_parts[0]
            ws["C8"] = start_parts[1]
        if len(end_parts) == 2:
            ws["E8"] = end_parts[0]
            ws["G8"] = end_parts[1]

    # Topic: H8:M8 merged (below the header H7:M7)
    display_topic = data.topic
    if data.training_subject:
        display_topic = f"{data.training_subject} — {data.topic}"
    ws["H8"] = display_topic
    ws["H8"].alignment = Alignment(horizontal="center", vertical="center")

    # Instructor: N8:N12 merged (below the header N7)
    if data.instructor:
        ws["N8"] = data.instructor
        ws["N8"].alignment = Alignment(horizontal="center", vertical="center")

    # Remarks: A30:N30 merged
    if data.remarks:
        ws["A30"] = f"备注：{data.remarks}"


def _fill_new_factory(ws, data: TrainingSignInSheetInput) -> None:
    """新厂签到表模板填充 (R-GN-2002 K)."""
    # Training date: A5 merged — append to template's "培训日期："
    if data.training_date:
        ws["A5"] = f"培训日期：{data.training_date.strftime('%Y年%m月%d日')}"

    # Training method: A6:E6 merged — tick checkboxes
    # Template: "培训方式：□ 面授 □ 现场 □ 远程教育 □ 自学 □其他："
    if data.training_method:
        method_map = {
            "面授": "□ 面授",
            "现场": "□ 现场",
            "远程教育": "□ 远程教育",
            "自学": "□ 自学",
            "其他": "□其他：",
        }
        a6_value = ws["A6"].value or ""
        for label, placeholder in method_map.items():
            if label in data.training_method:
                a6_value = a6_value.replace(placeholder, placeholder.replace("□", "☑"))
        ws["A6"] = a6_value

    # Department: A7:E7 merged
    if data.department:
        ws["A7"] = f"培训部门/班组：{data.department}"

    # Attendance count: A8:E8 merged — preserve underline (\xa0) around number
    # Template: "应培训人数 \xa0\xa0\xa0 人           ..."
    total = len(data.employee_names)
    old_value = ws["A8"].value or ""
    # Replace the \xa0\xa0\xa0 with \xa0[number]\xa0 to keep underline
    import re
    new_value = re.sub(r"\xa0{2,}(?=\s*人)", f"\xa0 {total} \xa0", old_value, count=1)
    ws["A8"] = new_value

    # Training time: A10 — "08：00 ～ 09：00" format (fullwidth ～, ： from template)
    if data.training_time_start and data.training_time_end:
        ws["A10"] = f"{data.training_time_start.replace(':', '：')} ～ {data.training_time_end.replace(':', '：')}"

    # Topic: B10:D10 merged
    display_topic = data.topic
    if data.training_subject:
        display_topic = f"{data.training_subject} — {data.topic}"
    ws["B10"] = display_topic
    ws["B10"].alignment = Alignment(horizontal="center", vertical="center")

    # Instructor: E10:E12 merged (E9 is the fixed header "授课人")
    if data.instructor:
        ws["E10"] = data.instructor
        ws["E10"].alignment = Alignment(horizontal="center", vertical="center")

    # Remarks: A30:D30 merged
    if data.remarks:
        ws["A30"] = f"备注：{data.remarks}"
