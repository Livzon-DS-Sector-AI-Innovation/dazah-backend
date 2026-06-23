"""培训效果评估表 Excel 文档生成器."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from pydantic import BaseModel, Field


class TrainingEvaluationInput(BaseModel):
    subject: str = Field(..., max_length=128, description="培训主题")
    training_date: date | None = Field(None, description="培训日期")
    training_time_start: str | None = Field(None, max_length=32, description="培训开始时间")
    training_time_end: str | None = Field(None, max_length=32, description="培训结束时间")
    duration_hours: float | None = Field(None, description="学时")
    training_method: str | None = Field(None, max_length=32, description="培训方式")
    is_exam: bool = Field(False, description="是否考试")
    trainer_type: str | None = Field(None, max_length=64, description="培训人员类型")
    trainer: str | None = Field(None, max_length=64, description="授课人")
    department_personnel: str | None = Field(None, max_length=256, description="部门/班组/人员")
    expected_count: int | None = Field(None, description="应出席人数")
    actual_count: int | None = Field(None, description="实际出席人数")
    absent_count: int | None = Field(None, description="缺席人数")
    textbook: str | None = Field(None, max_length=256, description="培训教材")
    makeup_training: bool | None = Field(None, description="是否补课")
    assessment_method: str | None = Field(None, max_length=32, description="考核方式")
    pass_count: int | None = Field(None, description="合格人数")
    fail_count: int | None = Field(None, description="不合格人数")
    absent_exam_count: int | None = Field(None, description="缺考人数")
    absent_exam_handling: str | None = Field(None, max_length=512, description="缺考人员处理方式和原因")
    excellent_count: int | None = Field(None, description="优秀人数")
    qualified_count: int | None = Field(None, description="合格人数")
    unqualified_count: int | None = Field(None, description="不合格人数")
    evaluation_conclusion: str | None = Field(None, max_length=1024, description="培训效果评估及结论")
    organizer: str | None = Field(None, max_length=64, description="培训组织人")
    organizer_date: date | None = Field(None, description="组织日期")
    remarks: str | None = Field(None, max_length=512, description="备注")


def _cell_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_training_evaluation(data: TrainingEvaluationInput) -> BytesIO:
    """根据填写的培训信息生成培训效果评估表 Excel 文档."""
    wb = Workbook()
    ws = wb.active
    ws.title = "培训效果评估表"

    # 列宽设置
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16

    # 标题行
    ws.merge_cells("A1:E1")
    ws["A1"] = "QR.SOP.PM.003/18（格式）  P8/12"
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    ws.merge_cells("A2:E2")
    ws["A2"] = "丽珠集团新北江制药股份有限公司"
    ws["A2"].font = Font(name="宋体", size=14, bold=True)
    ws["A2"].alignment = _center_align()
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A3:E3")
    ws["A3"] = "培训效果评估表"
    ws["A3"].font = Font(name="宋体", size=16, bold=True)
    ws["A3"].alignment = _center_align()
    ws.row_dimensions[3].height = 32

    # 培训主题
    ws.merge_cells("A4:E4")
    ws["A4"] = f"培训主题：{data.subject}"
    ws["A4"].alignment = _left_align()
    ws["A4"].border = _cell_border()
    ws.row_dimensions[4].height = 24

    # 培训时间 + 学时
    time_str = ""
    if data.training_date:
        time_str = data.training_date.strftime("%Y年%m月%d日")
    if data.training_time_start and data.training_time_end:
        time_str += f" {data.training_time_start}~{data.training_time_end}"
    ws.merge_cells("A5:C5")
    ws["A5"] = f"培训时间：{time_str}"
    ws["A5"].alignment = _left_align()
    ws["A5"].border = _cell_border()
    ws["D5"] = f"学时：{data.duration_hours if data.duration_hours is not None else ''}"
    ws["D5"].alignment = _left_align()
    ws["D5"].border = _cell_border()
    ws["E5"].border = _cell_border()
    ws.row_dimensions[5].height = 24

    # 培训方式 + 是否考试
    method_map = {
        "面授": "☑面授  □函授  □远程教育  □自学  □其他方式",
        "函授": "□面授  ☑函授  □远程教育  □自学  □其他方式",
        "远程教育": "□面授  □函授  ☑远程教育  □自学  □其他方式",
        "自学": "□面授  □函授  □远程教育  ☑自学  □其他方式",
    }
    method_str = method_map.get(data.training_method, "□面授  □函授  □远程教育  □自学  □其他方式")
    ws.merge_cells("A6:C6")
    ws["A6"] = f"培训方式：{method_str}"
    ws["A6"].alignment = _left_align()
    ws["A6"].border = _cell_border()
    exam_str = "☑考试" if data.is_exam else "□考试"
    ws["D6"] = exam_str
    ws["D6"].alignment = _left_align()
    ws["D6"].border = _cell_border()
    ws["E6"].border = _cell_border()
    ws.row_dimensions[6].height = 28

    # 培训人员
    ws.merge_cells("A7:E7")
    trainer_type = data.trainer_type or ""
    ws["A7"] = f"培训人员：□讲师/专家/官员等    {trainer_type}"
    ws["A7"].alignment = _left_align()
    ws["A7"].border = _cell_border()
    ws.row_dimensions[7].height = 24

    # 应出席/实际出席/缺席人数
    ws.merge_cells("A8:E8")
    expected = data.expected_count if data.expected_count is not None else "___"
    actual = data.actual_count if data.actual_count is not None else "___"
    absent = data.absent_count if data.absent_count is not None else "___"
    ws["A8"] = f"应出席 {expected} 人；实际出席 {actual} 人；缺席 {absent} 人。"
    ws["A8"].alignment = _left_align()
    ws["A8"].border = _cell_border()
    ws.row_dimensions[8].height = 24

    # 培训教材
    ws.merge_cells("A9:E9")
    ws["A9"] = f"培训教材：{data.textbook or ''}"
    ws["A9"].alignment = _left_align()
    ws["A9"].border = _cell_border()
    ws.row_dimensions[9].height = 24

    # 缺席人员处理方式
    ws.merge_cells("A10:E10")
    makeup_str = ""
    if data.makeup_training is True:
        makeup_str = "是否进行补课培训，☑是 □否，未参加培训人员必须补上培训内容，（包括培训时间、地点、方式等）。"
    elif data.makeup_training is False:
        makeup_str = "是否进行补课培训，□是 ☑否，未参加培训人员必须补上培训内容，（包括培训时间、地点、方式等）。"
    else:
        makeup_str = "是否进行补课培训，□是 □否，未参加培训人员必须补上培训内容，（包括培训时间、地点、方式等）。"
    ws["A10"] = f"缺席人员处理方式：\n{makeup_str}"
    ws["A10"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws["A10"].border = _cell_border()
    ws.row_dimensions[10].height = 48

    # 考核方式
    ws.merge_cells("A11:E11")
    am = data.assessment_method or ""
    am_map = {
        "笔试": "☑ 笔试    □ 口试   □ 实操   □ 写总结",
        "口试": "□ 笔试    ☑ 口试   □ 实操   □ 写总结",
        "实操": "□ 笔试    □ 口试   ☑ 实操   □ 写总结",
        "写总结": "□ 笔试    □ 口试   □ 实操   ☑ 写总结",
    }
    am_str = am_map.get(am, "□ 笔试    □ 口试   □ 实操   □ 写总结")
    ws["A11"] = f"考核方式：{am_str}"
    ws["A11"].alignment = _left_align()
    ws["A11"].border = _cell_border()
    ws.row_dimensions[11].height = 24

    # 考核结果
    ws.merge_cells("A12:E12")
    p_cnt = data.pass_count if data.pass_count is not None else "___"
    f_cnt = data.fail_count if data.fail_count is not None else "___"
    ae_cnt = data.absent_exam_count if data.absent_exam_count is not None else "___"
    ws["A12"] = f"考核结果：□合格 {p_cnt} 人；□不合格 {f_cnt} 人；缺考 {ae_cnt} 人。"
    ws["A12"].alignment = _left_align()
    ws["A12"].border = _cell_border()
    ws.row_dimensions[12].height = 24

    # 缺考人员处理方式和原因
    ws.merge_cells("A13:E13")
    ws["A13"] = f"缺考人员处理方式和原因：{data.absent_exam_handling or ''}"
    ws["A13"].alignment = _left_align()
    ws["A13"].border = _cell_border()
    ws.row_dimensions[13].height = 36

    # 综合评分
    ws.merge_cells("A14:E14")
    ex_cnt = data.excellent_count if data.excellent_count is not None else "___"
    q_cnt = data.qualified_count if data.qualified_count is not None else "___"
    uq_cnt = data.unqualified_count if data.unqualified_count is not None else "___"
    ws["A14"] = f"综合评分：□优秀 {ex_cnt} 人；□合格 {q_cnt} 人；□不合格 {uq_cnt} 人。"
    ws["A14"].alignment = _left_align()
    ws["A14"].border = _cell_border()
    ws.row_dimensions[14].height = 24

    # 空行
    ws.merge_cells("A15:E15")
    ws["A15"] = ""
    ws["A15"].border = _cell_border()
    ws.row_dimensions[15].height = 24

    # 培训效果评估及结论
    ws.merge_cells("A16:E17")
    conclusion = data.evaluation_conclusion or ""
    ws["A16"] = f"培训效果评估及结论：\n{conclusion}"
    ws["A16"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws["A16"].border = _cell_border()
    ws["A17"].border = _cell_border()
    ws.row_dimensions[16].height = 24
    ws.row_dimensions[17].height = 48

    # 培训组织人/日期
    ws.merge_cells("A18:E18")
    org_date = ""
    if data.organizer:
        org_date += data.organizer
    if data.organizer_date:
        org_date += f" / {data.organizer_date.strftime('%Y年%m月%d日')}"
    ws["A18"] = f"培训组织人/日期：{org_date}"
    ws["A18"].alignment = _left_align()
    ws["A18"].border = _cell_border()
    ws.row_dimensions[18].height = 24

    # 备注
    ws.merge_cells("A19:E20")
    ws["A19"] = f"备注：\n{data.remarks or ''}"
    ws["A19"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws["A19"].border = _cell_border()
    ws["A20"].border = _cell_border()
    ws.row_dimensions[19].height = 24
    ws.row_dimensions[20].height = 36

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
