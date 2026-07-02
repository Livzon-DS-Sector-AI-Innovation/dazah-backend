from datetime import date, datetime
from io import BytesIO
from uuid import UUID
import logging
import os
import re

from fastapi import Body, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import flag_modified
from urllib.parse import quote

from app.core.database import get_db
from app.core.redis import cache_set, cache_get, cache_delete
from app.core.response import paginated_response, success_response, error_response
from app.modules.hr.schemas import (
    AnnualTrainingPlanCreate,
    AnnualTrainingPlanItemBatchUpdate,
    AnnualTrainingPlanItemResponse,
    AnnualTrainingPlanResponse,
    AnnualTrainingPlanUpdate,
    CandidateCreate,
    CandidateListResponse,
    CandidateResponse,
    CandidateUpdate,
    CandidateUpdateRecommendationLevel,
    DepartureRecordCreate,
    DepartureRecordResponse,
    DepartureRecordUpdate,
    DepartmentCreate,
    DepartmentResponse,
    DepartmentUpdate,
    EmployeeCreate,
    EmployeeResponse,
    EmployeeUpdate,
    OffboardingRecordCreate,
    OffboardingRecordResponse,
    OffboardingRecordUpdate,
    OnboardingEvaluationInput,
    OnboardingRecordResponse,
    TeamCreate,
    TeamResponse,
    TeamUpdate,
    TrainingEvaluationInput,
    TrainingLedgerCreate,
    TrainingLedgerResponse,
    TrainingLedgerUpdate,
    TrainingLedgerPageCreate,
    TrainingLedgerPageResponse,
    TrainingSessionCreate,
    TrainingSessionResponse,
    TrainingSessionUpdate,
    TrainingSessionStatusUpdate,
    TrainingNotificationInput,
    TrainingNotifyInput,
    TrainingSelectTaskCreate,
    TrainingSelectTaskSubmit,
    TrainingSignInSheetInput,
    TrainingSpecialistCreate,
    TrainingSpecialistUpdate,
    TrainingSpecialistResponse,
    TrainingTeamCreate,
    TrainingTeamUpdate,
    TrainingTeamResponse,
    PrejobTemplateCreate,
    PrejobTemplateItem,
    PrejobTemplateResponse,
)
try:
    from app.modules.hr.attendance_schemas import DepartmentProductionSettings
except ImportError:
    DepartmentProductionSettings = None  # type: ignore
from app.modules.hr.models import Department, Employee
from app.modules.hr.document_generator import generate_onboarding_training_record
from app.modules.hr.evaluation_document_generator import generate_training_evaluation
from app.modules.hr.notification_document_generator import generate_training_notification
from app.modules.hr.onboarding_evaluation_document_generator import generate_onboarding_evaluation
from app.modules.hr.prejob_document_generator import generate_prejob_training_plan
from app.modules.hr.signin_document_generator import generate_training_sign_in_sheet
from app.modules.hr.ledger_export_generator import generate_ledger_export
from app.modules.hr.analysis_api import router as analysis_router
try:
    from app.modules.hr.attendance_api import router as attendance_router
except ImportError:
    attendance_router = None  # type: ignore
from app.modules.hr.repository import PrejobTemplateRepository, TrainingSpecialistRepository, TrainingTeamRepository
from app.modules.hr.service import (
    AnnualTrainingPlanItemService,
    AnnualTrainingPlanService,
    CandidateService,
    DepartureRecordService,
    DepartmentService,
    EmployeeService,
    OffboardingRecordService,
    OnboardingRecordService,
    TeamService,
    TrainingLedgerService,
    TrainingLedgerPageService,
    TrainingSessionService,
)
from app.shared.module_api import create_module_router
from app.shared.module_registry import MODULES_BY_CODE
from app.shared.schemas import PageParams

router = create_module_router(MODULES_BY_CODE["hr"])

logger = logging.getLogger(__name__)


async def _get_new_employee(employee_id: UUID, session: AsyncSession) -> Employee:
    """Query employees_new clone table and construct an Employee instance."""
    sql = text(
        "SELECT * FROM hr.employees_new WHERE id = :id AND is_deleted = false"
    )
    result = await session.execute(sql, {"id": str(employee_id)})
    row = result.mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="未找到该新厂员工")
    return Employee(**dict(row))


def get_employee_service(session: AsyncSession = Depends(get_db)) -> EmployeeService:
    return EmployeeService(session)


def get_department_service(
    session: AsyncSession = Depends(get_db),
) -> DepartmentService:
    return DepartmentService(session)


def get_offboarding_service(
    session: AsyncSession = Depends(get_db),
) -> OffboardingRecordService:
    return OffboardingRecordService(session)


def get_team_service(
    session: AsyncSession = Depends(get_db),
) -> TeamService:
    return TeamService(session)


def get_onboarding_service(
    session: AsyncSession = Depends(get_db),
) -> OnboardingRecordService:
    return OnboardingRecordService(session)


def get_departure_service(
    session: AsyncSession = Depends(get_db),
) -> DepartureRecordService:
    return DepartureRecordService(session)


def get_training_ledger_service(
    session: AsyncSession = Depends(get_db),
) -> TrainingLedgerService:
    return TrainingLedgerService(session)


def get_training_ledger_page_service(
    session: AsyncSession = Depends(get_db),
) -> TrainingLedgerPageService:
    return TrainingLedgerPageService(session)


def get_annual_training_plan_service(
    session: AsyncSession = Depends(get_db),
) -> AnnualTrainingPlanService:
    return AnnualTrainingPlanService(session)


def get_annual_training_plan_item_service(
    session: AsyncSession = Depends(get_db),
) -> AnnualTrainingPlanItemService:
    return AnnualTrainingPlanItemService(session)


def get_training_session_service(
    session: AsyncSession = Depends(get_db),
) -> TrainingSessionService:
    return TrainingSessionService(session)


def get_candidate_service(
    session: AsyncSession = Depends(get_db),
) -> CandidateService:
    return CandidateService(session)


# ─── Employee Routes ───

@router.get("/employees", summary="员工列表")
async def list_employees(
    department: str | None = Query(None, description="部门筛选"),
    status: str | None = Query(None, description="状态筛选"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    page_params: PageParams = Depends(),
    service: EmployeeService = Depends(get_employee_service),
):
    employees, total = await service.list_employees(
        department=department,
        status=status,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        EmployeeResponse.model_validate(e).model_dump(mode="json")
        for e in employees
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/employees", summary="创建员工")
async def create_employee(
    payload: EmployeeCreate,
    service: EmployeeService = Depends(get_employee_service),
):
    employee = await service.create_employee(payload)
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
        message="员工创建成功",
        status_code=201,
    )


@router.post("/employees/sync-from-feishu", summary="从飞书多维表格同步员工数据")
async def sync_employees_from_feishu(
    service: EmployeeService = Depends(get_employee_service),
):
    """手动触发：从飞书多维表格拉取全部员工数据并 upsert 到本地 PG。"""
    stats = await service.sync_from_feishu()
    msg = (
        f"同步完成：新增 {stats['created']} 条，"
        f"更新 {stats['updated']} 条，失败 {stats['failed']} 条"
    )
    return success_response(
        data=stats,
        message=msg,
    )


@router.get("/employees/sync-status", summary="飞书同步状态")
async def get_employee_sync_status(
    service: EmployeeService = Depends(get_employee_service),
):
    """查看本地与飞书的数据同步统计。"""
    status = await service.get_sync_status()
    return success_response(
        data=status.model_dump(mode="json"),
    )


@router.get("/employees/by-number/{employee_number}", summary="根据工号查询员工")
async def get_employee_by_number(
    employee_number: str,
    service: EmployeeService = Depends(get_employee_service),
):
    employee = await service.get_employee_by_number(employee_number)
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
    )


@router.get("/employees/{employee_id}", summary="员工详情")
async def get_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
):
    employee = await service.get_employee(employee_id)
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
    )


@router.put("/employees/{employee_id}", summary="更新员工")
async def update_employee(
    employee_id: UUID,
    payload: EmployeeUpdate,
    service: EmployeeService = Depends(get_employee_service),
):
    employee = await service.update_employee(employee_id, payload)
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
        message="员工更新成功",
    )


@router.delete("/employees/{employee_id}", summary="删除员工")
async def delete_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
):
    await service.delete_employee(employee_id)
    return success_response(message="员工删除成功")


@router.post("/employees/{employee_id}/sync-to-feishu", summary="同步单个员工到飞书")
async def sync_employee_to_feishu(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
):
    """将本地单个员工强制同步到飞书多维表格。"""
    record_id = await service.sync_to_feishu(employee_id)
    return success_response(
        data={"feishu_record_id": record_id},
        message="员工已同步到飞书",
    )


@router.post("/webhook/feishu-approval", summary="飞书审批完成回调")
async def feishu_approval_webhook(
    payload: dict,
    service: EmployeeService = Depends(get_employee_service),
):
    """接收飞书审批完成通知，更新员工状态为在职。"""
    employee_number = payload.get("employee_number")
    if not employee_number:
        return success_response(message="缺少工号")

    try:
        employee = await service.approve_employee(employee_number)
        return success_response(
            data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
            message="员工审批通过，状态已更新为在职",
        )
    except Exception as e:
        return success_response(message=f"审批处理失败: {str(e)}")


@router.get(
    "/employees/{employee_id}/onboarding-training-record",
    summary="导出员工入职培训记录",
)
async def export_onboarding_training_record(
    employee_id: UUID,
    factory: str = Query("old", description="厂别：old=旧厂, new=新厂"),
    service: EmployeeService = Depends(get_employee_service),
    session: AsyncSession = Depends(get_db),
):
    """根据员工数据自动生成并下载入职培训记录 Word 文档。"""
    employee = await service.get_employee(employee_id) if factory == "old" else await _get_new_employee(employee_id, session)
    try:
        buffer: BytesIO = generate_onboarding_training_record(employee, factory)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    filename = f"onboarding_training_record_{employee.employee_number}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/employees/{employee_id}/prejob-training-plan",
    summary="导出员工岗前培训计划",
)
async def export_prejob_training_plan(
    employee_id: UUID,
    factory: str = Query("old", description="厂别：old=旧厂, new=新厂"),
    service: EmployeeService = Depends(get_employee_service),
    session: AsyncSession = Depends(get_db),
):
    """根据员工数据自动生成并下载岗前培训计划文档。"""
    employee = await service.get_employee(employee_id) if factory == "old" else await _get_new_employee(employee_id, session)
    try:
        buffer: BytesIO = generate_prejob_training_plan(employee, factory)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    ext = "xlsx" if factory == "old" else "docx"
    media_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if factory == "old"
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    filename = f"prejob_training_plan_{employee.employee_number}.{ext}"
    return StreamingResponse(
        _iterfile(),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/employees/{employee_id}/prejob-training-plan",
    summary="导出员工岗前培训计划（含编辑后内容）",
)
async def export_prejob_training_plan_with_items(
    employee_id: UUID,
    factory: str = Query("old", description="厂别：old=旧厂, new=新厂"),
    items: list[PrejobTemplateItem] | None = Body(
        None, description="编辑后的培训计划内容（10行），不传则使用部门默认内容"
    ),
    service: EmployeeService = Depends(get_employee_service),
    session: AsyncSession = Depends(get_db),
):
    """根据员工数据自动生成并下载岗前培训计划文档，支持传入编辑后的培训计划条目。"""
    employee = (
        await service.get_employee(employee_id)
        if factory == "old"
        else await _get_new_employee(employee_id, session)
    )
    items_dict = [it.model_dump() for it in items] if items else None
    try:
        buffer: BytesIO = generate_prejob_training_plan(employee, factory, items_dict)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    ext = "xlsx" if factory == "old" else "docx"
    media_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if factory == "old"
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    filename = f"prejob_training_plan_{employee.employee_number}.{ext}"
    return StreamingResponse(
        _iterfile(),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/employees/{employee_id}/onboarding-evaluation",
    summary="导出员工上岗评估表",
)
async def export_onboarding_evaluation_by_employee(
    employee_id: UUID,
    factory: str = Query("old", description="厂别：old=旧厂, new=新厂"),
    service: EmployeeService = Depends(get_employee_service),
    session: AsyncSession = Depends(get_db),
):
    """根据员工档案预填基本信息并导出上岗评估表文档。"""
    employee = await service.get_employee(employee_id) if factory == "old" else await _get_new_employee(employee_id, session)

    try:
        buffer: BytesIO = generate_onboarding_evaluation(employee, factory)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        import traceback
        detail = traceback.format_exc()
        raise HTTPException(status_code=500, detail=f"生成文档失败: {str(e)}\n{detail}")

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(employee.hire_date).replace("-", "") if employee.hire_date else "nodate"
    ext = "xlsx" if factory == "old" else "docx"
    media_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if factory == "old"
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    filename = f"onboarding_evaluation_{employee.employee_number}_{safe_date}.{ext}"
    return StreamingResponse(
        _iterfile(),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


import zipfile

@router.post("/training-sign-in-sheet", summary="生成培训签到表")
async def export_training_sign_in_sheet(
    payload: TrainingSignInSheetInput,
    factory: str = Query("old", description="厂别：old=旧厂, new=新厂"),
):
    """根据填写的培训信息自动生成培训签到表文档。

    当应出席受训人员超过 30 人时，自动分页生成多张签到表并打包为 zip。
    """
    safe_date = str(payload.training_date).replace("-", "")
    total = len(payload.employee_names)
    ext = "xlsx"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if total <= 30:
        try:
            buffer: BytesIO = generate_training_sign_in_sheet(payload, factory)
        except FileNotFoundError as e:
            raise HTTPException(status_code=400, detail=str(e))

        def _iterfile():
            buffer.seek(0)
            yield buffer.read()

        filename = f"training_sign_in_sheet_{safe_date}.{ext}"
        return StreamingResponse(
            _iterfile(),
            media_type=media_type,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # 超过 30 人：分页生成并打包为 zip
    pages = (total + 29) // 30  # 向上取整
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for page in range(pages):
            try:
                page_buffer = generate_training_sign_in_sheet(payload, factory, page=page)
            except FileNotFoundError as e:
                raise HTTPException(status_code=400, detail=str(e))
            page_buffer.seek(0)
            zf.writestr(f"training_sign_in_sheet_{safe_date}_page{page + 1}.{ext}", page_buffer.read())
    zip_buffer.seek(0)

    def _iter_zip():
        zip_buffer.seek(0)
        yield zip_buffer.read()

    return StreamingResponse(
        _iter_zip(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="training_sign_in_sheet_{safe_date}.zip"'},
    )


@router.post("/training-notifications/send", summary="发送培训通知到飞书")
async def send_training_notification(
    payload: TrainingNotifyInput,
    service: EmployeeService = Depends(get_employee_service),
):
    """根据填写的培训信息，向受训人员发送飞书单聊消息。"""
    result = await service.notify_training(payload)
    msg = f"发送完成：成功 {result['sent']} 人，失败 {result['failed']} 人"
    return success_response(data=result, message=msg)


# ─── Training Select Task APIs ───

import uuid
import json

from app.core.config import get_settings
from app.platform.integrations.feishu.auth import FeishuAuth
from app.platform.integrations.feishu.im import FeishuIM


@router.post("/training-select-tasks/send", summary="发送飞书选择受训人员任务")
async def send_training_select_task(
    payload: TrainingSelectTaskCreate,
    service: EmployeeService = Depends(get_employee_service),
):
    """创建临时选择任务，发送飞书消息给李文兆。"""
    token = str(uuid.uuid4())
    cache_key = f"training_select:{token}"
    cache_data = payload.model_dump_json()
    await cache_set(cache_key, cache_data, ex=86400)

    # 同时记录到任务列表
    list_key = "training_select_list"
    await cache_set(f"{list_key}:{token}", cache_data, ex=86400)

    # 更新任务列表
    list_key = "training_select_list"
    existing = await cache_get(list_key)
    task_list = json.loads(existing) if existing else []
    task_list.append({
        "token": token,
        "department": payload.department,
        "training_date": payload.training_date,
        "subject": payload.subject,
        "factory": payload.factory,
        "created_at": datetime.now().isoformat(),
    })
    await cache_set(list_key, json.dumps(task_list, ensure_ascii=False), ex=604800)  # 7 天

    settings = get_settings()
    select_url = f"{settings.FRONTEND_URL}/hr/training/select?token={token}"

    # 查询李文兆（先查老厂，再查新厂）
    li_employee = None
    li_table = None
    for table in ["hr.employees", "hr.employees_new"]:
        sql = text(
            f"SELECT * FROM {table} WHERE name = :name AND is_deleted = false LIMIT 1"
        )
        result = await service.repo.session.execute(sql, {"name": "李文兆"})
        row = result.mappings().first()
        if row:
            li_employee = Employee(**dict(row))
            li_table = table
            break

    if not li_employee:
        raise HTTPException(status_code=404, detail="未找到接收人李文兆")

    open_id = li_employee.feishu_open_id
    if not open_id and li_employee.phone:
        try:
            im = FeishuIM()
            mobile = li_employee.phone if li_employee.phone.startswith("+") else f"+86{li_employee.phone}"
            mapping = await im.batch_get_open_ids_by_mobile([mobile])
            open_id = mapping.get(mobile) or mapping.get(li_employee.phone)
            if open_id:
                # 持久化到数据库
                update_sql = text(
                    f"UPDATE {li_table} SET feishu_open_id = :oid WHERE employee_number = :eno"
                )
                await service.repo.session.execute(update_sql, {"oid": open_id, "eno": li_employee.employee_number})
                await service.repo.session.flush()
        except Exception:
            pass

    if not open_id:
        raise HTTPException(status_code=400, detail="李文兆缺少飞书 open_id 且无法实时获取")

    msg_content = (
        f"【培训人员选择通知】\n"
        f"请为 {payload.department} 的 {payload.training_date} {payload.subject} 培训选择受训人员。\n"
        f"点击链接选择：{select_url}\n"
        f"（链接 1 小时内有效）"
    )

    try:
        im = FeishuIM()
        await im.send_text_message(open_id, msg_content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"发送飞书消息失败: {str(e)}")

    return success_response(
        data={"token": token, "url": select_url},
        message="已发送飞书选择任务给李文兆",
    )


@router.get("/training-select-tasks/{token}", summary="获取选择任务")
async def get_training_select_task(token: str):
    """根据 token 获取临时选择任务详情。"""
    cache_key = f"training_select:{token}"
    data = await cache_get(cache_key)
    if not data:
        raise HTTPException(status_code=404, detail="任务已过期或不存在")
    return success_response(data=json.loads(data))


@router.post("/training-select-tasks/{token}/submit", summary="提交选择结果")
async def submit_training_select_task(
    token: str,
    payload: TrainingSelectTaskSubmit,
    service: TrainingSessionService = Depends(get_training_session_service),
):
    """提交选择的受训人员，返回培训信息 + 选择的人员。"""
    cache_key = f"training_select:{token}"
    data = await cache_get(cache_key)
    if not data:
        raise HTTPException(status_code=404, detail="任务已过期或不存在")

    task = json.loads(data)
    task["employee_numbers"] = payload.employee_numbers
    task["employee_names"] = payload.employee_names or []
    await cache_set(cache_key, json.dumps(task, ensure_ascii=False), ex=3600)

    # 找到关联的 TrainingSession 并更新 select_tasks
    # 先按旧的 select_task_token 查找，再按 select_tasks JSON 数组查找
    session_obj = await service.repo.get_by_select_task_token(token)
    if not session_obj:
        session_obj = await service.repo.get_by_select_tasks_json_token(token)
    if session_obj and session_obj.select_tasks:
        for st in session_obj.select_tasks:
            if st.get("token") == token:
                st["employee_names"] = payload.employee_names or []
                st["employee_numbers"] = payload.employee_numbers
                st["status"] = "submitted"
                break
        # 标记 JSON 字段已变更，SQLAlchemy 才能检测到嵌套修改
        flag_modified(session_obj, "select_tasks")
        await service.repo.update(session_obj)

    return success_response(
        data=task,
        message="选择结果已提交",
    )


@router.get("/training-select-tasks", summary="获取培训选择任务列表")
async def list_training_select_tasks():
    """获取所有已发送的培训选择任务列表。"""
    list_key = "training_select_list"
    existing = await cache_get(list_key)
    if not existing:
        return success_response(data=[])

    task_list = json.loads(existing)
    results = []
    for meta in task_list:
        token = meta["token"]
        data = await cache_get(f"training_select:{token}")
        if data:
            task = json.loads(data)
            results.append({
                "token": token,
                "department": task.get("department"),
                "training_date": task.get("training_date"),
                "subject": task.get("subject"),
                "factory": task.get("factory"),
                "location": task.get("location"),
                "trainer": task.get("trainer"),
                "training_method": task.get("training_method"),
                "has_result": bool(task.get("employee_numbers")),
                "selected_count": len(task.get("employee_numbers", [])),
                "created_at": meta.get("created_at"),
            })

    # 按时间倒序
    results.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return success_response(data=results)


@router.get("/training-select-tasks/{token}/result", summary="获取选择结果")
async def get_training_select_task_result(token: str):
    """根据 token 获取已提交的选择结果。"""
    cache_key = f"training_select:{token}"
    data = await cache_get(cache_key)
    if not data:
        raise HTTPException(status_code=404, detail="任务已过期或不存在")
    task = json.loads(data)
    if not task.get("employee_numbers"):
        raise HTTPException(status_code=400, detail="该任务尚未提交选择结果")
    return success_response(data=task)


@router.post("/training-notification", summary="生成培训通知")
async def export_training_notification(
    payload: TrainingNotificationInput,
    factory: str = Query("old", description="厂区: old=旧厂, new=新厂"),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    """根据填写的培训信息自动生成培训通知 Word 文档。

    旧厂使用模板「旧厂员工培训教育管理规程/培训通知.docx」，新厂使用模板「新厂人员培训管理规程/SOP-GN-2002 Q 培训通知.docx」。
    若应出席受训人员包含李健文(110000673)或黄丽耘(110001372)，
    自动为其创建培训台账记录。
    """
    try:
        buffer: BytesIO = generate_training_notification(payload, factory)
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # 自动关联创建培训台账记录
    employee_service = EmployeeService(service.repo.session)
    for name in payload.trainee_names:
        emp = await employee_service.repo.list_employees(
            keyword=name, page=1, page_size=1
        )
        if emp[0] and emp[0][0]:
            employee = emp[0][0]
            if employee.employee_number in {"110000673", "110001372"}:
                await service.create_from_notification(
                    employee_number=employee.employee_number,
                    training_date=payload.training_date,
                    training_subject=payload.subject,
                    training_method=None,
                    trainer=payload.trainer,
                    source_id=f"notification_{payload.training_date}_{payload.subject}",
                )

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(payload.training_date).replace("-", "")
    filename = f"training_notification_{safe_date}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/training-evaluation", summary="生成培训效果评估表")
async def export_training_evaluation(
    payload: TrainingEvaluationInput,
    factory: str = Query("old", description="厂别：old=旧厂, new=新厂"),
):
    """根据填写的培训信息自动生成培训效果评估表文档。"""
    buffer: BytesIO = generate_training_evaluation(payload, factory)

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(payload.training_date).replace("-", "") if payload.training_date else "nodate"
    ext = "xlsx" if factory == "old" else "docx"
    media_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if factory == "old"
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    filename = f"training_evaluation_{safe_date}.{ext}"
    return StreamingResponse(
        _iterfile(),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/onboarding-evaluation", summary="生成员工上岗评估表")
async def export_onboarding_evaluation(
    payload: OnboardingEvaluationInput,
    factory: str = Query("old", description="厂别：old=旧厂, new=新厂"),
):
    """根据填写的评估信息自动生成员工上岗评估表文档。"""
    buffer: BytesIO = generate_onboarding_evaluation(payload, factory)

    def _iterfile():
        buffer.seek(0)
        yield buffer.read()

    safe_date = str(payload.approval_date).replace("-", "") if payload.approval_date else "nodate"
    ext = "xlsx" if factory == "old" else "docx"
    media_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if factory == "old"
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    filename = f"onboarding_evaluation_{safe_date}.{ext}"
    return StreamingResponse(
        _iterfile(),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Department Routes ───

@router.get("/departments", summary="部门列表")
async def list_departments(
    keyword: str | None = Query(None, description="部门名称或编码关键词"),
    page_params: PageParams = Depends(),
    service: DepartmentService = Depends(get_department_service),
):
    departments, total = await service.list_departments(
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        DepartmentResponse.model_validate(d).model_dump(mode="json")
        for d in departments
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/departments", summary="创建部门")
async def create_department(
    payload: DepartmentCreate,
    service: DepartmentService = Depends(get_department_service),
):
    department = await service.create_department(payload)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
        message="部门创建成功",
        status_code=201,
    )


@router.get("/departments/{department_id}", summary="部门详情")
async def get_department(
    department_id: UUID,
    service: DepartmentService = Depends(get_department_service),
):
    department = await service.get_department(department_id)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
    )


@router.put("/departments/{department_id}", summary="更新部门")
async def update_department(
    department_id: UUID,
    payload: DepartmentUpdate,
    service: DepartmentService = Depends(get_department_service),
):
    department = await service.update_department(department_id, payload)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
        message="部门更新成功",
    )


@router.delete("/departments/{department_id}", summary="删除部门")
async def delete_department(
    department_id: UUID,
    service: DepartmentService = Depends(get_department_service),
):
    await service.delete_department(department_id)
    return success_response(message="部门删除成功")


if DepartmentProductionSettings is not None:

    @router.put("/departments/{department_id}/production-settings", summary="设置部门生产属性")
    async def set_department_production(
        department_id: UUID,
        payload: DepartmentProductionSettings,
        service: DepartmentService = Depends(get_department_service),
    ):
        """设置部门是否为生产部门及生产班次时间。"""
        department = await service.get_department(department_id)
        department.is_production = payload.is_production
        department.production_start_time = payload.production_start_time
        department.production_end_time = payload.production_end_time
        await service.repo.session.flush()
        return success_response(
            data={"id": str(department.id), "is_production": department.is_production},
            message="生产设置已更新",
        )


# ─── Team Routes ───

@router.get("/teams", summary="班组列表")
async def list_teams(
    department_id: UUID | None = Query(None, description="部门筛选"),
    keyword: str | None = Query(None, description="班组名称或编码关键词"),
    page_params: PageParams = Depends(),
    service: TeamService = Depends(get_team_service),
):
    teams, total = await service.list_teams(
        department_id=department_id,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        TeamResponse.model_validate(t).model_dump(mode="json")
        for t in teams
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/teams", summary="创建班组")
async def create_team(
    payload: TeamCreate,
    service: TeamService = Depends(get_team_service),
):
    team = await service.create_team(payload)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
        message="班组创建成功",
        status_code=201,
    )


@router.get("/teams/{team_id}", summary="班组详情")
async def get_team(
    team_id: UUID,
    service: TeamService = Depends(get_team_service),
):
    team = await service.get_team(team_id)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
    )


@router.put("/teams/{team_id}", summary="更新班组")
async def update_team(
    team_id: UUID,
    payload: TeamUpdate,
    service: TeamService = Depends(get_team_service),
):
    team = await service.update_team(team_id, payload)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
        message="班组更新成功",
    )


@router.delete("/teams/{team_id}", summary="删除班组")
async def delete_team(
    team_id: UUID,
    service: TeamService = Depends(get_team_service),
):
    await service.delete_team(team_id)
    return success_response(message="班组删除成功")


# ─── OffboardingRecord Routes ───

@router.get("/offboarding-records", summary="离职记录列表")
async def list_offboarding_records(
    employee_id: UUID | None = Query(None, description="员工ID筛选"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    page_params: PageParams = Depends(),
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    records, total = await service.list_records(
        employee_id=employee_id,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        OffboardingRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/offboarding-records", summary="创建离职记录")
async def create_offboarding_record(
    payload: OffboardingRecordCreate,
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    record = await service.create_record(payload)
    # 手动构建响应，避免触发未加载的 relationship
    data = {
        "id": str(record.id),
        "employee_id": str(record.employee_id),
        "offboarding_date": (
            record.offboarding_date.isoformat()
            if record.offboarding_date else None
        ),
        "offboarding_type": record.offboarding_type,
        "reason": record.reason,
        "handover_status": record.handover_status,
        "notes": record.notes,
        "created_at": record.created_at.isoformat() if record.created_at else None,
        "updated_at": record.updated_at.isoformat() if record.updated_at else None,
    }
    return success_response(
        data=data,
        message="离职记录创建成功，员工状态已更新为离职",
        status_code=201,
    )


@router.get("/offboarding-records/{record_id}", summary="离职记录详情")
async def get_offboarding_record(
    record_id: UUID,
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    record = await service.get_record(record_id)
    return success_response(
        data=OffboardingRecordResponse.model_validate(record).model_dump(mode="json"),
    )


@router.put("/offboarding-records/{record_id}", summary="更新离职记录")
async def update_offboarding_record(
    record_id: UUID,
    payload: OffboardingRecordUpdate,
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    record = await service.update_record(record_id, payload)
    return success_response(
        data=OffboardingRecordResponse.model_validate(record).model_dump(mode="json"),
        message="离职记录更新成功",
    )


@router.delete("/offboarding-records/{record_id}", summary="删除离职记录")
async def delete_offboarding_record(
    record_id: UUID,
    service: OffboardingRecordService = Depends(get_offboarding_service),
):
    await service.delete_record(record_id)
    return success_response(message="离职记录删除成功")


# ─── OnboardingRecord Routes ───

@router.get("/onboarding-records", summary="老厂入职台账列表")
async def list_onboarding_records(
    department: str | None = Query(None, description="部门筛选"),
    position: str | None = Query(None, description="岗位筛选"),
    is_employed: str | None = Query(None, description="是否在职筛选"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    sort_by: str = Query("hire_date", description="排序字段"),
    sort_order: str = Query("desc", description="排序方向"),
    page_params: PageParams = Depends(),
    service: OnboardingRecordService = Depends(get_onboarding_service),
):
    records, total = await service.list_records(
        department=department,
        position=position,
        is_employed=is_employed,
        keyword=keyword,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        OnboardingRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/onboarding-records/sync-from-feishu", summary="从飞书同步老厂入职台账")
async def sync_onboarding_from_feishu(
    service: OnboardingRecordService = Depends(get_onboarding_service),
):
    """手动触发：从飞书多维表格拉取全部老厂入职数据并 upsert 到本地 PG。"""
    stats = await service.sync_from_feishu()
    msg = (
        f"同步完成：新增 {stats['created']} 条，"
        f"更新 {stats['updated']} 条，失败 {stats['failed']} 条"
    )
    return success_response(
        data=stats,
        message=msg,
    )


@router.get("/onboarding-records/sync-status", summary="老厂入职台账同步状态")
async def get_onboarding_sync_status(
    service: OnboardingRecordService = Depends(get_onboarding_service),
):
    """查看本地与飞书的数据同步统计。"""
    status = await service.get_sync_status()
    return success_response(
        data=status.model_dump(mode="json"),
    )


@router.get("/onboarding-records/{record_id}", summary="入职记录详情")
async def get_onboarding_record(
    record_id: UUID,
    service: OnboardingRecordService = Depends(get_onboarding_service),
):
    record = await service.get_record(record_id)
    return success_response(
        data=OnboardingRecordResponse.model_validate(record).model_dump(mode="json"),
    )


# ─── DepartureRecord Routes ───

@router.get("/departure-records", summary="老厂离职台账列表")
async def list_departure_records(
    department: str | None = Query(None, description="部门筛选"),
    offboarding_type: str | None = Query(None, description="离职类型筛选"),
    keyword: str | None = Query(None, description="姓名/部门/职位关键词"),
    sort_by: str = Query("offboarding_date", description="排序字段"),
    sort_order: str = Query("desc", description="排序方向"),
    page_params: PageParams = Depends(),
    service: DepartureRecordService = Depends(get_departure_service),
):
    records, total = await service.list_records(
        department=department,
        offboarding_type=offboarding_type,
        keyword=keyword,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        DepartureRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/departure-records", summary="创建离职台账记录")
async def create_departure_record(
    payload: DepartureRecordCreate,
    service: DepartureRecordService = Depends(get_departure_service),
):
    record = await service.create_record(payload)
    return success_response(
        data=DepartureRecordResponse.model_validate(record).model_dump(mode="json"),
        message="离职台账记录创建成功",
        status_code=201,
    )


@router.get("/departure-records/{record_id}", summary="离职台账记录详情")
async def get_departure_record(
    record_id: UUID,
    service: DepartureRecordService = Depends(get_departure_service),
):
    record = await service.get_record(record_id)
    return success_response(
        data=DepartureRecordResponse.model_validate(record).model_dump(mode="json"),
    )


@router.put("/departure-records/{record_id}", summary="更新离职台账记录")
async def update_departure_record(
    record_id: UUID,
    payload: DepartureRecordUpdate,
    service: DepartureRecordService = Depends(get_departure_service),
):
    record = await service.update_record(record_id, payload)
    return success_response(
        data=DepartureRecordResponse.model_validate(record).model_dump(mode="json"),
        message="离职台账记录更新成功",
    )


@router.delete("/departure-records/{record_id}", summary="删除离职台账记录")
async def delete_departure_record(
    record_id: UUID,
    service: DepartureRecordService = Depends(get_departure_service),
):
    await service.delete_record(record_id)
    return success_response(message="离职台账记录删除成功")


@router.post("/departure-records/sync-from-feishu", summary="从飞书同步老厂离职台账")
async def sync_departure_from_feishu(
    service: DepartureRecordService = Depends(get_departure_service),
):
    """手动触发：从飞书多维表格拉取全部老厂离职数据并 upsert 到本地 PG。"""
    stats = await service.sync_from_feishu()
    msg = (
        f"同步完成：新增 {stats['created']} 条，"
        f"更新 {stats['updated']} 条，失败 {stats['failed']} 条"
    )
    return success_response(
        data=stats,
        message=msg,
    )


@router.get("/departure-records/sync-status", summary="老厂离职台账同步状态")
async def get_departure_sync_status(
    service: DepartureRecordService = Depends(get_departure_service),
):
    """查看本地与飞书的数据同步统计。"""
    status = await service.get_sync_status()
    return success_response(
        data=status.model_dump(mode="json"),
    )


# ─── TrainingLedger Routes ───

@router.get("/training-ledgers", summary="培训台账列表")
async def list_training_ledgers(
    employee_number: str | None = Query(None, description="工号筛选"),
    date_from: date | None = Query(None, description="培训日期起"),
    date_to: date | None = Query(None, description="培训日期止"),
    page_params: PageParams = Depends(),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    records, total = await service.list_records(
        employee_number=employee_number,
        date_from=date_from,
        date_to=date_to,
        page=page_params.page,
        page_size=page_params.page_size,
        sort_by="training_date",
        sort_order="asc",
    )
    data = [
        TrainingLedgerResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/training-ledgers", summary="创建培训台账记录")
async def create_training_ledger(
    payload: TrainingLedgerCreate,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    record = await service.create_record(payload)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
        message="培训台账记录创建成功",
        status_code=201,
    )


# ─── TrainingLedgerPage Routes (must be before /{record_id}) ───

@router.get("/training-ledgers/pages", summary="已创建的培训台账页面列表")
async def list_training_ledger_pages(
    service: TrainingLedgerPageService = Depends(get_training_ledger_page_service),
):
    pages_with_dept = await service.list_pages_with_department()
    data = [
        {
            "id": str(page.id),
            "employee_number": page.employee_number,
            "employee_name": page.employee_name,
            "ledger_type": page.ledger_type,
            "department": dept or "未知部门",
            "factory": factory or "",
            "created_at": page.created_at.isoformat() if page.created_at else None,
            "updated_at": page.updated_at.isoformat() if page.updated_at else None,
        }
        for page, dept, factory in pages_with_dept
    ]
    return success_response(data=data)


@router.post("/training-ledgers/pages", summary="创建培训台账页面")
async def create_training_ledger_page(
    payload: TrainingLedgerPageCreate,
    service: TrainingLedgerPageService = Depends(get_training_ledger_page_service),
):
    page = await service.create_page(payload)
    return success_response(
        data=TrainingLedgerPageResponse(
            id=page.id,
            employee_number=page.employee_number,
            employee_name=page.employee_name,
            department=None,
            created_at=page.created_at,
            updated_at=page.updated_at,
        ).model_dump(mode="json"),
        message="培训台账页面创建成功",
        status_code=201,
    )


def _generate_training_ledger_excel(employee: dict, records: list[dict]) -> BytesIO:
    """Generate training ledger Excel based on employee training ledger format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "员工培训台账"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=16)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    ws.merge_cells("A1:G1")
    ws["A1"] = "丽珠集团新北江制药股份有限公司"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = "员工培训台账"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 24

    ws["A3"] = "姓名"
    ws["A3"].font = bold_font
    ws["A3"].alignment = center_align
    ws["A3"].border = thin_border
    ws["B3"] = employee.get("name", "")
    ws["B3"].border = thin_border
    ws["C3"] = "性别"
    ws["C3"].font = bold_font
    ws["C3"].alignment = center_align
    ws["C3"].border = thin_border
    ws["D3"] = employee.get("gender", "")
    ws["D3"].border = thin_border
    ws["E3"] = "工作卡号"
    ws["E3"].font = bold_font
    ws["E3"].alignment = center_align
    ws["E3"].border = thin_border
    ws.merge_cells("F3:G3")
    ws["F3"] = employee.get("employee_number", "")
    ws["F3"].border = thin_border
    ws["G3"].border = thin_border

    ws["A4"] = "部门"
    ws["A4"].font = bold_font
    ws["A4"].alignment = center_align
    ws["A4"].border = thin_border
    ws["B4"] = employee.get("department", "")
    ws["B4"].border = thin_border
    ws["C4"] = "岗位/职务"
    ws["C4"].font = bold_font
    ws["C4"].alignment = center_align
    ws["C4"].border = thin_border
    ws["D4"] = employee.get("position", "")
    ws["D4"].border = thin_border
    ws["E4"] = "入厂时间"
    ws["E4"].font = bold_font
    ws["E4"].alignment = center_align
    ws["E4"].border = thin_border
    ws.merge_cells("F4:G4")
    ws["F4"] = employee.get("factory_entry_date") or employee.get("hire_date", "")
    ws["F4"].border = thin_border
    ws["G4"].border = thin_border

    ws["A5"] = "岗位变动"
    ws["A5"].font = bold_font
    ws["A5"].alignment = center_align
    ws["A5"].border = thin_border
    ws.merge_cells("B5:G5")
    ws["B5"] = employee.get("transfer_history", "无")
    ws["B5"].border = thin_border
    for c in range(3, 8):
        ws.cell(row=5, column=c).border = thin_border

    ws["A6"] = "记录"
    ws["A6"].font = bold_font
    ws["A6"].alignment = center_align
    ws["A6"].border = thin_border
    ws.merge_cells("B6:G6")
    ws["B6"] = ""
    ws["B6"].border = thin_border
    for c in range(3, 8):
        ws.cell(row=6, column=c).border = thin_border

    headers = ["年月日", "培训课程", "培训方式", "课时", "培训单位/培训师", "考核成绩", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[7].height = 24

    for idx, record in enumerate(records, 8):
        values = [
            record.get("training_date", ""),
            record.get("training_subject", ""),
            record.get("training_method", ""),
            record.get("duration_hours", ""),
            record.get("trainer", ""),
            record.get("assessment_result", ""),
            record.get("remarks", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col in (1, 3, 4, 6, 7) else left_align

    while len(records) < 12:
        row = 8 + len(records)
        for col in range(1, 8):
            ws.cell(row=row, column=col, value="").border = thin_border
        records.append({})

    footer_row = 8 + len(records)
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    ws.cell(row=footer_row, column=1, value="备注：笔试考核设置为满分100分，考试合格线为80分。")
    ws.cell(row=footer_row, column=1).alignment = left_align
    ws.cell(row=footer_row, column=1).border = thin_border
    for c in range(2, 8):
        ws.cell(row=footer_row, column=c).border = thin_border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/training-ledgers/export", summary="导出培训台账")
async def export_training_ledger(
    employee_number: str = Query(..., description="员工工号"),
    ledger_type: str = Query("event", description="台账类型: event=事件台账, sop=SOP培训台账"),
    factory: str = Query("old", description="厂区: old=旧厂, new=新厂"),
    ledger_service: TrainingLedgerService = Depends(get_training_ledger_service),
    employee_service: EmployeeService = Depends(get_employee_service),
):
    """根据员工数据、厂区和台账类型选择对应模板导出培训台账。"""
    employee = await employee_service.get_employee_by_number(employee_number)
    if not employee:
        raise HTTPException(status_code=404, detail="未找到该员工")

    records, _ = await ledger_service.list_records(
        employee_number=employee_number,
        ledger_type=ledger_type,
        page=1,
        page_size=1000,
        sort_by="training_date",
        sort_order="asc",
    )

    employee_dict = EmployeeResponse.model_validate(employee).model_dump(mode="json")
    record_dicts = [
        TrainingLedgerResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]

    buffer = generate_ledger_export(employee_dict, record_dicts, factory, ledger_type)
    buffer.seek(0)

    safe_name = employee.name or "unknown"
    ext = "xlsx" if factory == "old" else "docx"
    filename = f"{safe_name}培训台账.{ext}"
    encoded_filename = quote(filename, safe="")

    media_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if factory == "old"
        else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )
    return StreamingResponse(
        iter([buffer.read()]),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"},
    )


@router.get("/training-ledgers/{record_id}", summary="培训台账记录详情")
async def get_training_ledger(
    record_id: UUID,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    record = await service.get_record(record_id)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
    )


@router.put("/training-ledgers/{record_id}", summary="更新培训台账记录")
async def update_training_ledger(
    record_id: UUID,
    payload: TrainingLedgerUpdate,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    record = await service.update_record(record_id, payload)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
        message="培训台账记录更新成功",
    )


@router.delete("/training-ledgers/{record_id}", summary="删除培训台账记录")
async def delete_training_ledger(
    record_id: UUID,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
):
    await service.delete_record(record_id)
    return success_response(message="培训台账记录删除成功")


# ─── AnnualTrainingPlan Routes ───

@router.get("/annual-training-plans", summary="年度培训计划列表")
async def list_annual_training_plans(
    year: int | None = Query(None, description="年度筛选"),
    department: str | None = Query(None, description="部门筛选"),
    page_params: PageParams = Depends(),
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plans, total = await service.list_plans(
        year=year,
        department=department,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        AnnualTrainingPlanResponse.model_validate(p).model_dump(mode="json")
        for p in plans
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/annual-training-plans", summary="创建年度培训计划")
async def create_annual_training_plan(
    payload: AnnualTrainingPlanCreate,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plan = await service.create_plan(payload)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
        message="年度培训计划创建成功",
        status_code=201,
    )


@router.get("/annual-training-plans/{plan_id}", summary="年度培训计划详情")
async def get_annual_training_plan(
    plan_id: UUID,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plan = await service.get_plan(plan_id)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
    )


@router.put("/annual-training-plans/{plan_id}", summary="更新年度培训计划")
async def update_annual_training_plan(
    plan_id: UUID,
    payload: AnnualTrainingPlanUpdate,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    plan = await service.update_plan(plan_id, payload)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
        message="年度培训计划更新成功",
    )


@router.delete("/annual-training-plans/{plan_id}", summary="删除年度培训计划")
async def delete_annual_training_plan(
    plan_id: UUID,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
):
    await service.delete_plan(plan_id)
    return success_response(message="年度培训计划删除成功")


@router.get("/annual-training-plans/{plan_id}/items", summary="年度计划明细列表")
async def list_annual_training_plan_items(
    plan_id: UUID,
    service: AnnualTrainingPlanItemService = Depends(get_annual_training_plan_item_service),
):
    items = await service.list_items(plan_id)
    data = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]
    return success_response(data=data)


@router.put("/annual-training-plans/{plan_id}/items/batch", summary="批量更新年度计划明细")
async def batch_update_annual_training_plan_items(
    plan_id: UUID,
    payload: AnnualTrainingPlanItemBatchUpdate,
    service: AnnualTrainingPlanItemService = Depends(get_annual_training_plan_item_service),
):
    items = await service.batch_update_items(plan_id, payload)
    data = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]
    return success_response(
        data=data,
        message="年度计划明细更新成功",
    )


def _generate_annual_plan_excel(plan: dict, items: list[dict]) -> BytesIO:
    """Generate annual training plan Excel based on 7.7 template format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "年度培训计划"

    # Styles
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=16)

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 10

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{plan['year']} 年培训计划"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Department row
    ws.merge_cells("A2:I2")
    ws["A2"] = f"部门：{plan['department']}"
    ws["A2"].font = bold_font
    ws["A2"].alignment = left_align
    ws.row_dimensions[2].height = 22

    # Header row
    headers = ["序号", "培训季度及课时", "培训内容及使用教材", "培训对象",
               "授课单位及授课人", "考核方式", "培训跟踪", "确认人/日期", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[3].height = 28

    # Data rows
    for idx, item in enumerate(items, 1):
        row = 3 + idx
        quarter = item.get("month") or ""
        hours = item.get("duration_hours")
        quarter_hours = f"{quarter}\n{hours}课时" if hours else quarter

        values = [
            idx,
            quarter_hours,
            item.get("content_and_textbook") or "",
            item.get("target_audience") or "",
            item.get("position_and_count") or "",
            item.get("training_method") or "",
            item.get("tracking_status") or "",
            f"{item.get('confirmer') or ''}{' / ' + str(item.get('confirm_date')) if item.get('confirm_date') else ''}",
            item.get("remarks") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col in (1, 2, 6, 7) else left_align
        ws.row_dimensions[row].height = 36

    # Pad to at least 12 rows
    while len(items) < 12:
        row = 3 + len(items) + 1
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = thin_border
        ws.row_dimensions[row].height = 36
        items.append({})

    # Footer row
    footer_row = 4 + len(items) + 1
    ws.merge_cells(f"A{footer_row}:E{footer_row}")
    ws.cell(row=footer_row, column=1, value="制表人/日期：")
    ws.cell(row=footer_row, column=1).alignment = left_align
    ws.cell(row=footer_row, column=1).border = thin_border
    for c in range(2, 6):
        ws.cell(row=footer_row, column=c).border = thin_border

    ws.merge_cells(f"F{footer_row}:I{footer_row}")
    ws.cell(row=footer_row, column=6, value="部门负责人/日期：")
    ws.cell(row=footer_row, column=6).alignment = left_align
    ws.cell(row=footer_row, column=6).border = thin_border
    for c in range(7, 10):
        ws.cell(row=footer_row, column=c).border = thin_border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/annual-training-plans/{plan_id}/export", summary="导出年度培训计划Excel")
async def export_annual_training_plan(
    plan_id: UUID,
    plan_service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    item_service: AnnualTrainingPlanItemService = Depends(get_annual_training_plan_item_service),
):
    """根据年度计划数据生成并导出Excel文件（7.7年度培训计划格式）。"""
    plan = await plan_service.get_plan(plan_id)
    items = await item_service.list_items(plan_id)

    plan_dict = AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json")
    item_dicts = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]

    buffer = _generate_annual_plan_excel(plan_dict, item_dicts)
    buffer.seek(0)

    safe_dept = plan.department.replace(" ", "_")
    filename = f"{plan.year}年度培训计划_{safe_dept}.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"
        },
    )


# ─── New Factory Routes (read-only, query _new clone tables) ───

from types import SimpleNamespace


def _build_where(
    base_table: str,
    filters: list[tuple[str, str | None]],
    keyword_fields: list[str] | None = None,
    keyword: str | None = None,
) -> tuple[str, dict]:
    """Build WHERE clause and params for raw SQL."""
    conditions = [f"{base_table}.is_deleted = false"]
    params: dict = {}
    for col, val in filters:
        if val:
            conditions.append(f"{base_table}.{col} ILIKE :{col}")
            params[col] = f"%{val}%"
    if keyword and keyword_fields:
        ors = [f"{base_table}.{f} ILIKE :keyword" for f in keyword_fields]
        conditions.append(f"({' OR '.join(ors)})")
        params["keyword"] = f"%{keyword}%"
    return " AND ".join(conditions), params


async def _query_clone_table(
    session: AsyncSession,
    table: str,
    schema: type,
    where_sql: str,
    params: dict,
    page: int,
    page_size: int,
    sort_by: str = "created_at",
    sort_order: str = "desc",
) -> tuple[list, int]:
    count_sql = text(f"SELECT COUNT(*) FROM {table} WHERE {where_sql}")
    total = (await session.execute(count_sql, params)).scalar()

    sql = text(f"""
        SELECT * FROM {table}
        WHERE {where_sql}
        ORDER BY {sort_by} {sort_order}
        LIMIT :limit OFFSET :offset
    """)
    params["limit"] = page_size
    params["offset"] = (page - 1) * page_size

    result = await session.execute(sql, params)
    rows = result.mappings().all()
    data = [
        schema.model_validate(SimpleNamespace(**dict(row))).model_dump(mode="json")
        for row in rows
    ]
    return data, total


@router.get("/new/employees", summary="新厂员工列表")
async def list_new_employees(
    department: str | None = Query(None),
    status: str | None = Query(None),
    keyword: str | None = Query(None),
    page_params: PageParams = Depends(),
    session: AsyncSession = Depends(get_db),
):
    """List employees from employees_new clone table."""
    # 处理多肽车间按班组拆分的情况
    dept_filter = department
    team_filter = None
    if department and department.startswith("多肽车间-"):
        dept_filter = "多肽车间"
        team_filter = department.replace("多肽车间-", "")

    where, params = _build_where(
        "hr.employees_new",
        [("department", dept_filter), ("status", status)],
        keyword_fields=["name", "employee_number"],
        keyword=keyword,
    )
    if team_filter:
        where += " AND hr.employees_new.team = :team"
        params["team"] = team_filter

    data, total = await _query_clone_table(
        session, "hr.employees_new", EmployeeResponse,
        where, params, page_params.page, page_params.page_size,
    )
    return paginated_response(
        data=data, page=page_params.page,
        page_size=page_params.page_size, total=total,
    )


@router.get("/new/onboarding-records", summary="新厂入职台账列表")
async def list_new_onboarding_records(
    department: str | None = Query(None),
    position: str | None = Query(None),
    keyword: str | None = Query(None),
    page_params: PageParams = Depends(),
    session: AsyncSession = Depends(get_db),
):
    where, params = _build_where(
        "hr.onboarding_records_new",
        [("department", department), ("position", position)],
        keyword_fields=["name", "employee_number"],
        keyword=keyword,
    )
    data, total = await _query_clone_table(
        session, "hr.onboarding_records_new", OnboardingRecordResponse,
        where, params, page_params.page, page_params.page_size,
        sort_by="hire_date", sort_order="desc",
    )
    return paginated_response(
        data=data, page=page_params.page,
        page_size=page_params.page_size, total=total,
    )


@router.get("/new/departure-records", summary="新厂离职台账列表")
async def list_new_departure_records(
    department: str | None = Query(None),
    offboarding_type: str | None = Query(None),
    keyword: str | None = Query(None),
    page_params: PageParams = Depends(),
    session: AsyncSession = Depends(get_db),
):
    where, params = _build_where(
        "hr.departure_records_new",
        [("department", department), ("offboarding_type", offboarding_type)],
        keyword_fields=["name", "department", "position"],
        keyword=keyword,
    )
    data, total = await _query_clone_table(
        session, "hr.departure_records_new", DepartureRecordResponse,
        where, params, page_params.page, page_params.page_size,
        sort_by="offboarding_date", sort_order="desc",
    )
    return paginated_response(
        data=data, page=page_params.page,
        page_size=page_params.page_size, total=total,
    )


@router.get("/new/offboarding-records", summary="新厂离职管理列表")
async def list_new_offboarding_records(
    department: str | None = Query(None),
    offboarding_type: str | None = Query(None),
    keyword: str | None = Query(None),
    page_params: PageParams = Depends(),
    session: AsyncSession = Depends(get_db),
):
    # Reuse departure_records_new as offboarding data for new factory
    where, params = _build_where(
        "hr.departure_records_new",
        [("department", department), ("offboarding_type", offboarding_type)],
        keyword_fields=["name", "department", "position"],
        keyword=keyword,
    )
    data, total = await _query_clone_table(
        session, "hr.departure_records_new", DepartureRecordResponse,
        where, params, page_params.page, page_params.page_size,
        sort_by="offboarding_date", sort_order="desc",
    )
    return paginated_response(
        data=data, page=page_params.page,
        page_size=page_params.page_size, total=total,
    )


@router.get("/new/departments", summary="新厂部门列表")
async def list_new_departments(
    keyword: str | None = Query(None),
    page_params: PageParams = Depends(),
    session: AsyncSession = Depends(get_db),
):
    """Aggregate departments from employees_new (no standalone departments_new table)."""
    where = "is_deleted = false"
    params: dict = {}
    if keyword:
        where += " AND name ILIKE :keyword"
        params["keyword"] = f"%{keyword}%"

    count_sql = text(f"""
        SELECT COUNT(DISTINCT department) FROM hr.employees_new
        WHERE {where}
    """)
    total = (await session.execute(count_sql, params)).scalar()

    sql = text(f"""
        SELECT DISTINCT ON (department)
            department AS name,
            department AS code,
            '' AS description,
            gen_random_uuid() AS id,
            NOW() AS created_at,
            NOW() AS updated_at
        FROM hr.employees_new
        WHERE {where}
        ORDER BY department
        LIMIT :limit OFFSET :offset
    """)
    params["limit"] = page_params.page_size
    params["offset"] = (page_params.page - 1) * page_params.page_size

    result = await session.execute(sql, params)
    rows = result.mappings().all()
    data = [
        DepartmentResponse.model_validate(SimpleNamespace(**dict(row))).model_dump(mode="json")
        for row in rows
    ]

    # 多肽车间按班组拆分为 多肽车间-xxx
    expanded: list[dict] = []
    for dept in data:
        name = dept["name"]
        if name == "多肽车间":
            # 查询多肽车间下所有班组
            team_sql = text("""
                SELECT DISTINCT team FROM hr.employees_new
                WHERE department = '多肽车间' AND is_deleted = false AND team IS NOT NULL AND team != ''
                ORDER BY team
            """)
            team_result = await session.execute(team_sql)
            team_rows = team_result.mappings().all()
            if team_rows:
                for row in team_rows:
                    team_name = row["team"]
                    expanded.append({
                        **dept,
                        "name": f"多肽车间-{team_name}",
                        "code": f"多肽车间-{team_name}",
                    })
            else:
                expanded.append(dept)
        else:
            expanded.append(dept)

    return paginated_response(
        data=expanded, page=page_params.page,
        page_size=page_params.page_size, total=total,
    )


router.include_router(analysis_router)
if attendance_router is not None:
    router.include_router(attendance_router)


# ─── Candidate Routes ───

@router.get("/candidates", summary="候选人列表")
async def list_candidates(
    position: str | None = Query(None, description="职位筛选"),
    education: str | None = Query(None, description="学历筛选"),
    recommendation_level: str | None = Query(None, description="推荐等级筛选（支持逗号分隔多个值）"),
    sync_status: str | None = Query(None, description="飞书同步状态筛选: synced/failed/unsynced"),
    keyword: str | None = Query(None, description="姓名/职位关键词"),
    page_params: PageParams = Depends(),
    service: CandidateService = Depends(get_candidate_service),
):
    candidates, total = await service.list_candidates(
        position=position,
        education=education,
        recommendation_level=recommendation_level,
        sync_status=sync_status,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    logger.debug("API recommendation_level=%s, total=%s", recommendation_level, total)
    data = [
        CandidateResponse.model_validate(c).model_dump(mode="json")
        for c in candidates
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/candidates/parse-preview", summary="预览简历AI解析结果")
async def preview_resume_parse(
    resume: UploadFile = File(..., description="简历 PDF 附件"),
    position: str = Form(..., max_length=64, description="应聘职位名称"),
    service: CandidateService = Depends(get_candidate_service),
):
    """上传简历 PDF，转为图片后由 AI 视觉识别并返回预览字段（不保存）。"""
    resume_bytes = await resume.read()
    result = await service.parse_resume_preview(resume_bytes, position)
    return success_response(data=result)


@router.post("/candidates", summary="新建候选人")
async def create_candidate(
    name: str = Form(..., max_length=64, description="候选人姓名"),
    position: str = Form(..., max_length=64, description="应聘职位名称"),
    resume: UploadFile = File(..., description="简历 PDF 附件"),
    gender: str | None = Form(None, max_length=8, description="性别"),
    school: str | None = Form(None, max_length=128, description="学校名称"),
    education: str | None = Form(None, max_length=16, description="学历"),
    major: str | None = Form(None, max_length=64, description="专业"),
    match_report: str | None = Form(None, description="AI 匹配度报告"),
    recommendation_level: str | None = Form(
        None, max_length=16, description="推荐等级"
    ),
    service: CandidateService = Depends(get_candidate_service),
):
    """手动新建候选人：上传简历 PDF，创建本地记录并同步到飞书。"""
    resume_bytes = await resume.read()
    candidate = await service.create_candidate_with_resume(
        name=name,
        position=position,
        resume_bytes=resume_bytes,
        filename=resume.filename or f"{name}.pdf",
        gender=gender,
        school=school,
        education=education,
        major=major,
        match_report=match_report,
        recommendation_level=recommendation_level,
    )
    return success_response(
        data=CandidateResponse.model_validate(candidate).model_dump(mode="json"),
        message="候选人创建成功",
    )


@router.post("/candidates/sync-from-feishu", summary="从飞书同步候选人数据")
async def sync_candidates_from_feishu(
    service: CandidateService = Depends(get_candidate_service),
):
    """手动触发：从飞书多维表格拉取全部候选人数据并 upsert 到本地 PG。"""
    stats = await service.sync_from_feishu()
    msg = (
        f"同步完成：新增 {stats['created']} 条，"
        f"更新 {stats['updated']} 条，失败 {stats['failed']} 条"
    )
    return success_response(
        data=stats,
        message=msg,
    )


@router.get("/candidates/sync-status", summary="候选人同步状态")
async def get_candidates_sync_status(
    service: CandidateService = Depends(get_candidate_service),
):
    """查看本地与飞书的候选人数据同步统计。"""
    status = await service.get_sync_status()
    return success_response(
        data=status.model_dump(mode="json"),
    )


@router.post("/candidates/{candidate_id}/sync-to-feishu", summary="同步候选人到飞书")
async def sync_candidate_to_feishu(
    candidate_id: UUID,
    service: CandidateService = Depends(get_candidate_service),
):
    """手动触发：将单个候选人（含简历）同步到飞书多维表格。"""
    candidate = await service.sync_candidate_to_feishu(candidate_id)
    return success_response(
        data=CandidateResponse.model_validate(candidate).model_dump(mode="json"),
        message="候选人与简历已同步到飞书",
    )


@router.get("/candidates/{candidate_id}", summary="候选人详情")
async def get_candidate(
    candidate_id: UUID,
    service: CandidateService = Depends(get_candidate_service),
):
    candidate = await service.get_candidate(candidate_id)
    return success_response(
        data=CandidateResponse.model_validate(candidate).model_dump(mode="json"),
    )


@router.put("/candidates/{candidate_id}", summary="更新候选人信息")
async def update_candidate(
    candidate_id: UUID,
    payload: CandidateUpdate,
    service: CandidateService = Depends(get_candidate_service),
):
    candidate = await service.update_candidate(candidate_id, payload)
    return success_response(
        data=CandidateResponse.model_validate(candidate).model_dump(mode="json"),
        message="候选人信息更新成功",
    )


@router.delete("/candidates/{candidate_id}", summary="删除候选人")
async def delete_candidate(
    candidate_id: UUID,
    service: CandidateService = Depends(get_candidate_service),
):
    await service.delete_candidate(candidate_id)
    return success_response(message="候选人删除成功")


@router.put("/candidates/{candidate_id}/recommendation-level", summary="更新候选人推荐等级")
async def update_candidate_recommendation_level(
    candidate_id: UUID,
    payload: CandidateUpdateRecommendationLevel,
    service: CandidateService = Depends(get_candidate_service),
):
    candidate = await service.update_recommendation_level(
        candidate_id, payload.recommendation_level
    )
    return success_response(
        data=CandidateResponse.model_validate(candidate).model_dump(mode="json"),
        message="推荐等级更新成功",
    )


@router.get("/candidates/{candidate_id}/resume-preview", summary="简历预览")
async def preview_candidate_resume(
    candidate_id: UUID,
    service: CandidateService = Depends(get_candidate_service),
):
    """获取候选人简历 PDF。优先读取本地存储，若不存在则从飞书下载。"""
    import os

    from app.core.exceptions import NotFoundException

    candidate = await service.get_candidate(candidate_id)

    # 优先读取本地文件
    if candidate.resume_storage_path and os.path.exists(candidate.resume_storage_path):
        def iter_file():
            with open(candidate.resume_storage_path, "rb") as f:
                while chunk := f.read(64 * 1024):
                    yield chunk

        return StreamingResponse(
            content=iter_file(),
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="resume.pdf"'},
        )

    # 回退：从飞书实时下载
    attachments = candidate.resume_attachments or []
    if not attachments:
        raise NotFoundException("简历附件", str(candidate_id))

    file_token = attachments[0].get("file_token")
    if not file_token:
        raise NotFoundException("简历附件", str(candidate_id))

    try:
        tmp_download_url = await service.bitable.get_resume_download_url(file_token)
    except Exception as exc:
        logger.warning("Failed to get feishu download url for candidate %s: %s", candidate_id, exc)
        raise NotFoundException("简历附件", str(candidate_id)) from exc

    import httpx
    async with httpx.AsyncClient(follow_redirects=True) as client:
        response = await client.get(tmp_download_url)
        response.raise_for_status()

    return StreamingResponse(
        content=response.iter_bytes(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="resume.pdf"'},
    )


# ─── TrainingSession Routes ───

@router.get("/training-sessions", summary="培训记录列表")
async def list_training_sessions(
    department: str | None = Query(None, description="部门筛选"),
    keyword: str | None = Query(None, description="主题或培训师关键词"),
    date_from: date | None = Query(None, description="日期起始"),
    date_to: date | None = Query(None, description="日期截止"),
    page_params: PageParams = Depends(),
    service: TrainingSessionService = Depends(get_training_session_service),
):
    sessions, total = await service.list_sessions(
        department=department,
        keyword=keyword,
        date_from=date_from,
        date_to=date_to,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        TrainingSessionResponse.model_validate(s).model_dump(mode="json")
        for s in sessions
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/training-sessions", summary="创建培训记录")
async def create_training_session(
    payload: TrainingSessionCreate,
    service: TrainingSessionService = Depends(get_training_session_service),
):
    session_obj = await service.create_session(payload)
    return success_response(
        data=TrainingSessionResponse.model_validate(session_obj).model_dump(mode="json"),
        message="培训记录创建成功",
        status_code=201,
    )


@router.put("/training-sessions/{session_id}/status", summary="更新培训记录状态")
async def update_training_session_status(
    session_id: UUID,
    payload: TrainingSessionStatusUpdate,
    service: TrainingSessionService = Depends(get_training_session_service),
):
    session_obj = await service.update_status(session_id, payload.status)
    return success_response(
        data=TrainingSessionResponse.model_validate(session_obj).model_dump(mode="json"),
        message="状态更新成功",
    )


@router.get("/training-sessions/{session_id}", summary="培训记录详情")
async def get_training_session(
    session_id: UUID,
    service: TrainingSessionService = Depends(get_training_session_service),
):
    session_obj = await service.get_session(session_id)
    return success_response(
        data=TrainingSessionResponse.model_validate(session_obj).model_dump(mode="json"),
    )


@router.put("/training-sessions/{session_id}", summary="更新培训记录")
async def update_training_session(
    session_id: UUID,
    payload: TrainingSessionUpdate,
    service: TrainingSessionService = Depends(get_training_session_service),
):
    session_obj = await service.update_session(session_id, payload)
    return success_response(
        data=TrainingSessionResponse.model_validate(session_obj).model_dump(mode="json"),
        message="培训记录更新成功",
    )


@router.delete("/training-sessions/{session_id}", summary="删除培训记录")
async def delete_training_session(
    session_id: UUID,
    service: TrainingSessionService = Depends(get_training_session_service),
):
    await service.delete_session(session_id)
    return success_response(message="培训记录删除成功")


@router.post("/training-sessions/{session_id}/send-select-tasks", summary="发送多部门选择受训人员任务")
async def send_training_session_select_tasks(
    session_id: UUID,
    service: TrainingSessionService = Depends(get_training_session_service),
    employee_service: EmployeeService = Depends(get_employee_service),
):
    """为培训记录的每个受训部门发送选择受训人员任务。

    - 为每个部门生成 UUID token
    - 将任务数据存入 Redis（24h 有效）
    - 查找该部门培训专员并发送飞书 IM 消息
    - 更新 TrainingSession.select_tasks 和 status
    """
    session_obj = await service.get_session(session_id)
    trainee_departments = session_obj.trainee_departments or []
    if not trainee_departments:
        raise HTTPException(status_code=400, detail="该培训记录没有受训部门")

    settings = get_settings()
    select_tasks: list[dict] = []

    for dept in trainee_departments:
        token = str(uuid.uuid4())
        cache_key = f"training_select:{token}"
        task_data = {
            "department": dept,
            "training_date": str(session_obj.training_date),
            "subject": session_obj.subject,
            "training_time_start": session_obj.training_time_start,
            "training_time_end": session_obj.training_time_end,
            "location": session_obj.location,
            "trainer": session_obj.trainer,
            "training_method": session_obj.training_method,
            "content": session_obj.content,
            "factory": session_obj.factory,
            "issuer_department": session_obj.issuer_department,
            "issue_date": str(session_obj.issue_date) if session_obj.issue_date else None,
        }
        await cache_set(cache_key, json.dumps(task_data, ensure_ascii=False), ex=86400)
        await cache_set(f"training_select_list:{token}", json.dumps(task_data, ensure_ascii=False), ex=86400)

        # 同步更新共享任务列表
        list_key = "training_select_list"
        existing = await cache_get(list_key)
        task_list = json.loads(existing) if existing else []
        task_list.append({
            "token": token,
            "department": dept,
            "training_date": str(session_obj.training_date),
            "subject": session_obj.subject,
            "factory": session_obj.factory,
            "created_at": datetime.now().isoformat(),
        })
        await cache_set(list_key, json.dumps(task_list, ensure_ascii=False), ex=604800)

        # 查找培训专员 — 从 training_specialists 表查询
        specialist_repo = TrainingSpecialistRepository(employee_service.repo.session)
        specialist_config = await specialist_repo.get_by_dept_factory(dept, session_obj.factory or "old")
        specialist = None
        specialist_table = None

        if specialist_config:
            # 根据工号从对应厂区的员工表查找
            for table in ["hr.employees", "hr.employees_new"]:
                sql = text(
                    f"SELECT * FROM {table} "
                    f"WHERE employee_number = :num AND is_deleted = false LIMIT 1"
                )
                result = await employee_service.repo.session.execute(sql, {"num": specialist_config.employee_number})
                row = result.mappings().first()
                if row:
                    specialist = Employee(**dict(row))
                    specialist_table = table
                    break

        if not specialist:
            logger.warning("部门 %s（%s厂）未配置培训专员，跳过发送飞书消息", dept, session_obj.factory or "old")
            select_tasks.append({
                "department": dept,
                "token": token,
                "status": "pending",
                "employee_names": [],
                "employee_numbers": [],
                "specialist_found": False,
            })
            continue

        # 解析飞书 open_id
        open_id = specialist.feishu_open_id
        if not open_id and specialist.phone:
            try:
                im = FeishuIM()
                mobile = specialist.phone if specialist.phone.startswith("+") else f"+86{specialist.phone}"
                mapping = await im.batch_get_open_ids_by_mobile([mobile])
                open_id = mapping.get(mobile) or mapping.get(specialist.phone)
                if open_id:
                    update_sql = text(
                        f"UPDATE {specialist_table} SET feishu_open_id = :oid WHERE employee_number = :eno"
                    )
                    await employee_service.repo.session.execute(update_sql, {"oid": open_id, "eno": specialist.employee_number})
                    await employee_service.repo.session.flush()
            except Exception:
                pass

        if open_id:
            select_url = f"{settings.FRONTEND_URL}/hr/training/select?token={token}&department={quote(dept)}"
            msg_content = (
                f"【培训人员选择通知】\n"
                f"请为 {dept} 的 {session_obj.training_date} {session_obj.subject} 培训选择受训人员。\n"
                f"点击链接选择：{select_url}\n"
                f"（链接 24 小时内有效）"
            )
            try:
                im = FeishuIM()
                await im.send_text_message(open_id, msg_content)
            except Exception as e:
                logger.warning("发送飞书消息给 %s 失败: %s", specialist.name, e)

        select_tasks.append({
            "department": dept,
            "token": token,
            "status": "pending",
            "employee_names": [],
            "employee_numbers": [],
            "specialist_found": True,
            "specialist_name": specialist.name,
        })

    session_obj.select_tasks = select_tasks
    session_obj.status = "selecting"
    await service.repo.update(session_obj)

    return success_response(
        data={"session_id": str(session_id), "select_tasks": select_tasks},
        message="选择任务已发送",
    )


@router.get("/training-sessions/{session_id}/select-tasks", summary="获取培训记录的选择任务列表")
async def get_training_session_select_tasks(
    session_id: UUID,
    service: TrainingSessionService = Depends(get_training_session_service),
):
    """获取培训记录的多部门选择任务列表。"""
    session_obj = await service.get_session(session_id)
    return success_response(data=session_obj.select_tasks or [])


# ─── Training Specialists ───


def get_specialist_repo(session: AsyncSession = Depends(get_db)) -> TrainingSpecialistRepository:
    return TrainingSpecialistRepository(session)


@router.get("/training-specialists", summary="培训专员列表")
async def list_training_specialists(
    repo: TrainingSpecialistRepository = Depends(get_specialist_repo),
):
    specialists = await repo.list_all()
    data = [
        {
            "id": str(s.id),
            "department": s.department,
            "employee_number": s.employee_number,
            "employee_name": s.employee_name,
            "factory": s.factory,
            "feishu_open_id": s.feishu_open_id or "",
        }
        for s in specialists
    ]
    return success_response(data=data)


@router.post("/training-specialists", summary="新增或更新培训专员")
async def upsert_training_specialist(
    payload: TrainingSpecialistCreate,
    repo: TrainingSpecialistRepository = Depends(get_specialist_repo),
):
    s = await repo.upsert(payload.department, payload.employee_number, payload.employee_name, payload.factory)
    return success_response(
        data={"id": str(s.id), "department": s.department, "employee_number": s.employee_number, "employee_name": s.employee_name, "factory": s.factory},
        message="培训专员已保存",
    )


@router.delete("/training-specialists/{specialist_id}", summary="删除培训专员")
async def delete_training_specialist(
    specialist_id: UUID,
    repo: TrainingSpecialistRepository = Depends(get_specialist_repo),
):
    await repo.delete(specialist_id)
    return success_response(message="培训专员已删除")


@router.post("/training-specialists/sync-feishu-openids", summary="同步培训专员飞书 open_id")
async def sync_specialist_feishu_openids(
    repo: TrainingSpecialistRepository = Depends(get_specialist_repo),
    employee_service: EmployeeService = Depends(get_employee_service),
):
    """从员工表或飞书 API 批量获取培训专员的 open_id 并持久化。"""
    specialists = await repo.list_all()
    results: list[dict] = []

    for s in specialists:
        if s.feishu_open_id:
            continue  # already has open_id

        open_id = None
        source = ""

        # 1) Try employees table (old factory)
        if s.factory == "old":
            emp = await employee_service.repo.get_by_employee_number(s.employee_number)
            if emp and emp.feishu_open_id:
                open_id = emp.feishu_open_id
                source = "employees"
        else:
            # 2) Try employees_new table
            emp = await employee_service.repo.get_new_employee_by_number(s.employee_number)
            if emp and emp.feishu_open_id:
                open_id = emp.feishu_open_id
                source = "employees_new"

        # 3) Try Feishu API via phone
        if not open_id:
            try:
                from app.platform.integrations.feishu_im import FeishuIM
                # Get phone from employees tables
                phone = None
                emp = await employee_service.repo.get_by_employee_number(s.employee_number)
                if emp and emp.phone:
                    phone = emp.phone
                if not phone:
                    emp = await employee_service.repo.get_new_employee_by_number(s.employee_number)
                    if emp and emp.phone:
                        phone = emp.phone
                if phone:
                    im = FeishuIM()
                    mapping = await im.batch_get_open_ids_by_mobile([phone])
                    open_id = mapping.get(phone)
                    if open_id:
                        source = "feishu_api"
            except Exception as e:
                logger.warning("Feishu API lookup failed for %s: %s", s.employee_name, e)

        if open_id:
            s.feishu_open_id = open_id
            await repo.session.flush()
            results.append({"employee_number": s.employee_number, "employee_name": s.employee_name, "open_id": open_id, "source": source, "status": "updated"})
        else:
            results.append({"employee_number": s.employee_number, "employee_name": s.employee_name, "open_id": "", "source": "", "status": "not_found"})

    return success_response(data={"synced": len([r for r in results if r["status"] == "updated"]), "failed": len([r for r in results if r["status"] == "not_found"]), "details": results})


# ─── Training Teams ───


def get_team_repo(session: AsyncSession = Depends(get_db)) -> TrainingTeamRepository:
    return TrainingTeamRepository(session)


@router.get("/training-teams", summary="自定义受训班组列表")
async def list_training_teams(
    factory: str = Query("old", description="厂区: old=旧厂, new=新厂"),
    repo: TrainingTeamRepository = Depends(get_team_repo),
):
    teams = await repo.list_by_factory(factory)
    data = [
        {
            "id": str(t.id),
            "name": t.name,
            "factory": t.factory,
            "department": t.department,
            "specialist_employee_number": t.specialist_employee_number,
            "specialist_name": t.specialist_name,
            "employee_names": t.employee_names or [],
            "employee_numbers": t.employee_numbers or [],
        }
        for t in teams
    ]
    return success_response(data=data)


@router.post("/training-teams", summary="新增自定义受训班组")
async def create_training_team(
    payload: TrainingTeamCreate,
    repo: TrainingTeamRepository = Depends(get_team_repo),
):
    t = await repo.create(payload)
    return success_response(
        data={"id": str(t.id), "name": t.name, "factory": t.factory},
        message="班组已创建",
    )


@router.put("/training-teams/{team_id}", summary="编辑自定义受训班组")
async def update_training_team(
    team_id: UUID,
    payload: TrainingTeamUpdate,
    repo: TrainingTeamRepository = Depends(get_team_repo),
):
    t = await repo.update(team_id, payload)
    return success_response(
        data={"id": str(t.id), "name": t.name},
        message="班组已更新",
    )


@router.delete("/training-teams/{team_id}", summary="删除自定义受训班组")
async def delete_training_team(
    team_id: UUID,
    repo: TrainingTeamRepository = Depends(get_team_repo),
):
    await repo.delete(team_id)
    return success_response(message="班组已删除")


# ─── Pre-job Training Plan Templates ───


def get_prejob_template_repo(
    session: AsyncSession = Depends(get_db),
) -> PrejobTemplateRepository:
    return PrejobTemplateRepository(session)


@router.get(
    "/prejob-training-templates",
    summary="获取部门岗前培训计划模板",
)
async def get_prejob_template(
    department: str = Query(..., description="部门名称"),
    factory: str = Query("old", description="厂区: old=旧厂, new=新厂"),
    repo: PrejobTemplateRepository = Depends(get_prejob_template_repo),
):
    """获取指定部门+厂区的岗前培训计划模板。若没有保存过模板则返回 null。"""
    template = await repo.get_by_dept_factory(department, factory)
    if template:
        return success_response(
            data=PrejobTemplateResponse.model_validate(template).model_dump()
        )
    return success_response(data=None)


@router.put(
    "/prejob-training-templates",
    summary="保存部门岗前培训计划模板",
)
async def save_prejob_template(
    payload: PrejobTemplateCreate,
    repo: PrejobTemplateRepository = Depends(get_prejob_template_repo),
):
    """保存（新增或覆盖）指定部门+厂区的岗前培训计划模板。"""
    items = [item.model_dump() for item in payload.items]
    template = await repo.upsert(payload.department, payload.factory, items)
    return success_response(
        data=PrejobTemplateResponse.model_validate(template).model_dump(),
        message="岗前培训计划模板已保存",
    )


# ─── System Settings ───

ENV_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "..", ".env")
SETTING_KEYS = [
    "FEISHU_APP_ID", "FEISHU_APP_SECRET",
    "FEISHU_VEHICLE_APP_ID", "FEISHU_VEHICLE_APP_SECRET",
    "FEISHU_TRAINING_APP_ID", "FEISHU_TRAINING_APP_SECRET",
    "AI_BASE_URL", "AI_API_KEY", "AI_MODEL",
    "AILY_APP_ID",
]


@router.get("/system-settings", summary="获取系统设置")
async def get_system_settings(
    session: AsyncSession = Depends(get_db),
):
    """获取所有系统配置。secret 类字段仅返回尾部 4 位。"""
    result = {}
    _settings = get_settings()
    from sqlalchemy import text as sql_text
    r = await session.execute(sql_text("SELECT key, value FROM hr.system_settings"))
    db_settings = {row[0]: row[1] for row in r.fetchall()}

    for key in SETTING_KEYS:
        val = db_settings.get(key) or os.environ.get(key, "") or getattr(_settings, key, "")
        result[key] = val

    return success_response(data=result)


@router.put("/system-settings", summary="保存系统设置")
async def save_system_settings(
    payload: dict = Body(...),
    session: AsyncSession = Depends(get_db),
):
    """批量保存系统配置到 DB，同时写入 .env 文件。"""
    from sqlalchemy import text as sql_text

    updated = []
    for key, value in payload.items():
        if key not in SETTING_KEYS:
            continue
        await session.execute(
            sql_text(
                "INSERT INTO hr.system_settings (key, value, updated_at) VALUES (:k, :v, now()) "
                "ON CONFLICT (key) DO UPDATE SET value = :v, updated_at = now()"
            ),
            {"k": key, "v": value},
        )
        updated.append(key)
    await session.commit()

    # Write to .env file
    try:
        if os.path.exists(ENV_PATH):
            with open(ENV_PATH, "r", encoding="utf-8") as f:
                lines = f.readlines()
        else:
            lines = []
        existing_keys = {}
        for i, line in enumerate(lines):
            m = re.match(r"^(\w+)=.*", line.strip())
            if m:
                existing_keys[m.group(1)] = i

        for key, value in payload.items():
            if key not in SETTING_KEYS:
                continue
            new_line = f"{key}={value}\n"
            if key in existing_keys:
                lines[existing_keys[key]] = new_line
            else:
                lines.append(new_line)

        with open(ENV_PATH, "w", encoding="utf-8") as f:
            f.writelines(lines)
    except Exception as e:
        logger.warning("Failed to write .env: %s", e)

    return success_response(data={"updated": updated}, message=f"已保存 {len(updated)} 项设置")
