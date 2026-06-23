"""安全模块文档解析工具 — Markdown 提取。

从 app/platform/integrations/ai/document_parser.py 迁移过来，
避免安全模块直接修改平台层代码。
"""

from pathlib import Path


def extract_to_markdown(file_path: str, max_chars: int = 50000) -> str:
    """Extract content as Markdown-formatted text.

    - XLSX/XLS:  tab-separated rows → Markdown table (| col | col |)
    - DOCX:      paragraphs with heading detection
    - PDF/TXT/MD: same as extract_text
    """
    from app.platform.integrations.ai.document_parser import DocumentParser

    ext = Path(file_path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        text = _extract_xlsx_to_markdown(file_path)
    elif ext == ".docx":
        text = _extract_docx_to_markdown(file_path)
    else:
        text = DocumentParser.extract_text(file_path, max_chars=999999)

    if len(text) > max_chars:
        text = text[:max_chars] + "\n\n...（文档过长，已截断）"
    return text.strip()


def _extract_xlsx_to_markdown(path: str) -> str:
    """Extract Excel content as Markdown tables (one table per sheet)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    parts: list[str] = []
    # 限制：最多 10 个 sheet，每个 sheet 最多 500 行，避免 prompt 过长
    for sheet_name in wb.sheetnames[:10]:
        ws = wb[sheet_name]
        parts.append(f"## 📊 Sheet: {sheet_name}\n")
        rows: list[list[str]] = []
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(cells)
                row_count += 1
                if row_count >= 500:
                    parts.append("*(表格超过 500 行，已截断)*\n")
                    break
        if not rows:
            parts.append("*(空表格)*\n")
            continue
        # 构建 Markdown 表格
        max_cols = max(len(r) for r in rows)
        # 表头 = 第一行
        header = rows[0]
        # 补齐列数
        header += [""] * (max_cols - len(header))
        md_rows: list[str] = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * max_cols) + " |",
        ]
        for row in rows[1:]:
            row += [""] * (max_cols - len(row))
            md_rows.append("| " + " | ".join(row) + " |")
        parts.append("\n".join(md_rows) + "\n")
    wb.close()
    return "\n\n".join(parts)


def _extract_docx_to_markdown(path: str) -> str:
    """Extract DOCX content with basic heading detection."""
    from docx import Document

    doc = Document(path)
    lines: list[str] = []
    for p in doc.paragraphs:
        text = p.text.strip()
        if not text:
            continue
        # 检测标题样式
        if p.style and p.style.name and p.style.name.startswith("Heading"):
            level = p.style.name.replace("Heading", "").strip()
            try:
                lv = int(level)
            except ValueError:
                lv = 1
            lines.append(f"{'#' * min(lv, 4)} {text}")
        else:
            lines.append(text)
    return "\n\n".join(lines)
