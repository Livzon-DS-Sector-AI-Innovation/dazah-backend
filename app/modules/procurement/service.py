"""Procurement business workflows live here."""

import re
from datetime import UTC, date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties
from pypdf import PdfReader
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.procurement.models import (
    InvoiceRecognitionRecord,
    PurchaseRequest,
    PurchaseRequestApproval,
    PurchaseRequestItem,
)
from app.modules.procurement.repository import (
    InvoiceRecognitionRepository,
    PurchaseRequestRepository,
)
from app.modules.procurement.schemas import (
    InvoiceLineItem,
    InvoiceRecognitionResult,
    PurchaseApprovalRequest,
    PurchaseApprovalResult,
    PurchaseApprovalRole,
    PurchaseApprovalView,
    PurchaseOrderLineResponse,
    PurchaseRequestCreate,
    PurchaseRequestResponse,
    PurchaseRequestStatus,
    PurchaseRequestUpdate,
)

MONEY_PATTERN = r"[¥￥]?\s*([0-9]+(?:\.[0-9]+)?)"
NUMBER_PATTERN = r"[0-9]+(?:\.[0-9]+)?"
MONEY_QUANT = Decimal("0.01")

APPROVAL_ROLE_TO_PENDING_STATUS = {
    PurchaseApprovalRole.department_head: PurchaseRequestStatus.pending_department_head,
    PurchaseApprovalRole.responsible_leader: (
        PurchaseRequestStatus.pending_responsible_leader
    ),
}

PURCHASE_CATEGORY_LABELS = {
    "hardware": "五金材料",
    "computer": "电脑材料",
    "office": "办公用品",
    "raw-auxiliary": "原辅料",
    "chemical-glass": "化玻",
    "electrical": "电器",
    "labor-protection": "劳保",
}

PURCHASE_ORDER_EXPORT_HEADERS = [
    "序号",
    "商品名称",
    "规格型号",
    "用途",
    "材质",
    "品牌",
    "数量",
    "单位",
    "单价（元）",
    "总额（元）",
    "备注",
]

PURCHASE_ORDER_EXPORT_COLUMN_WIDTHS = {
    "A": 18,
    "B": 19.33,
    "C": 13.08,
    "D": 17.83,
    "E": 5.5,
    "F": 9.5,
    "G": 5.5,
    "H": 10.25,
    "I": 13,
    "J": 11.58,
    "K": 29.75,
}


class DuplicateInvoiceError(ValueError):
    def __init__(self, existing_record: InvoiceRecognitionRecord) -> None:
        self.existing_record = existing_record
        super().__init__(
            "发票已识别过，"
            f"记录 ID：{existing_record.id}，文件：{existing_record.file_name}"
        )


def recognize_invoice_pdf(
    pdf_bytes: bytes,
    *,
    include_details: bool = False,
) -> InvoiceRecognitionResult:
    """Extract invoice fields from an electronic VAT invoice PDF."""
    reader = PdfReader(BytesIO(pdf_bytes))
    text_parts: list[str] = []
    for page in reader.pages:
        text_parts.append(page.extract_text(extraction_mode="layout") or "")

    raw_text = "\n".join(text_parts).strip()
    if not raw_text:
        raise ValueError("未能从 PDF 中提取到文本，请确认文件为电子发票 PDF")

    return _parse_invoice_text(raw_text, include_details=include_details)


async def recognize_and_store_invoice_pdf(
    db: AsyncSession,
    pdf_bytes: bytes,
    *,
    file_name: str,
    include_details: bool = False,
) -> InvoiceRecognitionRecord:
    result = recognize_invoice_pdf(pdf_bytes, include_details=include_details)
    file_sha256 = sha256(pdf_bytes).hexdigest()
    duplicate_key = _build_invoice_duplicate_key(result)
    repository = InvoiceRecognitionRepository(db)
    duplicate = await repository.find_duplicate(
        duplicate_key=duplicate_key,
        source_file_sha256=file_sha256,
    )
    if duplicate:
        raise DuplicateInvoiceError(duplicate)

    record = InvoiceRecognitionRecord(
        file_name=file_name,
        include_details=include_details,
        invoice_number=result.invoice_number,
        duplicate_key=duplicate_key,
        source_file_sha256=file_sha256,
        invoice_date=result.invoice_date,
        seller_name=result.seller_name,
        total_tax_amount=result.total_tax_amount,
        total_amount_with_tax_small=result.total_amount_with_tax_small,
        line_items=[
            item.model_dump(mode="json", exclude_none=False)
            for item in result.line_items
        ],
        raw_text=result.raw_text,
    )
    return await repository.create(record)


async def list_invoice_recognition_records(
    db: AsyncSession,
    *,
    keyword: str | None = None,
    seller_name: str | None = None,
    invoice_number: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[InvoiceRecognitionRecord], int]:
    return await InvoiceRecognitionRepository(db).list_records(
        keyword=keyword,
        seller_name=seller_name,
        invoice_number=invoice_number,
        page=page,
        page_size=page_size,
    )


async def delete_invoice_recognition_record(
    db: AsyncSession,
    record_id: UUID,
) -> bool:
    return await InvoiceRecognitionRepository(db).delete_record(record_id)


async def batch_delete_invoice_recognition_records(
    db: AsyncSession,
    record_ids: list[UUID],
) -> int:
    return await InvoiceRecognitionRepository(db).batch_delete_records(record_ids)


async def create_purchase_request(
    db: AsyncSession,
    data: PurchaseRequestCreate,
) -> PurchaseRequestResponse:
    repository = PurchaseRequestRepository(db)
    items, total_amount = _build_purchase_request_items(data.items)
    now = datetime.now(UTC)
    request = PurchaseRequest(
        category=data.category.value,
        request_department=data.request_department,
        request_date=data.request_date,
        status=PurchaseRequestStatus.draft.value,
        total_amount=total_amount,
        status_updated_at=now,
    )
    created = await repository.create(request, items)
    return await _get_purchase_request_response(repository, created.id)


async def update_purchase_request(
    db: AsyncSession,
    request_id: UUID,
    data: PurchaseRequestUpdate,
) -> PurchaseRequestResponse:
    repository = PurchaseRequestRepository(db)
    request = await repository.get(request_id)
    if not request:
        raise ValueError("采购申请不存在")
    if request.status not in {
        PurchaseRequestStatus.draft.value,
        PurchaseRequestStatus.rejected.value,
    }:
        raise ValueError("只有草稿或已驳回的采购申请可以编辑")

    if data.request_department is not None:
        request.request_department = data.request_department
    if data.request_date is not None:
        request.request_date = data.request_date
    if data.items is not None:
        items, total_amount = _build_purchase_request_items(data.items)
        await repository.replace_items(request_id, items)
        request.total_amount = total_amount
    await db.flush()
    return await _get_purchase_request_response(repository, request_id)


async def get_purchase_request(
    db: AsyncSession,
    request_id: UUID,
) -> PurchaseRequestResponse:
    repository = PurchaseRequestRepository(db)
    return await _get_purchase_request_response(repository, request_id)


async def list_purchase_requests(
    db: AsyncSession,
    *,
    category: str | None = None,
    status: str | None = None,
    approval_role: PurchaseApprovalRole | None = None,
    approval_view: PurchaseApprovalView = PurchaseApprovalView.pending,
    keyword: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[PurchaseRequestResponse], int]:
    repository = PurchaseRequestRepository(db)
    if approval_role:
        if approval_view == PurchaseApprovalView.pending:
            status = APPROVAL_ROLE_TO_PENDING_STATUS[approval_role].value
        else:
            approval_result = (
                PurchaseApprovalResult.approved
                if approval_view == PurchaseApprovalView.completed
                else PurchaseApprovalResult.rejected
            )
            requests, total = await repository.list_requests_by_approval(
                approval_role=approval_role.value,
                result=approval_result.value,
                category=category,
                keyword=keyword,
                page=page,
                page_size=page_size,
            )
            responses = [
                await _get_purchase_request_response(repository, request.id)
                for request in requests
            ]
            return responses, total

    requests, total = await repository.list_requests(
        category=category,
        status=status,
        keyword=keyword,
        page=page,
        page_size=page_size,
    )
    responses = [
        await _get_purchase_request_response(repository, request.id)
        for request in requests
    ]
    return responses, total


async def list_purchase_order_lines(
    db: AsyncSession,
    *,
    category: str | None = None,
    year: int,
    month: int,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[PurchaseOrderLineResponse], int]:
    repository = PurchaseRequestRepository(db)
    start_date, end_date = _month_date_range(year, month)
    rows, total = await repository.list_purchase_order_lines(
        start_date=start_date,
        end_date=end_date,
        status=PurchaseRequestStatus.approved.value,
        category=category,
        page=page,
        page_size=page_size,
    )
    return [_build_purchase_order_line(request, item) for request, item in rows], total


async def export_purchase_order_lines_xlsx(
    db: AsyncSession,
    *,
    category: str | None = None,
    year: int,
    month: int,
) -> bytes:
    repository = PurchaseRequestRepository(db)
    start_date, end_date = _month_date_range(year, month)
    rows, _ = await repository.list_purchase_order_lines(
        start_date=start_date,
        end_date=end_date,
        status=PurchaseRequestStatus.approved.value,
        category=category,
    )
    lines = [_build_purchase_order_line(request, item) for request, item in rows]
    category_label = PURCHASE_CATEGORY_LABELS.get(category or "", "全部类别")
    workbook = _build_purchase_order_workbook(
        lines,
        year=year,
        month=month,
        category_label=category_label,
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()



async def submit_purchase_request(
    db: AsyncSession,
    request_id: UUID,
) -> PurchaseRequestResponse:
    repository = PurchaseRequestRepository(db)
    request = await repository.get(request_id)
    if not request:
        raise ValueError("采购申请不存在")
    if request.status not in {
        PurchaseRequestStatus.draft.value,
        PurchaseRequestStatus.rejected.value,
    }:
        raise ValueError("只有草稿或已驳回的采购申请可以提交")
    items = await repository.list_items(request_id)
    if not items:
        raise ValueError("采购申请至少需要一条明细")

    now = datetime.now(UTC)
    request.status = PurchaseRequestStatus.pending_department_head.value
    request.rejected_step = None
    request.status_updated_at = now
    await db.flush()
    return await _get_purchase_request_response(repository, request_id)


async def approve_purchase_request(
    db: AsyncSession,
    request_id: UUID,
    data: PurchaseApprovalRequest,
) -> PurchaseRequestResponse:
    if data.result != PurchaseApprovalResult.approved:
        raise ValueError("审批通过接口的结果必须为 approved")
    return await _review_purchase_request(db, request_id, data)


async def reject_purchase_request(
    db: AsyncSession,
    request_id: UUID,
    data: PurchaseApprovalRequest,
) -> PurchaseRequestResponse:
    if data.result != PurchaseApprovalResult.rejected:
        raise ValueError("审批驳回接口的结果必须为 rejected")
    return await _review_purchase_request(db, request_id, data)


async def _review_purchase_request(
    db: AsyncSession,
    request_id: UUID,
    data: PurchaseApprovalRequest,
) -> PurchaseRequestResponse:
    repository = PurchaseRequestRepository(db)
    request = await repository.get(request_id)
    if not request:
        raise ValueError("采购申请不存在")

    expected_status = APPROVAL_ROLE_TO_PENDING_STATUS[data.approval_role].value
    if request.status != expected_status:
        raise ValueError("当前采购申请不在该审批步骤")

    now = datetime.now(UTC)
    await repository.add_approval(
        PurchaseRequestApproval(
            purchase_request_id=str(request_id),
            approval_role=data.approval_role.value,
            result=data.result.value,
            opinion=data.opinion,
            approver_name=data.approver_name,
            approval_time=now,
        )
    )

    if data.result == PurchaseApprovalResult.rejected:
        request.status = PurchaseRequestStatus.rejected.value
        request.rejected_step = data.approval_role.value
    elif data.approval_role == PurchaseApprovalRole.department_head:
        request.status = PurchaseRequestStatus.pending_responsible_leader.value
        request.rejected_step = None
    else:
        request.status = PurchaseRequestStatus.approved.value
        request.rejected_step = None
    request.status_updated_at = now
    await db.flush()
    return await _get_purchase_request_response(repository, request_id)


def _build_purchase_request_items(
    item_inputs: list,
) -> tuple[list[PurchaseRequestItem], Decimal]:
    items: list[PurchaseRequestItem] = []
    total_amount = Decimal("0")
    for index, item in enumerate(item_inputs, start=1):
        line_amount = _calculate_line_amount(item.quantity, item.unit_price)
        total_amount += line_amount
        items.append(
            PurchaseRequestItem(
                purchase_request_id="",
                sequence=index,
                product_name=item.product_name,
                specification=item.specification,
                purpose=item.purpose,
                material=item.material,
                brand=item.brand,
                quantity=item.quantity,
                unit=item.unit,
                unit_price=item.unit_price,
                total_amount=line_amount,
                remarks=item.remarks,
            )
        )
    return items, total_amount.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _calculate_line_amount(quantity: Decimal, unit_price: Decimal) -> Decimal:
    return (quantity * unit_price).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _month_date_range(year: int, month: int) -> tuple[date, date]:
    start_date = date(year, month, 1)
    if month == 12:
        return start_date, date(year + 1, 1, 1)
    return start_date, date(year, month + 1, 1)


def _build_purchase_order_line(
    request: PurchaseRequest,
    item: PurchaseRequestItem,
) -> PurchaseOrderLineResponse:
    category_label = PURCHASE_CATEGORY_LABELS.get(request.category, request.category)
    return PurchaseOrderLineResponse.model_validate(
        {
            "request_id": request.id,
            "category": request.category,
            "category_label": category_label,
            "request_department": request.request_department,
            "request_date": request.request_date,
            "item_id": item.id,
            "item_sequence": item.sequence,
            "product_name": item.product_name,
            "specification": item.specification,
            "purpose": item.purpose,
            "material": item.material,
            "brand": item.brand,
            "quantity": item.quantity,
            "unit": item.unit,
            "unit_price": item.unit_price,
            "total_amount": item.total_amount,
            "remarks": item.remarks,
        }
    )


def _build_purchase_order_workbook(
    lines: list[PurchaseOrderLineResponse],
    *,
    year: int,
    month: int,
    category_label: str,
) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    _apply_purchase_order_sheet_setup(worksheet)

    worksheet.row_dimensions[1].height = 9
    _write_merged_title(
        worksheet,
        2,
        "丽珠集团（宁夏）制药有限公司",
    )
    _write_merged_title(
        worksheet,
        3,
        f"{year}年{month:02d}月份{category_label}申购单汇总",
    )

    row_index = 4
    total_rows: list[int] = []
    for department, department_lines in _group_purchase_order_lines_by_department(
        lines
    ):
        _write_department_row(worksheet, row_index, department)
        row_index += 1
        _write_purchase_order_header_row(worksheet, row_index)
        row_index += 1

        first_detail_row = row_index
        for index, line in enumerate(department_lines, start=1):
            _write_purchase_order_detail_row(worksheet, row_index, index, line)
            row_index += 1

        total_row = row_index
        total_rows.append(total_row)
        _write_purchase_order_total_row(
            worksheet,
            total_row,
            "合计",
            f"=SUM(J{first_detail_row}:J{row_index - 1})",
        )
        row_index += 1

    total_formula = (
        f"=SUM({','.join(f'J{row}' for row in total_rows)})"
        if total_rows
        else "0"
    )
    _write_purchase_order_total_row(worksheet, row_index, "总计", total_formula)
    row_index += 1
    _write_signature_row(worksheet, row_index)
    worksheet.print_area = f"A1:K{row_index}"
    return workbook


def _build_purchase_order_row_values(
    index: int,
    line: PurchaseOrderLineResponse,
) -> list[str | int | Decimal]:
    return [
        index,
        line.product_name,
        line.specification,
        line.purpose,
        line.material,
        line.brand,
        line.quantity,
        line.unit,
        line.unit_price,
        line.total_amount,
        line.remarks,
    ]


def _group_purchase_order_lines_by_department(
    lines: list[PurchaseOrderLineResponse],
) -> list[tuple[str, list[PurchaseOrderLineResponse]]]:
    groups: dict[str, list[PurchaseOrderLineResponse]] = {}
    for line in sorted(
        lines,
        key=lambda item: (
            item.request_department,
            item.request_date,
            item.category,
            str(item.request_id),
            item.item_sequence,
        ),
    ):
        groups.setdefault(line.request_department, []).append(line)
    return list(groups.items())


def _apply_purchase_order_sheet_setup(worksheet) -> None:
    for column_letter, width in PURCHASE_ORDER_EXPORT_COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column_letter].width = width
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    worksheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.75,
        bottom=0.75,
        header=0.3,
        footer=0.3,
    )


def _write_merged_title(worksheet, row_index: int, value: str) -> None:
    worksheet.merge_cells(
        start_row=row_index,
        start_column=1,
        end_row=row_index,
        end_column=11,
    )
    worksheet.row_dimensions[row_index].height = 27
    cell = worksheet.cell(row_index, 1, value)
    cell.font = Font(name="宋体", size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_department_row(worksheet, row_index: int, department: str) -> None:
    worksheet.row_dimensions[row_index].height = 32.15
    worksheet.cell(row_index, 1, f"申购部门：{department}")
    fill = PatternFill("solid", fgColor="D9E1F4")
    border = _purchase_order_border(top=True, bottom=True)
    for cell in worksheet[row_index][0:11]:
        cell.font = Font(name="宋体", size=12, bold=True)
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=False,
        )
        cell.fill = fill
        cell.border = border


def _write_purchase_order_header_row(worksheet, row_index: int) -> None:
    worksheet.row_dimensions[row_index].height = 32.15
    for column_index, header in enumerate(PURCHASE_ORDER_EXPORT_HEADERS, start=1):
        cell = worksheet.cell(row_index, column_index, header)
        cell.font = Font(name="宋体", size=12, bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = _purchase_order_border()


def _write_purchase_order_detail_row(
    worksheet,
    row_index: int,
    index: int,
    line: PurchaseOrderLineResponse,
) -> None:
    worksheet.row_dimensions[row_index].height = 32.15
    for column_index, value in enumerate(
        _build_purchase_order_row_values(index, line),
        start=1,
    ):
        cell = worksheet.cell(row_index, column_index, value)
        cell.font = Font(name="宋体", size=12)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = _purchase_order_border()
    for column_index in (2, 3, 4, 5, 6, 11):
        worksheet.cell(row_index, column_index).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
    worksheet.cell(row_index, 9).number_format = "0.00"
    worksheet.cell(row_index, 10).number_format = "0.00"


def _write_purchase_order_total_row(
    worksheet,
    row_index: int,
    label: str,
    total_formula: str,
) -> None:
    worksheet.row_dimensions[row_index].height = 32.15
    worksheet.cell(row_index, 1, label)
    worksheet.cell(row_index, 10, total_formula)
    for cell in worksheet[row_index][0:11]:
        cell.font = Font(name="宋体", size=12, bold=(label == "合计"))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _purchase_order_border(top=True, bottom=True)
    worksheet.cell(row_index, 10).font = Font(name="宋体", size=12, bold=True)
    worksheet.cell(row_index, 10).number_format = "0.00"


def _write_signature_row(worksheet, row_index: int) -> None:
    worksheet.merge_cells(
        start_row=row_index,
        start_column=1,
        end_row=row_index,
        end_column=11,
    )
    worksheet.row_dimensions[row_index].height = 32.15
    cell = worksheet.cell(
        row_index,
        1,
        " 总经理：                       财务总监："
        "                          部门负责人："
        "                       统计人：",
    )
    cell.font = Font(name="宋体", size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _purchase_order_border(*, top: bool = True, bottom: bool = True) -> Border:
    side = Side(style="thin", color="000000")
    return Border(
        left=side,
        right=side,
        top=side if top else Side(style=None),
        bottom=side if bottom else Side(style=None),
    )


async def _get_purchase_request_response(
    repository: PurchaseRequestRepository,
    request_id: UUID,
) -> PurchaseRequestResponse:
    request = await repository.get(request_id)
    if not request:
        raise ValueError("采购申请不存在")
    items = await repository.list_items(request_id)
    approvals = await repository.list_approvals(request_id)
    return PurchaseRequestResponse.model_validate(
        {
            **request.__dict__,
            "items": items,
            "approvals": approvals,
        }
    )


def _parse_invoice_text(
    raw_text: str,
    *,
    include_details: bool = False,
) -> InvoiceRecognitionResult:
    lines = [line.rstrip() for line in raw_text.splitlines() if line.strip()]

    return InvoiceRecognitionResult(
        invoice_number=_search_first(r"发票号码[:：]\s*([0-9]{8,})", raw_text),
        invoice_date=_search_first(
            r"开票日期[:：]\s*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日)",
            raw_text,
        ),
        seller_name=_extract_seller_name(lines),
        total_tax_amount=_extract_total_tax_amount(lines),
        total_amount_with_tax_small=_extract_total_amount(raw_text),
        line_items=_extract_line_items(lines) if include_details else [],
        raw_text=raw_text,
    )


def _build_invoice_duplicate_key(result: InvoiceRecognitionResult) -> str | None:
    if not result.invoice_number:
        return None

    parts = [
        _normalize_duplicate_part(result.invoice_number),
        _normalize_duplicate_part(result.invoice_date),
        _normalize_duplicate_part(result.seller_name),
        _normalize_duplicate_part(
            _format_money_for_key(result.total_amount_with_tax_small)
        ),
    ]
    return "|".join(parts)


def _normalize_duplicate_part(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", "", value).strip().lower()


def _format_money_for_key(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return f"{value.quantize(Decimal('0.01')):.2f}"


def _search_first(pattern: str, text: str) -> str | None:
    match = re.search(pattern, text)
    if not match:
        return None
    return match.group(1).strip()


def _extract_seller_name(lines: list[str]) -> str | None:
    for line in lines:
        if "名称" not in line:
            continue
        names = re.findall(
            r"名称[:：]\s*([^\n]+?)(?=\s{2,}\S+\s+名称[:：]|$)",
            line,
        )
        if len(names) >= 2:
            return names[1].strip()

        seller_name = _extract_seller_name_before_marker(line)
        if seller_name:
            return seller_name
    return None


def _extract_seller_name_before_marker(line: str) -> str | None:
    match = re.search(r"(.+?)\s*销\s*名称[:：]", line)
    if not match:
        return None

    candidates = [
        part.strip()
        for part in re.split(r"\s{2,}", match.group(1).strip())
        if part.strip()
    ]
    if not candidates:
        return None

    return candidates[-1]


def _extract_total_amount(text: str) -> Decimal | None:
    for line in text.splitlines():
        if "小写" not in line:
            continue

        amounts = re.findall(MONEY_PATTERN, line)
        if amounts:
            return _to_decimal(amounts[-1])

    return None


def _extract_total_tax_amount(lines: list[str]) -> Decimal | None:
    total_tax_amount: Decimal | None = None
    for line in lines:
        compact_line = re.sub(r"\s+", "", line)
        if "合计" not in compact_line or "小计" in compact_line:
            continue

        amounts = re.findall(MONEY_PATTERN, line)
        if len(amounts) < 2:
            continue

        total_tax_amount = _to_decimal(amounts[-1])

    return total_tax_amount


def _extract_line_items(lines: list[str]) -> list[InvoiceLineItem]:
    items: list[InvoiceLineItem] = []
    for line in lines:
        stripped_line = line.strip()
        if not stripped_line.startswith("*"):
            continue

        item = _parse_line_item(stripped_line)
        if item:
            items.append(item)

    return items


def _parse_line_item(line: str) -> InvoiceLineItem | None:
    parts = [part.strip() for part in re.split(r"\s{2,}", line) if part.strip()]
    tax_rate_index = _find_tax_rate_index(parts)
    if len(parts) < 4 or tax_rate_index is None:
        return None

    pre_tax_parts = parts[: tax_rate_index + 1]
    numeric_start = len(pre_tax_parts)
    for index in range(len(pre_tax_parts) - 1, -1, -1):
        if not _is_numeric_cluster(pre_tax_parts[index], allow_percent=True):
            break
        numeric_start = index

    if numeric_start == 0 or numeric_start >= len(pre_tax_parts):
        return None

    tax_amount = _to_decimal(
        _last_match(NUMBER_PATTERN, " ".join(parts[tax_rate_index:])),
    )
    tax_rate = _extract_tax_rate(parts[tax_rate_index])
    numeric_text = " ".join(pre_tax_parts[numeric_start:])
    dense_quantity = _extract_dense_quantity(numeric_text, tax_amount, tax_rate)
    if dense_quantity is not None:
        quantity = dense_quantity
    else:
        numeric_values = re.findall(
            NUMBER_PATTERN,
            _remove_tax_rate(numeric_text, tax_rate),
        )
        if len(numeric_values) < 2:
            return None

        amount = _to_decimal(numeric_values[-1])
        quantity = (
            _to_decimal(numeric_values[0])
            if len(numeric_values) >= 3
            else _extract_quantity(numeric_values[0], amount)
        )

    return InvoiceLineItem(
        project_name=parts[0],
        unit=pre_tax_parts[numeric_start - 1],
        quantity=quantity,
    )


def _find_tax_rate_index(parts: list[str]) -> int | None:
    for index, part in enumerate(parts):
        if "%" in part:
            return index
    return None


def _extract_tax_rate(value: str) -> Decimal | None:
    compact_value = re.sub(r"\s+", "", value)
    if "%" not in compact_value:
        return None

    before_percent = compact_value.split("%", 1)[0]
    for rate in ("13", "9", "6", "5", "3", "1", "0"):
        if before_percent.endswith(rate):
            return Decimal(rate)

    return None


def _extract_dense_quantity(
    numeric_text: str,
    tax_amount: Decimal | None,
    tax_rate: Decimal | None,
) -> Decimal | None:
    if re.search(r"\s", numeric_text.strip()):
        return None

    compact_text = re.sub(r"\s+", "", numeric_text)
    if compact_text.count(".") < 2 or tax_amount is None or not tax_rate:
        return None

    before_percent = compact_text.split("%", 1)[0]
    rate_text = _format_decimal(tax_rate)
    if not before_percent.endswith(rate_text):
        return None

    dense_text = before_percent[: -len(rate_text)]
    estimated_amount = tax_amount / (tax_rate / Decimal(100))
    best_quantity: Decimal | None = None
    best_delta: Decimal | None = None

    for quantity_end in range(1, len(dense_text)):
        quantity_text = dense_text[:quantity_end]
        if quantity_text.count(".") != 1:
            continue
        decimal_part = quantity_text.rsplit(".", 1)[1]
        if len(decimal_part) > 4:
            continue

        quantity = _to_decimal(quantity_text)
        if quantity is None or quantity <= 0:
            continue

        remaining_text = dense_text[quantity_end:]
        for unit_price_end in range(1, len(remaining_text) + 1):
            unit_price_text = remaining_text[:unit_price_end]
            if unit_price_text.count(".") != 1:
                continue

            unit_price = _to_decimal(unit_price_text)
            if unit_price is None:
                continue

            delta = abs(quantity * unit_price - estimated_amount)
            if delta > Decimal("0.2"):
                continue

            if best_delta is None or delta < best_delta:
                best_delta = delta
                best_quantity = quantity

    return best_quantity


def _remove_tax_rate(numeric_text: str, tax_rate: Decimal | None) -> str:
    if tax_rate is None or "%" not in numeric_text:
        return numeric_text

    rate_text = _format_decimal(tax_rate)
    before_percent, after_percent = numeric_text.split("%", 1)
    return before_percent.removesuffix(rate_text) + after_percent


def _last_match(pattern: str, text: str) -> str | None:
    matches = re.findall(pattern, text)
    return matches[-1] if matches else None


def _format_decimal(value: Decimal) -> str:
    return format(value.normalize(), "f")


def _is_numeric_cluster(value: str, *, allow_percent: bool = False) -> bool:
    pattern = r"[0-9¥￥\s.,%]+"
    if not allow_percent:
        pattern = r"[0-9¥￥\s.,]+"
    return bool(re.fullmatch(pattern, value))


def _extract_quantity(
    quantity_and_price: str,
    amount: Decimal | None,
) -> Decimal | None:
    numbers = re.findall(NUMBER_PATTERN, quantity_and_price.replace(",", ""))
    if not numbers:
        return None

    if len(numbers) >= 2:
        return _to_decimal(numbers[0])

    token = numbers[0]
    if "." not in token:
        return _to_decimal(token)

    return _infer_conjoined_quantity(token, amount)


def _infer_conjoined_quantity(token: str, amount: Decimal | None) -> Decimal | None:
    if amount is None:
        return None

    decimal_index = token.find(".")
    best_quantity: Decimal | None = None
    best_delta: Decimal | None = None
    for split_index in range(1, decimal_index):
        quantity = _to_decimal(token[:split_index])
        unit_price = _to_decimal(token[split_index:])
        if quantity is None or unit_price is None:
            continue

        delta = abs(quantity * unit_price - amount)
        if best_delta is None or delta < best_delta:
            best_delta = delta
            best_quantity = quantity

    if best_delta is not None and best_delta <= Decimal("0.02"):
        return best_quantity

    return None


def _to_decimal(value: str | None) -> Decimal | None:
    if not value:
        return None
    try:
        return Decimal(value.replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None
