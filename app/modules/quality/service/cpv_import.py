"""CPV Import service layer."""

import uuid
from datetime import date, datetime
from io import BytesIO

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundException
from app.modules.quality import repository as repo
from app.modules.quality.schemas import (
    CpvImportConfirmRequest,
    CpvImportPreviewResponse,
    CpvImportTaskResponse,
)


def _parse_date(value: str) -> date | None:
    """解析日期"""
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _check_abnormal(
    value: str | None,
    lower_limit: float | None,
    upper_limit: float | None,
) -> bool:
    """检查是否异常"""
    if value is None or value == "":
        return False
    
    try:
        num_val = float(value)
    except (ValueError, TypeError):
        return False
    
    if lower_limit is not None and num_val < lower_limit:
        return True
    if upper_limit is not None and num_val > upper_limit:
        return True
    
    return False


async def preview_import(
    db: AsyncSession,
    file_content: bytes,
    product_id: uuid.UUID,
    data_type: str,
    import_mode: str,
) -> CpvImportPreviewResponse:
    """预览导入"""
    # 使用 openpyxl 读取 Excel
    import openpyxl
    
    wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    ws = wb.active
    
    # 获取表头
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    # 获取参数定义
    parameters = await repo.get_parameters(db, product_id, data_type)
    param_map = {p.name: p for p in parameters}
    param_map.update({p.code: p for p in parameters if p.code})
    
    # 匹配列
    matched_params = {}
    unmatched_cols = []
    
    for col_idx, header in enumerate(headers):
        if header and str(header).strip() in param_map:
            matched_params[col_idx] = param_map[str(header).strip()]
        elif header and str(header).strip() not in ("批号", "batch_no", "生产日期", "production_date"):
            unmatched_cols.append(str(header))
    
    # 解析数据行
    error_rows = []
    valid_count = 0
    batch_no_col = next((i for i, h in enumerate(headers) if h in ("批号", "batch_no")), None)
    date_col = next((i for i, h in enumerate(headers) if h in ("生产日期", "production_date")), None)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))
        errors = []
        
        # 校验批号
        batch_no = row[batch_no_col] if batch_no_col is not None else None
        if not batch_no:
            errors.append("批号不能为空")
        
        # 校验日期
        prod_date = row[date_col] if date_col is not None else None
        if not prod_date:
            errors.append("生产日期不能为空")
        else:
            parsed_date = _parse_date(str(prod_date))
            if not parsed_date:
                errors.append(f"生产日期格式错误: {prod_date}")
        
        # 校验数值
        for col_idx, param in matched_params.items():
            value = row[col_idx] if col_idx < len(row) else None
            if value is not None and value != "":
                try:
                    float(value)
                except (ValueError, TypeError):
                    if value not in ("未检出", "-"):
                        errors.append(f"{param.name} 必须是数字: {value}")
        
        if errors:
            error_rows.append({
                "row_number": row_idx,
                "error_message": "; ".join(errors),
                "row_data": row_data,
            })
        else:
            valid_count += 1
    
    wb.close()
    
    return CpvImportPreviewResponse(
        total_rows=valid_count + len(error_rows),
        valid_rows=valid_count,
        error_rows=error_rows,
        matched_parameters=[p.name for p in matched_params.values()],
        unmatched_columns=unmatched_cols,
    )


async def confirm_import(
    db: AsyncSession,
    file_content: bytes,
    request: CpvImportConfirmRequest,
    current_user_id: uuid.UUID | None = None,
) -> CpvImportTaskResponse:
    """确认导入"""
    import openpyxl
    
    # 创建导入任务
    task = await repo.create_import_task(
        db,
        {
            "file_name": request.file_name,
            "product_id": request.product_id,
            "data_type": request.data_type,
            "import_mode": request.import_mode,
            "status": "processing",
            "created_by": current_user_id,
        },
    )
    
    # 读取 Excel
    wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    # 获取参数定义
    parameters = await repo.get_parameters(db, request.product_id, request.data_type)
    param_map = {p.name: p for p in parameters}
    param_map.update({p.code: p for p in parameters if p.code})
    
    # 匹配列
    matched_params = {}
    for col_idx, header in enumerate(headers):
        if header and str(header).strip() in param_map:
            matched_params[col_idx] = param_map[str(header).strip()]
    
    batch_no_col = next((i for i, h in enumerate(headers) if h in ("批号", "batch_no")), None)
    date_col = next((i for i, h in enumerate(headers) if h in ("生产日期", "production_date")), None)
    
    # 覆盖模式：删除旧数据
    if request.import_mode == "overwrite":
        await repo.delete_batches_by_product(db, request.product_id, request.data_type)
    
    # 导入数据
    total_rows = 0
    success_rows = 0
    failed_rows = 0
    error_details = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        total_rows += 1
        
        batch_no = row[batch_no_col] if batch_no_col is not None else None
        prod_date = row[date_col] if date_col is not None else None
        
        if not batch_no or not prod_date:
            failed_rows += 1
            error_details.append({"row": row_idx, "error": "批号或日期为空"})
            continue
        
        parsed_date = _parse_date(str(prod_date))
        if not parsed_date:
            failed_rows += 1
            error_details.append({"row": row_idx, "error": "日期格式错误"})
            continue
        
        # 检查批号是否已存在
        existing_batch = await repo.get_batch_by_no(
            db, request.product_id, str(batch_no), request.data_type
        )
        
        if existing_batch:
            if request.import_mode == "create":
                failed_rows += 1
                error_details.append({"row": row_idx, "error": f"批号已存在: {batch_no}"})
                continue
            batch = existing_batch
        else:
            batch = await repo.create_batch(
                db,
                {
                    "product_id": request.product_id,
                    "batch_no": str(batch_no),
                    "production_date": parsed_date,
                    "data_type": request.data_type,
                    "source": "excel",
                    "import_task_id": task.id,
                    "created_by": current_user_id,
                },
            )
        
        # 创建参数值
        values_data = []
        for col_idx, param in matched_params.items():
            value = row[col_idx] if col_idx < len(row) else None
            if value is not None and value != "":
                str_value = str(value)
                is_abnormal = _check_abnormal(str_value, param.lower_limit, param.upper_limit)
                values_data.append({
                    "batch_id": batch.id,
                    "parameter_id": param.id,
                    "actual_value": str_value,
                    "is_abnormal": is_abnormal,
                    "created_by": current_user_id,
                })
        
        if values_data:
            await repo.create_values_bulk(db, values_data)
        
        success_rows += 1
    
    wb.close()
    
    # 更新任务状态
    task = await repo.update_import_task(
        db,
        task.id,
        {
            "status": "completed",
            "total_rows": total_rows,
            "success_rows": success_rows,
            "failed_rows": failed_rows,
            "error_details": error_details[:100] if error_details else None,
        },
    )
    
    return CpvImportTaskResponse.model_validate(task)


async def get_import_tasks(
    db: AsyncSession,
    product_id: uuid.UUID | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[CpvImportTaskResponse], int]:
    """获取导入任务列表"""
    tasks, total = await repo.get_import_tasks(db, product_id, page, page_size)
    return [CpvImportTaskResponse.model_validate(t) for t in tasks], total


async def get_import_task_by_id(
    db: AsyncSession,
    task_id: uuid.UUID,
) -> CpvImportTaskResponse:
    """获取导入任务"""
    task = await repo.get_import_task_by_id(db, task_id)
    if not task:
        raise NotFoundException("导入任务", str(task_id))
    return CpvImportTaskResponse.model_validate(task)
