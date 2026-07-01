"""Excel export utility for regulatory documents."""

import io
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_STYLE = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

COLUMNS = [
    ("序号", 6),
    ("标题", 50),
    ("发布机构", 15),
    ("所属栏目", 20),
    ("发布日期", 12),
    ("状态", 10),
    ("分类", 12),
    ("AI 摘要", 50),
    ("相关性评分", 10),
    ("原文链接", 40),
    ("首次发现时间", 18),
]


def _format_cell_value(value) -> str:
    """格式化单元格值"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return f"{value:.0%}" if value <= 1 else f"{value:.1f}"
    return str(value)


def generate_regulatory_excel(
    documents: list,
    source_map: dict | None = None,
    channel_map: dict | None = None,
) -> io.BytesIO:
    """
    生成法规文档 Excel 文件。

    Args:
        documents: RegulatoryDocument 对象列表
        source_map: {source_id: source_name} 映射
        channel_map: {channel_id: channel_name} 映射

    Returns:
        BytesIO 对象，包含 Excel 文件内容
    """
    source_map = source_map or {}
    channel_map = channel_map or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "法规文档"

    # 标题行
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"法规文档导出 — {date.today().isoformat()}"
    title_cell.font = Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 表头
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = HEADER_STYLE
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 25

    # 数据行
    for row_idx, doc in enumerate(documents, 1):
        row_num = row_idx + 2
        values = [
            row_idx,
            doc.title,
            source_map.get(doc.source_id, ""),
            channel_map.get(doc.channel_id, ""),
            _format_cell_value(doc.publish_date),
            doc.status_text or "",
            doc.classification or "",
            doc.ai_summary or "",
            _format_cell_value(doc.ai_relevance_score),
            doc.original_url or "",
            _format_cell_value(doc.first_found_at),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            cell.font = Font(name="Microsoft YaHei", size=10)

        # 交替行背景色
        if row_idx % 2 == 0:
            alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
            for col_idx in range(1, len(values) + 1):
                ws.cell(row=row_num, column=col_idx).fill = alt_fill

    # 冻结表头
    ws.freeze_panes = "A3"

    # 自动筛选
    ws.auto_filter.ref = f"A2:K{len(documents) + 2}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
