import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.modules.procurement import service as procurement_service
from app.modules.procurement.schemas import (
    PurchaseApprovalRequest,
    PurchaseApprovalResult,
    PurchaseApprovalRole,
    PurchaseApprovalView,
    PurchaseRequestCategory,
    PurchaseRequestCreate,
    PurchaseRequestItemInput,
    PurchaseRequestStatus,
)


class FakeDb:
    async def flush(self) -> None:
        return None


class FakePurchaseRequestRepository:
    requests = {}
    items = {}
    approvals = {}

    def __init__(self, session) -> None:
        self.session = session

    @classmethod
    def reset(cls) -> None:
        cls.requests = {}
        cls.items = {}
        cls.approvals = {}

    async def create(self, request, items):
        request.id = uuid.uuid4()
        request.created_at = datetime.now(UTC)
        request.updated_at = request.created_at
        self.requests[request.id] = request
        for item in items:
            item.id = uuid.uuid4()
            item.purchase_request_id = str(request.id)
            item.created_at = request.created_at
            item.updated_at = request.created_at
        self.items[request.id] = items
        self.approvals[request.id] = []
        return request

    async def get(self, request_id):
        return self.requests.get(request_id)

    async def list_items(self, request_id):
        return self.items.get(request_id, [])

    async def list_approvals(self, request_id):
        return self.approvals.get(request_id, [])

    async def list_requests(
        self,
        *,
        category=None,
        status=None,
        keyword=None,
        page=1,
        page_size=20,
    ):
        records = list(self.requests.values())
        if category:
            records = [record for record in records if record.category == category]
        if status:
            records = [record for record in records if record.status == status]
        if keyword:
            records = [
                record
                for record in records
                if keyword in record.request_department
            ]
        return records[(page - 1) * page_size : page * page_size], len(records)

    async def list_requests_by_approval(
        self,
        *,
        approval_role,
        result,
        category=None,
        keyword=None,
        page=1,
        page_size=20,
    ):
        matching_ids = []
        for request_id, approvals in self.approvals.items():
            if any(
                approval.approval_role == approval_role
                and approval.result == result
                for approval in approvals
            ):
                matching_ids.append(request_id)

        records = [
            self.requests[request_id]
            for request_id in matching_ids
            if request_id in self.requests
        ]
        if category:
            records = [record for record in records if record.category == category]
        if keyword:
            records = [
                record
                for record in records
                if keyword in record.request_department
            ]
        return records[(page - 1) * page_size : page * page_size], len(records)

    async def list_purchase_order_lines(
        self,
        *,
        start_date,
        end_date,
        status,
        category=None,
        page=None,
        page_size=None,
    ):
        rows = []
        for request_id, request in self.requests.items():
            if request.status != status:
                continue
            if not start_date <= request.request_date < end_date:
                continue
            if category and request.category != category:
                continue
            for item in self.items.get(request_id, []):
                rows.append((request, item))

        rows.sort(
            key=lambda row: (
                row[0].request_date,
                row[0].category,
                row[0].request_department,
                row[1].sequence,
            )
        )
        total = len(rows)
        if page is not None and page_size is not None:
            rows = rows[(page - 1) * page_size : page * page_size]
        return rows, total

    async def replace_items(self, request_id, items):
        for item in items:
            item.id = uuid.uuid4()
            item.purchase_request_id = str(request_id)
            item.created_at = datetime.now(UTC)
            item.updated_at = item.created_at
        self.items[request_id] = items

    async def add_approval(self, approval):
        approval.id = uuid.uuid4()
        approval.created_at = datetime.now(UTC)
        approval.updated_at = approval.created_at
        self.approvals.setdefault(uuid.UUID(approval.purchase_request_id), []).append(
            approval
        )
        return approval


@pytest.fixture(autouse=True)
def fake_purchase_request_repository(monkeypatch):
    FakePurchaseRequestRepository.reset()
    monkeypatch.setattr(
        procurement_service,
        "PurchaseRequestRepository",
        FakePurchaseRequestRepository,
    )


def _create_payload() -> PurchaseRequestCreate:
    return PurchaseRequestCreate(
        category=PurchaseRequestCategory.hardware,
        request_department="102一车间",
        request_date=date(2026, 6, 28),
        items=[
            PurchaseRequestItemInput(
                product_name="碳鼓",
                specification="M1005",
                purpose="更换打印机碳鼓",
                material="",
                brand="惠普",
                quantity=Decimal("2"),
                unit="个",
                unit_price=Decimal("60.005"),
                remarks="申购人：郭娇",
            )
        ],
    )


def _create_payload_for(
    *,
    category: PurchaseRequestCategory = PurchaseRequestCategory.hardware,
    request_department: str = "102一车间",
    request_date: date = date(2026, 6, 28),
    product_name: str = "碳鼓",
) -> PurchaseRequestCreate:
    payload = _create_payload()
    payload.category = category
    payload.request_department = request_department
    payload.request_date = request_date
    payload.items[0].product_name = product_name
    return payload


async def _approve_request(request_id: uuid.UUID) -> None:
    await procurement_service.submit_purchase_request(FakeDb(), request_id)
    await procurement_service.approve_purchase_request(
        FakeDb(),
        request_id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.department_head,
            approver_name="部门负责人",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )
    await procurement_service.approve_purchase_request(
        FakeDb(),
        request_id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.responsible_leader,
            approver_name="分管领导",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )


@pytest.mark.anyio
async def test_purchase_request_amount_and_two_step_approval_flow() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )

    assert created.request_date == date(2026, 6, 28)
    assert created.status == PurchaseRequestStatus.draft
    assert created.total_amount == Decimal("120.01")
    assert created.items[0].total_amount == Decimal("120.01")

    submitted = await procurement_service.submit_purchase_request(
        FakeDb(),
        created.id,
    )
    assert submitted.status == PurchaseRequestStatus.pending_department_head

    department_approved = await procurement_service.approve_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.department_head,
            approver_name="部门负责人",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )
    assert (
        department_approved.status
        == PurchaseRequestStatus.pending_responsible_leader
    )
    assert department_approved.approvals[0].approval_role == (
        PurchaseApprovalRole.department_head
    )

    leader_approved = await procurement_service.approve_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.responsible_leader,
            approver_name="分管领导",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )
    assert leader_approved.status == PurchaseRequestStatus.approved
    assert len(leader_approved.approvals) == 2


@pytest.mark.anyio
async def test_purchase_request_reject_persists_approval_record() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )
    await procurement_service.submit_purchase_request(FakeDb(), created.id)

    rejected = await procurement_service.reject_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.department_head,
            approver_name="部门负责人",
            opinion="用途不清",
            result=PurchaseApprovalResult.rejected,
        ),
    )

    assert rejected.status == PurchaseRequestStatus.rejected
    assert rejected.rejected_step == PurchaseApprovalRole.department_head
    assert rejected.approvals[0].result == PurchaseApprovalResult.rejected
    assert rejected.approvals[0].opinion == "用途不清"


@pytest.mark.anyio
async def test_purchase_request_role_approval_views() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )
    await procurement_service.submit_purchase_request(FakeDb(), created.id)
    await procurement_service.approve_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.department_head,
            approver_name="部门负责人",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )

    department_completed, department_completed_total = (
        await procurement_service.list_purchase_requests(
            FakeDb(),
            category=PurchaseRequestCategory.hardware.value,
            approval_role=PurchaseApprovalRole.department_head,
            approval_view=PurchaseApprovalView.completed,
        )
    )
    leader_pending, leader_pending_total = (
        await procurement_service.list_purchase_requests(
            FakeDb(),
            category=PurchaseRequestCategory.hardware.value,
            approval_role=PurchaseApprovalRole.responsible_leader,
            approval_view=PurchaseApprovalView.pending,
        )
    )

    assert department_completed_total == 1
    assert department_completed[0].id == created.id
    assert (
        department_completed[0].status
        == PurchaseRequestStatus.pending_responsible_leader
    )
    assert leader_pending_total == 1
    assert leader_pending[0].id == created.id


@pytest.mark.anyio
async def test_purchase_request_role_rejected_view() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )
    await procurement_service.submit_purchase_request(FakeDb(), created.id)
    await procurement_service.reject_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.department_head,
            approver_name="部门负责人",
            opinion="用途不清",
            result=PurchaseApprovalResult.rejected,
        ),
    )

    department_rejected, department_rejected_total = (
        await procurement_service.list_purchase_requests(
            FakeDb(),
            category=PurchaseRequestCategory.hardware.value,
            approval_role=PurchaseApprovalRole.department_head,
            approval_view=PurchaseApprovalView.rejected,
        )
    )
    leader_rejected, leader_rejected_total = (
        await procurement_service.list_purchase_requests(
            FakeDb(),
            category=PurchaseRequestCategory.hardware.value,
            approval_role=PurchaseApprovalRole.responsible_leader,
            approval_view=PurchaseApprovalView.rejected,
        )
    )

    assert department_rejected_total == 1
    assert department_rejected[0].id == created.id
    assert department_rejected[0].status == PurchaseRequestStatus.rejected
    assert leader_rejected_total == 0
    assert leader_rejected == []


@pytest.mark.anyio
async def test_purchase_order_lines_include_only_approved_requests_in_month() -> None:
    approved = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(product_name="当月已通过"),
    )
    await _approve_request(approved.id)

    draft = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(product_name="当月草稿"),
    )
    before_month = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(request_date=date(2026, 5, 31), product_name="上月"),
    )
    after_month = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(request_date=date(2026, 7, 1), product_name="下月"),
    )
    await _approve_request(before_month.id)
    await _approve_request(after_month.id)

    lines, total = await procurement_service.list_purchase_order_lines(
        FakeDb(),
        year=2026,
        month=6,
    )

    assert total == 1
    assert lines[0].request_id == approved.id
    assert lines[0].product_name == "当月已通过"
    assert draft.id not in {line.request_id for line in lines}


@pytest.mark.anyio
async def test_purchase_order_lines_filter_category_and_paginate() -> None:
    hardware = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(
            category=PurchaseRequestCategory.hardware,
            product_name="五金材料",
        ),
    )
    office = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(
            category=PurchaseRequestCategory.office,
            product_name="办公用品",
        ),
    )
    await _approve_request(hardware.id)
    await _approve_request(office.id)

    all_lines, all_total = await procurement_service.list_purchase_order_lines(
        FakeDb(),
        year=2026,
        month=6,
        page=1,
        page_size=1,
    )
    office_lines, office_total = await procurement_service.list_purchase_order_lines(
        FakeDb(),
        category=PurchaseRequestCategory.office.value,
        year=2026,
        month=6,
    )

    assert all_total == 2
    assert len(all_lines) == 1
    assert office_total == 1
    assert office_lines[0].category == PurchaseRequestCategory.office
    assert office_lines[0].category_label == "办公用品"


@pytest.mark.anyio
async def test_purchase_order_xlsx_export_uses_reference_layout() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(
            category=PurchaseRequestCategory.hardware,
            request_department="102一车间",
            product_name="碳鼓",
        ),
    )
    other_department = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(
            category=PurchaseRequestCategory.hardware,
            request_department="103车间",
            product_name="粉盒",
        ),
    )
    await _approve_request(created.id)
    await _approve_request(other_department.id)

    xlsx_bytes = await procurement_service.export_purchase_order_lines_xlsx(
        FakeDb(),
        category=PurchaseRequestCategory.hardware.value,
        year=2026,
        month=6,
    )

    workbook = load_workbook(BytesIO(xlsx_bytes), data_only=False)
    worksheet = workbook.active
    merged_ranges = {str(cell_range) for cell_range in worksheet.merged_cells.ranges}

    assert worksheet.title == "Sheet1"
    assert "A2:K2" in merged_ranges
    assert "A3:K3" in merged_ranges
    assert worksheet["A2"].value == "丽珠集团（宁夏）制药有限公司"
    assert worksheet["A3"].value == "2026年06月份五金材料申购单汇总"
    assert worksheet["A4"].value == "申购部门：102一车间"
    assert [worksheet.cell(5, column).value for column in range(1, 12)] == (
        procurement_service.PURCHASE_ORDER_EXPORT_HEADERS
    )
    assert [worksheet.cell(6, column).value for column in range(1, 12)] == [
        1,
        "碳鼓",
        "M1005",
        "更换打印机碳鼓",
        None,
        "惠普",
        2,
        "个",
        60.005,
        120.01,
        "申购人：郭娇",
    ]
    assert worksheet["A7"].value == "合计"
    assert worksheet["J7"].value == "=SUM(J6:J6)"
    assert worksheet["A8"].value == "申购部门：103车间"
    assert worksheet["A12"].value == "总计"
    assert worksheet["J12"].value == "=SUM(J7,J11)"
    assert worksheet["A13"].value.startswith(" 总经理：")
    assert "A13:K13" in merged_ranges
    assert worksheet.column_dimensions["A"].width == 18
    assert worksheet.column_dimensions["B"].width == 19.33
    assert worksheet["A4"].fill.fgColor.rgb == "00D9E1F4"
    assert worksheet["A4"].alignment.wrap_text is not True
    assert worksheet["A5"].border.left.style == "thin"
    assert worksheet["K6"].alignment.wrap_text is True
    assert worksheet.page_setup.orientation == "landscape"
