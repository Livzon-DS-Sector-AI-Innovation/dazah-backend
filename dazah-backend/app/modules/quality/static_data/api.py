"""Static Data Module - API Routes

RESTful API endpoints for static data.
Prefix: /api/v1/quality/static-data/
"""

from datetime import date
from typing import Optional, List
import io

from fastapi import APIRouter, Depends, Query, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import CurrentUser, get_current_user
from app.core.response import ApiResponse
from app.modules.quality.static_data.service import StaticDataService
from app.modules.quality.static_data import schemas as s
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter(prefix="/static-data", tags=["Static Data"])


def _get_service(db: AsyncSession = Depends(get_db)) -> StaticDataService:
    return StaticDataService(db)


def _user_id(current_user: CurrentUser | None = Depends(get_current_user)) -> int:
    """Get current user ID"""
    return current_user.id if current_user else 0


# ========== 字典接口（固定选项） ==========

DICT_OPTIONS = {
    "equipment-category": [
        {"label": "色谱类", "value": "色谱类"},
        {"label": "称量类", "value": "称量类"},
        {"label": "灭菌类", "value": "灭菌类"},
        {"label": "微生物类", "value": "微生物类"},
        {"label": "其他", "value": "其他"},
    ],
    "equipment-status": [
        {"label": "在用", "value": "0"},
        {"label": "维修", "value": "1"},
        {"label": "封存", "value": "2"},
        {"label": "报废", "value": "3"},
    ],
    "verify-status": [
        {"label": "已完成", "value": "已完成"},
        {"label": "待验证", "value": "待验证"},
        {"label": "过期", "value": "过期"},
    ],
    "medium-type": [
        {"label": "干粉培养基", "value": "干粉培养基"},
        {"label": "颗粒培养基", "value": "颗粒培养基"},
        {"label": "液体培养基", "value": "液体培养基"},
        {"label": "显色培养基", "value": "显色培养基"},
        {"label": "环境监测培养基", "value": "环境监测培养基"},
    ],
    "reagent-purity": [
        {"label": "AR（分析纯）", "value": "AR"},
        {"label": "GR（优级纯）", "value": "GR"},
        {"label": "CP（化学纯）", "value": "CP"},
        {"label": "LR（实验纯）", "value": "LR"},
        {"label": "HPLC级", "value": "HPLC"},
        {"label": "农残级", "value": "农残级"},
    ],
    "danger-type": [
        {"label": "普通", "value": "普通"},
        {"label": "易制毒", "value": "易制毒"},
        {"label": "剧毒", "value": "剧毒"},
        {"label": "易燃易爆", "value": "易燃易爆"},
        {"label": "腐蚀性", "value": "腐蚀性"},
    ],
    "std-type": [
        {"label": "法定标准品", "value": "法定"},
        {"label": "工作标准品", "value": "工作"},
        {"label": "自制标准品", "value": "自制"},
        {"label": "参考标准品", "value": "参考"},
    ],
    "unit-type": [
        {"label": "质量", "value": "质量"},
        {"label": "体积", "value": "体积"},
        {"label": "浓度", "value": "浓度"},
        {"label": "微生物", "value": "微生物"},
        {"label": "比率", "value": "比率"},
        {"label": "其他", "value": "其他"},
    ],
    "chrom-column-status": [
        {"label": "在用", "value": "0"},
        {"label": "待清洗", "value": "1"},
        {"label": "封存", "value": "2"},
        {"label": "报废", "value": "3"},
    ],
    "material-type": [
        {"label": "原料", "value": "原料"},
        {"label": "辅料", "value": "辅料"},
        {"label": "包材", "value": "包材"},
        {"label": "中间体", "value": "中间体"},
    ],
    "standard-source": [
        {"label": "中国药典", "value": "中国药典"},
        {"label": "USP", "value": "USP"},
        {"label": "EP", "value": "EP"},
        {"label": "JP", "value": "JP"},
        {"label": "内控标准", "value": "内控标准"},
    ],
    "limit-type": [
        {"label": "上限", "value": "上限"},
        {"label": "下限", "value": "下限"},
        {"label": "区间", "value": "区间"},
        {"label": "不得检出", "value": "不得检出"},
        {"label": "等于", "value": "等于"},
    ],
    "test-item-category": [
        {"label": "理化检验", "value": "理化"},
        {"label": "仪器分析", "value": "仪器分析"},
        {"label": "微生物检验", "value": "微生物"},
    ],
}


@router.get("/dict/{dict_type}", summary="Get dictionary options by type")
async def get_dict_options(dict_type: str):
    """Get dictionary options - returns hardcoded options for various dict types"""
    if dict_type in DICT_OPTIONS:
        return ApiResponse(data=DICT_OPTIONS[dict_type])
    return ApiResponse(code=404, message=f"Dictionary type '{dict_type}' not found")


@router.get("/storage-condition/options", summary="Get storage condition options")
async def get_storage_condition_options(db: AsyncSession = Depends(get_db)):
    """Get storage condition options for dropdown selection"""
    from sqlalchemy import select, and_
    from app.modules.quality.static_data.models import StorageCondition
    
    result = await db.execute(
        select(StorageCondition).where(
            and_(StorageCondition.del_flag == 0, StorageCondition.status == 0)
        ).order_by(StorageCondition.id)
    )
    items = result.scalars().all()
    return ApiResponse(data=[{"label": x.cond_name, "value": x.cond_code} for x in items])


@router.get("/unit/options", summary="Get unit options")
async def get_unit_options(db: AsyncSession = Depends(get_db)):
    """Get unit options for dropdown selection"""
    from sqlalchemy import select, and_
    from app.modules.quality.static_data.models import Unit
    
    result = await db.execute(
        select(Unit).where(
            and_(Unit.del_flag == 0, Unit.status == 0)
        ).order_by(Unit.id)
    )
    items = result.scalars().all()
    return ApiResponse(data=[{"label": x.unit_name, "value": x.unit_code} for x in items])


# ========== Template Download ==========

@router.get("/hplc-reference/template", summary="Download HPLC reference template")
async def download_hplc_reference_template():
    """Download Excel template for HPLC reference substance import"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HPLC Reference Template"
    
    # Headers
    headers = [
        '对照品编号(ref_code)*', '对照品名称(ref_name)*', '检测项目(project_name)',
        '厂内批号(internal_batch)', 'CAS号(cas_no)', '供应商货号(cat_no)',
        '厂家批号(manufacturer_batch)', '供应商/来源(manufacturer)', '规格(spec)',
        '纯度(purity)', '含量(content)', '数量(quantity)', '库存状态(stock_status)',
        '到货日期(arrival_date)', '生产/标定日期(produce_date)', '有效期至(expire_date)',
        '复标周期天(recal_cycle_days)', '开瓶日期(open_date)', '开瓶有效期天(open_expire_days)',
        '贮存条件编码(storage_cond_code)', '存放位置(location)', '是否有COA(has_coa)',
        '交接单号(handover_no)', '状态(ref_status:0在用/1用完/2过期/3报废)', '备注(remark)'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True)
    
    # Sample row
    sample_data = [
        'REF0001', '样品名称', '检测项目A',
        'BATCH001', '12345-67-8', 'CAT-001',
        'MFG-BATCH', '供应商名称', '100mg',
        '99.5', '99.0', '1',
        '在用', '2026-01-01', '2026-01-01', '2027-01-01',
        '180', '2026-06-01', '30',
        'SC001', '冷藏柜A', 'TRUE',
        'HJ20260001', '0', '测试备注'
    ]
    for col, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=hplc_reference_template.xlsx"}
    )


# ========== Batch Import ==========

@router.post("/hplc-reference/batch-import", summary="Batch import HPLC reference substances")
async def batch_import_hplc_reference(
    file: UploadFile = File(...),
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    """Import HPLC reference substances from Excel file"""
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        return ApiResponse(code=400, message="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Field mapping
        field_map = {
            '对照品编号(ref_code)*': 'ref_code',
            '对照品名称(ref_name)*': 'ref_name',
            '检测项目(project_name)': 'project_name',
            '厂内批号(internal_batch)': 'internal_batch',
            'CAS号(cas_no)': 'cas_no',
            '供应商货号(cat_no)': 'cat_no',
            '厂家批号(manufacturer_batch)': 'manufacturer_batch',
            '供应商/来源(manufacturer)': 'manufacturer',
            '规格(spec)': 'spec',
            '纯度(purity)': 'purity',
            '含量(content)': 'content',
            '数量(quantity)': 'quantity',
            '库存状态(stock_status)': 'stock_status',
            '到货日期(arrival_date)': 'arrival_date',
            '生产/标定日期(produce_date)': 'produce_date',
            '有效期至(expire_date)': 'expire_date',
            '复标周期天(recal_cycle_days)': 'recal_cycle_days',
            '开瓶日期(open_date)': 'open_date',
            '开瓶有效期天(open_expire_days)': 'open_expire_days',
            '贮存条件编码(storage_cond_code)': 'storage_cond_code',
            '存放位置(location)': 'location',
            '是否有COA(has_coa)': 'has_coa',
            '交接单号(handover_no)': 'handover_no',
            '状态(ref_status:0在用/1用完/2过期/3报废)': 'ref_status',
            '备注(remark)': 'remark',
        }
        
        success_count = 0
        error_count = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or not row[1]:  # Skip empty rows
                continue
            
            data = {'create_by': user_id}
            for col, header in enumerate(headers):
                if header in field_map and row[col] is not None:
                    field = field_map[header]
                    value = row[col]
                    # Convert boolean
                    if field == 'has_coa':
                        value = str(value).upper() in ['TRUE', '是', 'YES', '1']
                    # Convert status
                    elif field == 'ref_status':
                        if isinstance(value, str):
                            if '用完' in value:
                                value = 1
                            elif '过期' in value:
                                value = 2
                            elif '报废' in value:
                                value = 3
                            else:
                                value = 0
                    # Convert date strings
                    elif field in ['arrival_date', 'produce_date', 'expire_date', 'open_date']:
                        if isinstance(value, str):
                            try:
                                value = date.fromisoformat(value.split()[0])
                            except:
                                value = None
                    data[field] = value
            
            try:
                await service.create_hplc_reference(s.HplcReferenceCreate(**data), user_id)
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append(f"Row {row_num}: {str(e)}")
        
        message = f"Import completed: {success_count} success, {error_count} failed"
        if errors:
            message += f"\nErrors: {'; '.join(errors[:5])}"
        
        return ApiResponse(message=message, data={"success": success_count, "failed": error_count})
    except Exception as e:
        return ApiResponse(code=500, message=f"Import failed: {str(e)}")


# ========== 11. HPLC Reference Substance ==========

@router.get("/hplc-reference", summary="List HPLC reference substances")
async def list_hplc_reference(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    ref_code: Optional[str] = Query(None, description="Reference code"),
    ref_name: Optional[str] = Query(None, description="Reference name"),
    project_name: Optional[str] = Query(None, description="Test project"),
    ref_status: Optional[int] = Query(None, description="Status"),
    has_coa: Optional[bool] = Query(None, description="Has COA"),
    service: StaticDataService = Depends(_get_service),
):
    skip = (page - 1) * page_size
    items, total = await service.list_hplc_reference(
        skip, page_size,
        ref_code=ref_code, ref_name=ref_name, project_name=project_name,
        ref_status=ref_status, has_coa=has_coa,
    )
    return ApiResponse(
        data=[s.HplcReferenceResponse.model_validate(x) for x in items],
        meta={"page": page, "page_size": page_size, "total": total},
    )


@router.get("/hplc-reference/need-recal", summary="查询需要复标的对照品")
async def get_hplc_references_need_recal(
    service: StaticDataService = Depends(_get_service),
):
    """查询剩余量低于复标阈值、需要复标的对照品列表"""
    items = await service.get_hplc_references_need_recal()
    return ApiResponse(
        data=[s.HplcReferenceResponse.model_validate(x) for x in items],
        meta={'count': len(items)},
    )


@router.get("/hplc-reference/{id}", summary="Get HPLC reference substance by ID")
async def get_hplc_reference(
    id: int,
    service: StaticDataService = Depends(_get_service),
):
    obj = await service.get_hplc_reference(id)
    if not obj:
        return ApiResponse(code=404, message="Record not found")
    return ApiResponse(data=s.HplcReferenceResponse.model_validate(obj))


@router.post("/hplc-reference", summary="Create HPLC reference substance")
async def create_hplc_reference(
    data: s.HplcReferenceCreate,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.create_hplc_reference(data, user_id)
        return ApiResponse(data=s.HplcReferenceResponse.model_validate(obj), message="Created successfully")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.put("/hplc-reference/{id}", summary="Update HPLC reference substance")
async def update_hplc_reference(
    id: int,
    data: s.HplcReferenceUpdate,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.update_hplc_reference(id, data, user_id)
        return ApiResponse(data=s.HplcReferenceResponse.model_validate(obj), message="Updated successfully")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.delete("/hplc-reference/{id}", summary="Delete HPLC reference substance")
async def delete_hplc_reference(
    id: int,
    service: StaticDataService = Depends(_get_service),
):
    try:
        await service.delete_hplc_reference(id)
        return ApiResponse(message="Deleted successfully")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.post("/hplc-reference/{id}/adjust-quantity", summary="Adjust HPLC reference quantity")
async def adjust_hplc_reference_quantity(
    id: int,
    quantity_change: int = Body(..., embed=True, description="Quantity change (positive = in, negative = out)"),
    service: StaticDataService = Depends(_get_service),
    current_user: dict = Depends(get_current_user),
):
    try:
        user_id = current_user.get('id', 0) if current_user else 0
        obj = await service.adjust_hplc_reference_quantity(id, quantity_change, user_id)
        return ApiResponse(data=s.HplcReferenceResponse.model_validate(obj), message="Quantity adjusted")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.post("/hplc-reference/{id}/use", summary="使用/领用对照品")
async def use_hplc_reference(
    id: int,
    usage_amount: float = Body(..., embed=True, description="领用量 (mg/g)"),
    usage_unit: str = Body('mg', embed=True, description="领用单位"),
    usage_person: Optional[str] = Body(None, embed=True, description="领用人"),
    usage_purpose: Optional[str] = Body(None, embed=True, description="领用用途/项目"),
    remark: Optional[str] = Body(None, embed=True, description="备注"),
    service: StaticDataService = Depends(_get_service),
    current_user: dict = Depends(get_current_user),
):
    """领用对照品，扣减剩余量并记录领用历史"""
    try:
        user_id = current_user.get('id', 0) if current_user else 0
        obj, usage_log = await service.use_hplc_reference(
            id, usage_amount, usage_unit,
            usage_person, usage_purpose, remark,
            user_id,
        )
        return ApiResponse(
            data={
                'reference': s.HplcReferenceResponse.model_validate(obj),
                'usage': s.HplcReferenceUsageResponse.model_validate(usage_log),
            },
            message="领用成功",
        )
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.get("/hplc-reference/{id}/usage-history", summary="查询对照品领用历史")
async def get_hplc_reference_usage_history(
    id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    service: StaticDataService = Depends(_get_service),
):
    """查询指定对照品的领用历史记录"""
    skip = (page - 1) * page_size
    items, total = await service.list_hplc_reference_usage(ref_id=id, skip=skip, limit=page_size)
    return ApiResponse(
        data=[s.HplcReferenceUsageResponse.model_validate(x) for x in items],
        meta={'page': page, 'page_size': page_size, 'total': total},
    )


# ========== 5. Chromatography Column ==========

@router.get("/chrom-column", summary="List chromatography columns")
async def list_chrom_column(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    col_code: Optional[str] = Query(None, description="Column code"),
    col_type: Optional[str] = Query(None, description="Column type"),
    manufacturer: Optional[str] = Query(None, description="Manufacturer"),
    spec: Optional[str] = Query(None, description="Specification"),
    col_status: Optional[int] = Query(None, description="Status: 0-active 1-waiting_clean 2-sealed 3-scrapped"),
    column_category: Optional[int] = Query(None, description="Category: 0-HPLC 1-GC"),
    service: StaticDataService = Depends(_get_service),
):
    skip = (page - 1) * page_size
    items, total = await service.list_chrom_column(
        skip, page_size,
        col_code=col_code, col_type=col_type,
        manufacturer=manufacturer, spec=spec,
        col_status=col_status,
        column_category=column_category,
    )
    return ApiResponse(
        data=[s.ChromColumnResponse.model_validate(x) for x in items],
        meta={"page": page, "page_size": page_size, "total": total},
    )


@router.get("/chrom-column/template", summary="Download chromatography column template")
async def download_chrom_column_template():
    """Download Excel template for chromatography column import"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "液相色谱柱"
    
    headers = [
        '品牌', '型号', '色谱柱类型*', '色谱柱内径(mm)', '柱长(mm)', '填料粒径(μm)',
        '验收日期*', '色谱柱编号*', '货号P.N', '序列号S.N*', '产品编号', '批次号L.N',
        '启用日期', '适用检测项目', '最大使用次数', '存放位置*', '贮存条件编码*',
        '状态(0在用/1待清洗/2封存/3报废)', '备注'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D6E4FF", end_color="D6E4FF", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, horizontal='center')
    
    sample_data = [
        '安捷伦', 'Agilent', 'ZORBAX SB-Aq', '4.6', '250', '5',
        '2024-05-04', 'LC24001', '880975-914', 'USAG025278', '/', 'B23485',
        '2024-05-31', '枸橼酸铋钾有关物质', '100', '液相室A柜', 'ROOM_TEMP',
        '0', '示例数据，请删除后再导入'
    ]
    for col, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
    
    ws2 = wb.create_sheet("气相色谱柱")
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, horizontal='center')
    
    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=chromatography_column_template.xlsx"}
    )


@router.post("/chrom-column/batch-import", summary="Batch import chromatography columns")
async def batch_import_chrom_column(
    file: UploadFile = File(...),
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    """Import chromatography columns from Excel file (supports both 液相 and 气相 sheets)"""
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        return ApiResponse(code=400, message="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        
        field_map = {
            '品牌': 'brand',
            '型号': 'model',
            '色谱柱信息': 'col_type',
            '色谱柱类型*': 'col_type',
            '色谱柱类型': 'col_type',
            '色谱柱内径(mm)': 'inner_diameter',
            '色谱柱内径': 'inner_diameter',
            '柱长(mm)': 'column_length',
            '柱长': 'column_length',
            '填料粒径(μm)': 'particle_size',
            '填料粒径': 'particle_size',
            '验收日期*': 'purchase_date',
            '验收日期': 'purchase_date',
            '色谱柱编号*': 'col_code',
            '色谱柱编号': 'col_code',
            '货号P.N': 'cat_no',
            '序列号S.N*': 'serial_no',
            '序列号S.N': 'serial_no',
            '产品编号': 'product_no',
            '编号C-N0': 'product_no',
            '批次号L.N': 'batch_no',
            '启用日期': 'use_start_date',
            '适用检测项目': 'apply_method',
            '品名/检验项目': 'apply_method',
            '最大使用次数': 'max_use_times',
            '存放位置*': 'location',
            '存放位置': 'location',
            '贮存条件编码*': 'storage_cond_code',
            '贮存条件编码': 'storage_cond_code',
            '状态(0在用/1待清洗/2封存/3报废)': 'col_status',
            '状态': 'col_status',
            '备注': 'remark',
        }
        
        success_count = 0
        error_count = 0
        errors = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue
            
            headers = [cell.value for cell in ws[1]]
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                
                data = {'create_by': user_id}
                
                if '气相' in sheet_name or 'GC' in sheet_name.upper():
                    data['column_category'] = 1
                else:
                    data['column_category'] = 0
                
                spec_parts = []
                manufacturer_parts = []
                
                for col, header in enumerate(headers):
                    if not header or row[col] is None:
                        continue
                    header_str = str(header).strip()
                    if header_str not in field_map:
                        continue
                    
                    field = field_map[header_str]
                    value = row[col]
                    
                    if field == 'brand':
                        manufacturer_parts.append(str(value))
                    elif field == 'model':
                        manufacturer_parts.append(str(value))
                    elif field == 'inner_diameter':
                        spec_parts.append(f"{value}")
                    elif field == 'column_length':
                        spec_parts.append(f"*{value}mm")
                    elif field == 'particle_size':
                        spec_parts.append(f" {value}μm")
                    elif field == 'col_type':
                        data['col_type'] = str(value).strip()
                    elif field == 'col_code':
                        data['col_code'] = str(value).strip()
                    elif field == 'serial_no':
                        data['serial_no'] = str(value).strip()
                    elif field == 'location':
                        data['location'] = str(value).strip()
                    elif field == 'storage_cond_code':
                        data['storage_cond_code'] = str(value).strip()
                    elif field == 'apply_method':
                        data['apply_method'] = str(value).strip()
                    elif field == 'purchase_date':
                        if isinstance(value, date):
                            data['purchase_date'] = value
                        elif isinstance(value, (int, float)):
                            try:
                                from datetime import timedelta
                                base = date(1899, 12, 30)
                                data['purchase_date'] = base + timedelta(days=int(value))
                            except:
                                pass
                        elif isinstance(value, str) and value.strip() and value.strip() != '/':
                            try:
                                v = value.strip()
                                if len(v) == 8 and v.isdigit():
                                    data['purchase_date'] = date(int(v[:4]), int(v[4:6]), int(v[6:8]))
                                else:
                                    data['purchase_date'] = date.fromisoformat(v.replace('/', '-').split()[0])
                            except:
                                pass
                    elif field == 'use_start_date':
                        if isinstance(value, date):
                            data['use_start_date'] = value
                        elif isinstance(value, (int, float)):
                            try:
                                from datetime import timedelta
                                base = date(1899, 12, 30)
                                data['use_start_date'] = base + timedelta(days=int(value))
                            except:
                                pass
                        elif isinstance(value, str) and value.strip() and value.strip() != '/':
                            try:
                                v = value.strip()
                                if len(v) == 8 and v.isdigit():
                                    data['use_start_date'] = date(int(v[:4]), int(v[4:6]), int(v[6:8]))
                                else:
                                    data['use_start_date'] = date.fromisoformat(v.replace('/', '-').split()[0])
                            except:
                                pass
                    elif field == 'max_use_times':
                        try:
                            data['max_use_times'] = int(float(value))
                        except:
                            data['max_use_times'] = 100
                    elif field == 'col_status':
                        v = str(value).strip()
                        if '待清洗' in v:
                            data['col_status'] = 1
                        elif '封存' in v:
                            data['col_status'] = 2
                        elif '报废' in v:
                            data['col_status'] = 3
                        else:
                            data['col_status'] = 0
                    elif field == 'remark':
                        data['remark'] = str(value).strip()
                
                if manufacturer_parts:
                    data['manufacturer'] = ' '.join(manufacturer_parts)
                if spec_parts:
                    data['spec'] = ''.join(spec_parts).replace('*', '×')
                
                if 'col_code' not in data or not data['col_code'] or data['col_code'] == '/':
                    error_count += 1
                    errors.append(f"Sheet[{sheet_name}] Row {row_num}: 缺少色谱柱编号")
                    continue
                if 'col_type' not in data or not data['col_type']:
                    error_count += 1
                    errors.append(f"Sheet[{sheet_name}] Row {row_num} ({data.get('col_code','')}): 缺少色谱柱类型")
                    continue
                if 'manufacturer' not in data or not data['manufacturer']:
                    data['manufacturer'] = '未知'
                if 'serial_no' not in data or not data['serial_no'] or data['serial_no'] == '/':
                    data['serial_no'] = data['col_code']
                if 'location' not in data or not data['location']:
                    data['location'] = '未指定'
                if 'storage_cond_code' not in data or not data['storage_cond_code']:
                    data['storage_cond_code'] = 'ROOM_TEMP'
                if 'max_use_times' not in data:
                    data['max_use_times'] = 100
                if 'col_status' not in data:
                    data['col_status'] = 0
                if 'purchase_date' not in data:
                    if 'use_start_date' in data:
                        data['purchase_date'] = data['use_start_date']
                    else:
                        data['purchase_date'] = date.today()
                if 'spec' not in data or not data['spec']:
                    data['spec'] = '未指定'
                
                try:
                    await service.create_chrom_column(s.ChromColumnCreate(**data), user_id)
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f"Sheet[{sheet_name}] Row {row_num} ({data.get('col_code','')}): {str(e)}")
        
        message = f"导入完成: 成功 {success_count} 条，失败 {error_count} 条"
        if errors:
            message += f"\n错误: {'; '.join(errors[:10])}"
            if len(errors) > 10:
                message += f" ...等共{len(errors)}条错误"
        
        return ApiResponse(message=message, data={"success": success_count, "failed": error_count, "errors": errors})
    except Exception as e:
        return ApiResponse(code=500, message=f"导入失败: {str(e)}")


@router.get("/chrom-column/{id}", summary="Get chromatography column by ID")
async def get_chrom_column(
    id: int,
    service: StaticDataService = Depends(_get_service),
):
    obj = await service.get_chrom_column(id)
    if not obj:
        return ApiResponse(code=404, message="Record not found")
    return ApiResponse(data=s.ChromColumnResponse.model_validate(obj))


@router.post("/chrom-column", summary="Create chromatography column")
async def create_chrom_column(
    data: s.ChromColumnCreate,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.create_chrom_column(data, user_id)
        return ApiResponse(data=s.ChromColumnResponse.model_validate(obj), message="Created successfully")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.put("/chrom-column/{id}", summary="Update chromatography column")
async def update_chrom_column(
    id: int,
    data: s.ChromColumnUpdate,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.update_chrom_column(id, data, user_id)
        return ApiResponse(data=s.ChromColumnResponse.model_validate(obj), message="Updated successfully")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.delete("/chrom-column/{id}", summary="Delete chromatography column")
async def delete_chrom_column(
    id: int,
    service: StaticDataService = Depends(_get_service),
):
    try:
        await service.delete_chrom_column(id)
        return ApiResponse(message="Deleted successfully")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.post("/chrom-column/{id}/increment-usage", summary="Increment column usage count")
async def increment_chrom_column_usage(
    id: int,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.increment_chrom_column_usage(id, user_id)
        return ApiResponse(data=s.ChromColumnResponse.model_validate(obj), message="Usage incremented")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


# ========== 6. Medium (培养基) ==========

@router.get("/medium", summary="List medium")
async def list_medium(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    medium_code: Optional[str] = Query(None, description="Medium code"),
    medium_name: Optional[str] = Query(None, description="Medium name"),
    medium_type: Optional[str] = Query(None, description="Medium type"),
    manufacturer: Optional[str] = Query(None, description="Manufacturer"),
    verify_status: Optional[str] = Query(None, description="Verify status"),
    status: Optional[int] = Query(None, description="Status: 0-active 1-inactive"),
    service: StaticDataService = Depends(_get_service),
):
    skip = (page - 1) * page_size
    items, total = await service.list_medium(
        skip, page_size,
        medium_code=medium_code, medium_name=medium_name,
        medium_type=medium_type, manufacturer=manufacturer,
        verify_status=verify_status, status=status,
    )
    return ApiResponse(
        data=[s.MediumResponse.model_validate(x) for x in items],
        meta={"page": page, "page_size": page_size, "total": total},
    )


@router.get("/medium/{id}", summary="Get medium by ID")
async def get_medium(
    id: int,
    service: StaticDataService = Depends(_get_service),
):
    obj = await service.get_medium(id)
    if not obj:
        return ApiResponse(code=404, message="Medium not found")
    return ApiResponse(data=s.MediumResponse.model_validate(obj))


@router.post("/medium", summary="Create medium")
async def create_medium(
    data: s.MediumCreate,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.create_medium(data, user_id)
        return ApiResponse(data=s.MediumResponse.model_validate(obj), message="Medium created")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.put("/medium/{id}", summary="Update medium")
async def update_medium(
    id: int,
    data: s.MediumUpdate,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.update_medium(id, data, user_id)
        return ApiResponse(data=s.MediumResponse.model_validate(obj), message="Medium updated")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.delete("/medium/{id}", summary="Delete medium")
async def delete_medium(
    id: int,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        await service.delete_medium(id)
        return ApiResponse(message="Medium deleted")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.post("/medium/{id}/adjust-stock", summary="Adjust medium stock quantity")
async def adjust_medium_stock(
    id: int,
    quantity: int = Body(..., embed=True, description="Quantity change (positive or negative)"),
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.adjust_medium_stock(id, quantity, user_id)
        return ApiResponse(data=s.MediumResponse.model_validate(obj), message="Stock adjusted")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


# ========== 7. Standard (标准品) ==========

@router.get("/standard", summary="List standards")
async def list_standard(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    std_code: Optional[str] = Query(None, description="Standard code"),
    std_name: Optional[str] = Query(None, description="Standard name"),
    std_type: Optional[str] = Query(None, description="Type: national/working/international"),
    manufacturer: Optional[str] = Query(None, description="Manufacturer"),
    std_status: Optional[int] = Query(None, description="Status: 0-active 1-used_up 2-expired 3-scrapped"),
    service: StaticDataService = Depends(_get_service),
):
    skip = (page - 1) * page_size
    items, total = await service.list_standard(
        skip, page_size,
        std_code=std_code, std_name=std_name,
        std_type=std_type, manufacturer=manufacturer,
        std_status=std_status,
    )
    return ApiResponse(
        data=[s.StandardResponse.model_validate(x) for x in items],
        meta={"page": page, "page_size": page_size, "total": total},
    )


@router.get("/standard/{id}", summary="Get standard by ID")
async def get_standard(
    id: int,
    service: StaticDataService = Depends(_get_service),
):
    obj = await service.get_standard(id)
    if not obj:
        return ApiResponse(code=404, message="Standard not found")
    return ApiResponse(data=s.StandardResponse.model_validate(obj))


@router.post("/standard", summary="Create standard")
async def create_standard(
    data: s.StandardCreate,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.create_standard(data, user_id)
        return ApiResponse(data=s.StandardResponse.model_validate(obj), message="Standard created")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.put("/standard/{id}", summary="Update standard")
async def update_standard(
    id: int,
    data: s.StandardUpdate,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.update_standard(id, data, user_id)
        return ApiResponse(data=s.StandardResponse.model_validate(obj), message="Standard updated")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.delete("/standard/{id}", summary="Delete standard")
async def delete_standard(
    id: int,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        await service.delete_standard(id)
        return ApiResponse(message="Standard deleted")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.post("/standard/{id}/adjust-quantity", summary="Adjust standard quantity")
async def adjust_standard_quantity(
    id: int,
    quantity: int = Body(..., embed=True, description="Quantity change (positive or negative)"),
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.adjust_standard_quantity(id, quantity, user_id)
        return ApiResponse(data=s.StandardResponse.model_validate(obj), message="Quantity adjusted")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


# ========== 8. Storage Condition (贮存条件) ==========

@router.get("/storage-condition", summary="List storage conditions")
async def list_storage_condition(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    cond_code: Optional[str] = Query(None, description="Condition code"),
    cond_name: Optional[str] = Query(None, description="Condition name"),
    status: Optional[int] = Query(None, description="Status: 0-enabled 1-disabled"),
    service: StaticDataService = Depends(_get_service),
):
    skip = (page - 1) * page_size
    items, total = await service.list_storage_condition(
        skip, page_size,
        cond_code=cond_code, cond_name=cond_name, status=status,
    )
    return ApiResponse(
        data=[s.StorageConditionResponse.model_validate(x) for x in items],
        meta={"page": page, "page_size": page_size, "total": total},
    )


@router.get("/storage-condition/{id}", summary="Get storage condition by ID")
async def get_storage_condition(
    id: int,
    service: StaticDataService = Depends(_get_service),
):
    obj = await service.get_storage_condition(id)
    if not obj:
        return ApiResponse(code=404, message="Storage condition not found")
    return ApiResponse(data=s.StorageConditionResponse.model_validate(obj))


@router.post("/storage-condition", summary="Create storage condition")
async def create_storage_condition(
    data: s.StorageConditionCreate,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.create_storage_condition(data, user_id)
        return ApiResponse(data=s.StorageConditionResponse.model_validate(obj), message="Storage condition created")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.put("/storage-condition/{id}", summary="Update storage condition")
async def update_storage_condition(
    id: int,
    data: s.StorageConditionUpdate,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        obj = await service.update_storage_condition(id, data, user_id)
        return ApiResponse(data=s.StorageConditionResponse.model_validate(obj), message="Storage condition updated")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))


@router.delete("/storage-condition/{id}", summary="Delete storage condition")
async def delete_storage_condition(
    id: int,
    service: StaticDataService = Depends(_get_service),
    user_id: int = Depends(_user_id),
):
    try:
        await service.delete_storage_condition(id)
        return ApiResponse(message="Storage condition deleted")
    except ValueError as e:
        return ApiResponse(code=400, message=str(e))